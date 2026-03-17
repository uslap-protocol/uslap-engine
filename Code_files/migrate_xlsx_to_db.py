#!/usr/bin/env python3
"""
Migrate ALL sheets from USLaP_Final_Data_Consolidated_Master_v3.xlsx
into uslap_database_v3.db as indexed SQLite tables.

Existing tables are preserved. New tables are created for sheets
not yet in the DB. If a sheet's table already exists, it is REPLACED
(dropped and recreated) to ensure full sync.
"""

import sqlite3
import openpyxl
import os
import re
import shutil
from datetime import datetime

MASTER_XLSX = "/Users/mmsetubal/Documents/USLaP workplace/USLaP_Final_Data_Consolidated_Master_v3.xlsx"
DB_PATH = "/Users/mmsetubal/Documents/USLaP workplace/Code_files/uslap_database_v3.db"

# Tables that already exist and should NOT be touched by this migration
# (they have their own schema and data)
PRESERVE_TABLES = {
    'umd_operations', 'child_schema', 'dp_register', 'att_terms',
    'phonetic_reversal', 'session_index', 'protocol_corrections',
    'scholar_warnings', 'cross_reference', 'sqlite_sequence'
}

def sanitize_table_name(sheet_name):
    """Convert sheet name to valid SQLite table name."""
    name = sheet_name.strip()
    # Replace spaces, hyphens, special chars with underscore
    name = re.sub(r'[^a-zA-Z0-9_\u0400-\u04FF\u0600-\u06FF]', '_', name)
    # Remove leading/trailing underscores
    name = name.strip('_')
    # Ensure it doesn't start with a number
    if name and name[0].isdigit():
        name = 'sheet_' + name
    return name.lower()

def sanitize_column_name(col_name):
    """Convert column header to valid SQLite column name."""
    if col_name is None:
        return None
    name = str(col_name).strip()
    name = re.sub(r'[^a-zA-Z0-9_\u0400-\u04FF\u0600-\u06FF]', '_', name)
    name = name.strip('_')
    if not name:
        return None
    if name[0].isdigit():
        name = 'col_' + name
    return name.lower()

def get_column_type(values):
    """Infer SQLite column type from sample values."""
    has_int = False
    has_float = False
    has_text = False
    for v in values:
        if v is None:
            continue
        if isinstance(v, bool):
            has_text = True
        elif isinstance(v, int):
            has_int = True
        elif isinstance(v, float):
            has_float = True
        else:
            has_text = True
    if has_text:
        return 'TEXT'
    if has_float:
        return 'REAL'
    if has_int:
        return 'INTEGER'
    return 'TEXT'

def migrate():
    # Backup DB first
    backup_dir = os.path.dirname(DB_PATH)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(backup_dir, f'backups/uslap_db_v3_backup_{timestamp}.db')
    os.makedirs(os.path.dirname(backup_path), exist_ok=True)
    shutil.copy2(DB_PATH, backup_path)
    print(f"Backup created: {backup_path}")

    # Load workbook (read_only for speed, data_only to get values not formulas)
    print(f"Loading {MASTER_XLSX}...")
    wb = openpyxl.load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    print(f"Sheets found: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"  - {s}")

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    migrated = []
    skipped = []
    errors = []

    for sheet_name in wb.sheetnames:
        table_name = sanitize_table_name(sheet_name)

        # Check if this would collide with a preserved table
        if table_name in PRESERVE_TABLES:
            # Use a prefixed name to avoid collision
            table_name = 'xlsx_' + table_name
            print(f"  Sheet '{sheet_name}' -> table '{table_name}' (prefixed to avoid collision)")

        print(f"\nProcessing: '{sheet_name}' -> '{table_name}'")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            print(f"  SKIP: empty sheet")
            skipped.append(sheet_name)
            continue

        # First row = headers
        raw_headers = rows[0]
        headers = []
        seen = {}
        for i, h in enumerate(raw_headers):
            col_name = sanitize_column_name(h)
            if col_name is None:
                col_name = f'col_{i}'
            # Handle duplicates
            if col_name in seen:
                seen[col_name] += 1
                col_name = f'{col_name}_{seen[col_name]}'
            else:
                seen[col_name] = 0
            headers.append(col_name)

        data_rows = rows[1:]

        if not data_rows:
            print(f"  SKIP: headers only, no data")
            skipped.append(sheet_name)
            continue

        # Filter out completely empty rows
        data_rows = [r for r in data_rows if any(v is not None for v in r)]

        if not data_rows:
            print(f"  SKIP: all rows empty")
            skipped.append(sheet_name)
            continue

        # Ensure all rows have same column count as headers
        n_cols = len(headers)
        clean_rows = []
        for r in data_rows:
            row = list(r)
            if len(row) < n_cols:
                row.extend([None] * (n_cols - len(row)))
            elif len(row) > n_cols:
                row = row[:n_cols]
            clean_rows.append(tuple(row))

        # Infer column types
        col_types = []
        for i in range(n_cols):
            sample = [r[i] for r in clean_rows[:100]]
            col_types.append(get_column_type(sample))

        # Build CREATE TABLE
        col_defs = []
        for h, t in zip(headers, col_types):
            col_defs.append(f'"{h}" {t}')

        try:
            # Drop if exists (full sync)
            cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            create_sql = f'CREATE TABLE "{table_name}" ({", ".join(col_defs)})'
            cursor.execute(create_sql)

            # Insert data
            placeholders = ', '.join(['?' for _ in headers])
            insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'

            # Convert all values to strings for TEXT columns to avoid type issues
            final_rows = []
            for r in clean_rows:
                row = []
                for i, v in enumerate(r):
                    if v is not None and col_types[i] == 'TEXT':
                        row.append(str(v))
                    else:
                        row.append(v)
                final_rows.append(tuple(row))

            cursor.executemany(insert_sql, final_rows)

            # Create indexes on common key columns
            key_columns = ['entry_id', 'root_id', 'en_term', 'score', 'network_id',
                          'shift_id', 'dp_id', 'allah_name_id', 'root_letters',
                          'запись_id', 'ru_term', 'корень_id']
            for kc in key_columns:
                if kc in headers:
                    idx_name = f'idx_{table_name}_{kc}'
                    try:
                        cursor.execute(f'CREATE INDEX IF NOT EXISTS "{idx_name}" ON "{table_name}"("{kc}")')
                    except:
                        pass  # Skip if index creation fails

            conn.commit()
            print(f"  OK: {len(final_rows)} rows, {n_cols} columns")
            migrated.append((sheet_name, table_name, len(final_rows), n_cols))

        except Exception as e:
            errors.append((sheet_name, str(e)))
            print(f"  ERROR: {e}")
            conn.rollback()

    wb.close()

    # Print summary
    print("\n" + "="*60)
    print("MIGRATION SUMMARY")
    print("="*60)
    print(f"\nMigrated: {len(migrated)} sheets")
    for sheet, table, rows, cols in migrated:
        print(f"  {sheet:40s} -> {table:35s} ({rows} rows, {cols} cols)")
    print(f"\nSkipped: {len(skipped)} sheets")
    for s in skipped:
        print(f"  {s}")
    print(f"\nErrors: {len(errors)} sheets")
    for s, e in errors:
        print(f"  {s}: {e}")

    # Show final table list
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
    all_tables = cursor.fetchall()
    print(f"\nTotal tables in DB: {len(all_tables)}")
    for t in all_tables:
        cursor.execute(f'SELECT COUNT(*) FROM "{t[0]}"')
        count = cursor.fetchone()[0]
        print(f"  {t[0]:40s} {count:>6d} rows")

    conn.close()
    print(f"\nDone. DB: {DB_PATH}")

if __name__ == '__main__':
    migrate()
