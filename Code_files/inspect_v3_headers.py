#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook("USLaP_Final_Data_Consolidated_Master_v3.xlsx", read_only=True)

for sheet_name in ["PROTOCOL_CORRECTIONS", "SCHOLAR_WARNINGS"]:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        print(f"\n=== {sheet_name} ===")
        for i, row in enumerate(rows[:10]):
            print(f"Row {i}: {row}")
        
        # Try to guess header row
        header_candidate = None
        for i, row in enumerate(rows):
            if row and any(isinstance(cell, str) and ('ID' in cell.upper() or 'TYPE' in cell.upper() or 'NAME' in cell.upper()) for cell in row):
                print(f"\nPotential header row {i}: {row}")
                header_candidate = (i, row)
                break
        if header_candidate:
            print(f"\nHeader row index: {header_candidate[0]}")
            print(f"Header row: {header_candidate[1]}")
        else:
            print("No clear header row found")

wb.close()