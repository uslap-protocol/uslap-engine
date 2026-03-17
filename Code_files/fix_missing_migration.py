#!/usr/bin/env python3
"""
Fix missing data in USLaP database migration:
1. Populate cross_refs from A5_CROSS_REFS sheet
2. Update ru_term in existing entries from A1_ЗАПИСИ sheet
3. Update fa_term in existing entries from PERSIAN_A1_MADĀKHIL sheet
4. Insert new entries for BITIG_A1_ENTRIES with ORIG2 prefix
"""

import sqlite3
import openpyxl
import re
import sys
import os
from datetime import datetime

EXCEL_PATH = "USLaP_Final_Data_Consolidated_Master_v3.xlsx"
DB_PATH = "Code_files/uslap_lattice.db"

def clean_column_name(col):
    """Convert Excel column header to valid SQLite column name."""
    if col is None:
        return "unknown"
    col = str(col).strip()
    col = re.sub(r'[^\w\s]', '', col)  # Remove punctuation
    col = re.sub(r'\s+', '_', col)     # Replace spaces with underscore
    col = col.lower()
    if not col:
        return "unknown"
    return col

def fix_cross_refs(conn, cursor, wb):
    """Fix cross_refs migration - insert missing cross-references."""
    print("  Fixing cross_refs migration...")
    
    if "A5_CROSS_REFS" not in wb.sheetnames:
        print("    Sheet A5_CROSS_REFS not found")
        return 0
    
    ws = wb["A5_CROSS_REFS"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A5_CROSS_REFS")
        return 0
    
    # First row is headers
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map columns according to Excel structure
        xref_id = row_dict.get('xref_id')
        from_id = row_dict.get('from_id')
        to_id = row_dict.get('to_id')
        link_type = row_dict.get('link_type')
        description = row_dict.get('description')
        layer_ref = row_dict.get('layer_ref')
        
        # Skip if missing critical data
        if not xref_id or from_id is None or to_id is None:
            continue
        
        # Convert IDs to integers if needed
        try:
            from_id = int(from_id)
            to_id = int(to_id)
        except (ValueError, TypeError):
            print(f"    Row {i}: Invalid ID format - from_id={from_id}, to_id={to_id}")
            continue
        
        # Check if entry IDs exist
        cursor.execute('SELECT 1 FROM entries WHERE entry_id = ?', (from_id,))
        if not cursor.fetchone():
            print(f"    Row {i}: from_entry_id {from_id} not found in entries")
            continue
        
        cursor.execute('SELECT 1 FROM entries WHERE entry_id = ?', (to_id,))
        if not cursor.fetchone():
            print(f"    Row {i}: to_entry_id {to_id} not found in entries")
            continue
        
        # Check if cross-ref already exists
        cursor.execute('''
            SELECT 1 FROM cross_refs 
            WHERE from_entry_id = ? AND to_entry_id = ? AND link_type = ?
        ''', (from_id, to_id, link_type))
        
        if cursor.fetchone():
            # Already exists, skip
            continue
        
        try:
            cursor.execute('''
                INSERT INTO cross_refs (
                    xref_id, from_entry_id, to_entry_id, link_type, description, layer_ref
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                xref_id, from_id, to_id, link_type, description, layer_ref
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: xref_id={xref_id}, from_id={from_id}, to_id={to_id}, link_type={link_type}")
    
    print(f"    Inserted {count} cross-references")
    return count

def fix_russian_entries(conn, cursor, wb):
    """Fix Russian entries - update ru_term in existing entries."""
    print("  Fixing Russian entries (A1_ЗАПИСИ)...")
    
    if "A1_ЗАПИСИ" not in wb.sheetnames:
        print("    Sheet A1_ЗАПИСИ not found")
        return 0
    
    ws = wb["A1_ЗАПИСИ"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A1_ЗАПИСИ")
        return 0
    
    # First row is headers (Russian column names)
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Get entry data
        entry_id = row_dict.get('запись_id')
        ru_term = row_dict.get('рус_термин')
        
        if not entry_id or not ru_term:
            continue
        
        # Convert entry_id to int if needed
        try:
            entry_id = int(entry_id)
        except (ValueError, TypeError):
            print(f"    Row {i}: Invalid entry_id format: {entry_id}")
            continue
        
        # Check if entry exists
        cursor.execute('SELECT 1 FROM entries WHERE entry_id = ?', (entry_id,))
        if not cursor.fetchone():
            print(f"    Row {i}: entry_id {entry_id} not found in entries")
            continue
        
        # Update ru_term
        try:
            cursor.execute('''
                UPDATE entries 
                SET ru_term = ?, modified_at = CURRENT_TIMESTAMP
                WHERE entry_id = ?
            ''', (ru_term, entry_id))
            
            if cursor.rowcount > 0:
                count += 1
        except Exception as e:
            print(f"    Error updating row {i}: {e}")
            print(f"    Data: entry_id={entry_id}, ru_term={ru_term}")
    
    print(f"    Updated {count} Russian entries")
    return count

def fix_persian_entries(conn, cursor, wb):
    """Fix Persian entries - update fa_term in existing entries."""
    print("  Fixing Persian entries (PERSIAN_A1_MADĀKHIL)...")
    
    if "PERSIAN_A1_MADĀKHIL" not in wb.sheetnames:
        print("    Sheet PERSIAN_A1_MADĀKHIL not found")
        return 0
    
    ws = wb["PERSIAN_A1_MADĀKHIL"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in PERSIAN_A1_MADĀKHIL")
        return 0
    
    # First row has complex headers with Persian and English
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map to entries table schema
        entry_id = row_dict.get('madkhal_identry_id')
        fa_term = row_dict.get('vazhe_farsipersian_term')
        source_word = row_dict.get('kalame_aslisource_word')
        
        if not entry_id:
            continue
        
        # Convert entry_id to int if needed
        try:
            entry_id = int(entry_id)
        except (ValueError, TypeError):
            print(f"    Row {i}: Invalid entry_id format: {entry_id}")
            continue
        
        # Check if entry exists
        cursor.execute('SELECT 1 FROM entries WHERE entry_id = ?', (entry_id,))
        if not cursor.fetchone():
            print(f"    Row {i}: entry_id {entry_id} not found in entries")
            continue
        
        # Update fa_term and ar_word if source_word is provided
        try:
            if fa_term:
                cursor.execute('''
                    UPDATE entries 
                    SET fa_term = ?, modified_at = CURRENT_TIMESTAMP
                    WHERE entry_id = ?
                ''', (fa_term, entry_id))
                
                if cursor.rowcount > 0:
                    count += 1
                    
            # Also update ar_word if source_word is provided
            if source_word:
                cursor.execute('''
                    UPDATE entries 
                    SET ar_word = ?, modified_at = CURRENT_TIMESTAMP
                    WHERE entry_id = ? AND (ar_word IS NULL OR ar_word = '')
                ''', (source_word, entry_id))
                
        except Exception as e:
            print(f"    Error updating row {i}: {e}")
            print(f"    Data: entry_id={entry_id}, fa_term={fa_term}")
    
    print(f"    Updated {count} Persian entries")
    return count

def fix_bitig_entries(conn, cursor, wb):
    """Fix Bitig entries - insert as new rows with ORIG2 prefix."""
    print("  Fixing Bitig entries (BITIG_A1_ENTRIES)...")
    
    if "BITIG_A1_ENTRIES" not in wb.sheetnames:
        print("    Sheet BITIG_A1_ENTRIES not found")
        return 0
    
    ws = wb["BITIG_A1_ENTRIES"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in BITIG_A1_ENTRIES")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Get the original entry_id from Excel
        excel_entry_id = row_dict.get('entry_id')
        
        # Check if this entry_id already exists
        cursor.execute('SELECT 1 FROM entries WHERE entry_id = ?', (excel_entry_id,))
        if cursor.fetchone():
            # Entry already exists, skip
            print(f"    Row {i}: entry_id {excel_entry_id} already exists, skipping")
            continue
        
        # Get next available entry_id
        cursor.execute('SELECT MAX(entry_id) FROM entries')
        max_id = cursor.fetchone()[0] or 0
        new_entry_id = max_id + 1
        
        # Prepare entry data
        entry_data = {
            'entry_id': new_entry_id,
            'score': row_dict.get('score', 5),
            'en_term': row_dict.get('orig2_term'),
            'ar_word': row_dict.get('orig2_script'),
            'root_letters': row_dict.get('root_letters'),
            'phonetic_chain': row_dict.get('phonetic_chain'),
            'notes': row_dict.get('notes'),
        }
        
        # Extract root ID if present
        root_id = None
        root_letters = entry_data['root_letters']
        if root_letters:
            # Try to find existing root with these letters
            cursor.execute('SELECT root_id FROM roots WHERE root_letters = ?', (root_letters,))
            result = cursor.fetchone()
            if result:
                root_id = result[0]
            else:
                # Create new root if needed
                cursor.execute('SELECT COUNT(*) FROM roots')
                root_count = cursor.fetchone()[0]
                root_id = f"R{root_count + 1:03d}"
                
                # Create bare root (without hyphens)
                root_bare = re.sub(r'[\-\s]', '', root_letters) if root_letters else ''
                root_type = 'TRILITERAL' if len(root_bare) == 3 else 'QUADRILITERAL' if len(root_bare) == 4 else 'OTHER'
                
                cursor.execute('''
                    INSERT INTO roots (root_id, root_letters, root_bare, root_type, notes)
                    VALUES (?, ?, ?, ?, ?)
                ''', (
                    root_id, root_letters, root_bare, root_type,
                    f"Created during Bitig entries fix for entry {new_entry_id}"
                ))
        
        # For ORIG2 entries, mark with ORIG2 prefix in foundation_refs
        kashgari_attestation = row_dict.get('kashgari_attestation', '')
        foundation_refs = f"ORIG2: {kashgari_attestation}"
        
        try:
            cursor.execute('''
                INSERT INTO entries (
                    entry_id, score, en_term, ar_word, root_id, root_letters,
                    phonetic_chain, foundation_refs, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                entry_data['entry_id'], entry_data['score'], entry_data['en_term'],
                entry_data['ar_word'], root_id, entry_data['root_letters'],
                entry_data['phonetic_chain'], foundation_refs, entry_data['notes']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting Bitig row {i}: {e}")
            print(f"    Data: entry_id={new_entry_id}, en_term={entry_data['en_term']}")
    
    print(f"    Inserted {count} Bitig entries")
    return count

def main():
    print("\n" + "═" * 70)
    print("  USLaP Migration Fix: Missing Data")
    print("  بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ")
    print("═" * 70)
    
    # Check if files exist
    if not os.path.exists(EXCEL_PATH):
        print(f"\n❌ ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)
    
    if not os.path.exists(DB_PATH):
        print(f"\n❌ ERROR: Database file not found: {DB_PATH}")
        sys.exit(1)
    
    # Load Excel workbook
    print(f"\n📖 Loading Excel file: {EXCEL_PATH}")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        print(f"\n❌ ERROR: Failed to load Excel file: {e}")
        sys.exit(1)
    
    # Connect to database
    print(f"\n🗄️  Connecting to database: {DB_PATH}")
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Enable foreign keys
        cursor.execute("PRAGMA foreign_keys = ON")
        
        # Begin transaction
        conn.execute("BEGIN TRANSACTION")
        
        print("\n🔧 Fixing migration issues...")
        
        # Fix cross_refs
        cross_refs_count = fix_cross_refs(conn, cursor, wb)
        
        # Fix Russian entries
        russian_count = fix_russian_entries(conn, cursor, wb)
        
        # Fix Persian entries
        persian_count = fix_persian_entries(conn, cursor, wb)
        
        # Fix Bitig entries
        bitig_count = fix_bitig_entries(conn, cursor, wb)
        
        # Commit transaction
        conn.commit()
        
        # Generate statistics
        print("\n📈 Fix Statistics:")
        print("─" * 40)
        
        cursor.execute("SELECT COUNT(*) FROM entries")
        entries_count = cursor.fetchone()[0]
        print(f"  Total entries: {entries_count}")
        
        cursor.execute("SELECT COUNT(*) FROM cross_refs")
        crossrefs_count = cursor.fetchone()[0]
        print(f"  Total cross-references: {crossrefs_count}")
        
        cursor.execute("SELECT COUNT(*) FROM entries WHERE ru_term IS NOT NULL AND ru_term != ''")
        ru_count = cursor.fetchone()[0]
        print(f"  Entries with ru_term: {ru_count}")
        
        cursor.execute("SELECT COUNT(*) FROM entries WHERE fa_term IS NOT NULL AND fa_term != ''")
        fa_count = cursor.fetchone()[0]
        print(f"  Entries with fa_term: {fa_count}")
        
        cursor.execute("SELECT COUNT(*) FROM entries WHERE foundation_refs LIKE 'ORIG2:%'")
        orig2_count = cursor.fetchone()[0]
        print(f"  ORIG2 entries: {orig2_count}")
        
        # Verify foreign keys
        print("\n🔍 Verifying foreign key constraints...")
        cursor.execute("PRAGMA foreign_key_check")
        fk_errors = cursor.fetchall()
        
        if fk_errors:
            print("❌ Foreign key errors found:")
            for error in fk_errors:
                print(f"  Table: {error[0]}, Row: {error[1]}, Referenced: {error[2]}, FK Index: {error[3]}")
        else:
            print("✓ All foreign key constraints satisfied")
        
        # Close workbook and connection
        wb.close()
        conn.close()
        
        print("\n" + "═" * 70)
        print("✅ MIGRATION FIX COMPLETED SUCCESSFULLY")
        print("═" * 70)
        
    except Exception as e:
        print(f"\n❌ ERROR during migration fix: {e}")
        import traceback
        traceback.print_exc()
        
        # Rollback if connection is still open
        try:
            conn.rollback()
            conn.close()
        except:
            pass
        
        print("\n⚠️  Migration fix failed. Database has been rolled back.")
        sys.exit(1)

if __name__ == "__main__":
    main()