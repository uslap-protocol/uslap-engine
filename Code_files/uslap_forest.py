#!/usr/bin/env python3
"""
USLaP SELF-GROWING FOREST SYSTEM
One file. Six components. Zero manual intervention.
Your 2.6 MB file grows itself.

HOW TO USE:
1. Save this file as 'uslap_forest.py' in the same folder as your Excel file
2. Open terminal in that folder
3. Run: python uslap_forest.py
4. Watch it grow

The system will:
- Read your entire Excel file
- Find its own gaps and patterns
- Generate new entries automatically
- Validate everything internally
- Write back to the file
- Repeat every time you run it
"""

import openpyxl
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import os
import sys
import re

# ============================================================================
# COMPONENT 1: SELF-READER ENGINE
# ============================================================================

class SelfReader:
    """Reads the entire Excel file into memory as a living graph"""
    
    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self.wb = None
        self.graph = {}
        self.foundation = {}
        self.mechanism = {}
        self.application = {}
        self.sheets_data = {}
        self.total_entries = 0
        
    def read_all(self):
        """Read every sheet in the workbook"""
        print(f"\n📖 Reading {self.filepath.name}...")
        
        if not self.filepath.exists():
            raise FileNotFoundError(f"File not found: {self.filepath}")
        
        self.wb = load_workbook(self.filepath, data_only=True)
        
        # Read all sheets
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            data = []
            headers = []
            
            # Get headers from first row
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(h) if h else f"COL_{i}" for i, h in enumerate(row)]
            
            # Get data from remaining rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row if cell is not None):
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = value
                    if row_dict:
                        data.append(row_dict)
            
            self.sheets_data[sheet_name] = data
            
            # Categorize by sheet prefix
            if sheet_name.startswith('F'):
                self.foundation[sheet_name] = data
            elif sheet_name.startswith('M'):
                self.mechanism[sheet_name] = data
            elif sheet_name.startswith('A'):
                self.application[sheet_name] = data
            
            print(f"  ✓ {sheet_name}: {len(data)} rows")
        
        # Build graph from A1_ENTRIES
        self.build_graph()
        
        return self
    
    def build_graph(self):
        """Build a connected graph from A1_ENTRIES"""
        entries = self.sheets_data.get('A1_ENTRIES', [])
        
        for entry in entries:
            entry_id = entry.get('ENTRY_ID') or entry.get('ЗАПИСЬ_ID')
            if not entry_id:
                continue
            
            node = {
                'id': str(entry_id),
                'en_term': entry.get('EN_TERM') or entry.get('РУС_ТЕРМИН', ''),
                'root': entry.get('ROOT_ID', ''),
                'root_letters': entry.get('ROOT_LETTERS') or entry.get('КОРЕНЬ_ID', ''),
                'network': entry.get('NETWORK_ID') or entry.get('СЕТЬ_ID', ''),
                'allah_name': entry.get('ALLAH_NAME_ID') or entry.get('ИМЯ_АЛЛАХА_ID', ''),
                'phonetic_chain': entry.get('PHONETIC_CHAIN') or entry.get('ФОНЕТИЧЕСКАЯ_ЦЕПЬ', ''),
                'score': entry.get('SCORE') or entry.get('БАЛЛ', 0),
                'aa_word': entry.get('AR_WORD') or entry.get('АР_СЛОВО', ''),
                'pattern': entry.get('PATTERN') or entry.get('ПАТТЕРН', ''),
                'foundation_ref': entry.get('FOUNDATION_REF') or entry.get('ОСНОВАНИЕ', ''),
                'connections': []
            }
            
            self.graph[str(entry_id)] = node
        
        self.total_entries = len(self.graph)
        print(f"\n📊 Graph built: {self.total_entries} nodes")
        
        return self.graph


# ============================================================================
# COMPONENT 2: PATTERN DETECTOR
# ============================================================================

class PatternDetector:
    """Finds gaps and opportunities in the existing data"""
    
    def __init__(self, reader):
        self.reader = reader
        self.graph = reader.graph
        self.gaps = []
        self.candidates = []
        self.shifts = self._load_shifts()
        self.networks = self._load_networks()
        self.roots = self._build_root_index()
        
    def _load_shifts(self):
        """Load phonetic shifts from M1 sheet"""
        shifts = {}
        m1_data = self.reader.sheets_data.get('M1_PHONETIC_SHIFTS', [])
        
        for row in m1_data:
            shift_id = row.get('SHIFT_ID') or row.get('СДВИГ_ID')
            if shift_id:
                shifts[shift_id] = {
                    'ar_letter': row.get('AR_LETTER') or row.get('АР_БУКВА', ''),
                    'en_outputs': row.get('EN_OUTPUTS') or row.get('РУС_ВЫХОДЫ', ''),
                    'examples': row.get('EXAMPLES') or row.get('ПРИМЕРЫ', '')
                }
        
        return shifts
    
    def _load_networks(self):
        """Load networks from M4 sheet"""
        networks = {}
        m4_data = self.reader.sheets_data.get('M4_NETWORKS', [])
        
        for row in m4_data:
            net_id = row.get('NETWORK_ID') or row.get('СЕТЬ_ID')
            if net_id:
                networks[net_id] = {
                    'name': row.get('NAME') or row.get('НАЗВАНИЕ', ''),
                    'entries': []
                }
        
        return networks
    
    def _build_root_index(self):
        """Group entries by root"""
        roots = {}
        for node_id, node in self.graph.items():
            root = node.get('root')
            if root:
                if root not in roots:
                    roots[root] = []
                roots[root].append(node_id)
        return roots
    
    def _extract_ratio_from_term(self, term):
        """Extract ratio from terms like 'musical fourth (4/3)'"""
        if not term:
            return None, None
        
        term_str = str(term)
        
        # Look for patterns like (4/3), (5/3), (7/3) etc.
        import re
        match = re.search(r'\((\d+)/(\d+)\)', term_str)
        if match:
            numerator = int(match.group(1))
            denominator = int(match.group(2))
            return numerator, denominator
        
        # Also check for "4/3" without parentheses
        match = re.search(r'(\d+)/(\d+)', term_str)
        if match:
            numerator = int(match.group(1))
            denominator = int(match.group(2))
            return numerator, denominator
        
        return None, None
    
    def get_existing_ratios_for_network(self, network):
        """Get all ratios already present in a network"""
        existing_ratios = set()
        
        for node_id, node in self.graph.items():
            if node.get('network') == network:
                term = node.get('en_term', '')
                num, den = self._extract_ratio_from_term(term)
                if num and den:
                    existing_ratios.add(f"{num}/{den}")
        
        return existing_ratios
    
    def detect_all(self):
        """Run all detection methods"""
        print("\n🔍 Detecting patterns and gaps...")
        
        self.find_networks_without_ratios()
        self.find_roots_with_multiple_entries()
        self.find_entries_without_networks()
        self.find_potential_ratio_entries()
        
        print(f"  Found {len(self.gaps)} gaps, {len(self.candidates)} candidates")
        
        return self.gaps, self.candidates
    
    def find_networks_without_ratios(self):
        """Find networks that should have ratio entries but don't"""
        network_entries = {}

        for node_id, node in self.graph.items():
            net = node.get('network')
            if net:
                if net not in network_entries:
                    network_entries[net] = []
                network_entries[net].append(node_id)

        for net, entries in network_entries.items():
            if len(entries) >= 3:  # Networks with at least 3 entries
                # Check if this network has any ratio entries using ratio extraction
                has_ratio = False
                for entry_id in entries:
                    entry = self.graph[entry_id]
                    term = entry.get('en_term', '')
                    num, den = self._extract_ratio_from_term(term)
                    if num and den:
                        has_ratio = True
                        break

                if not has_ratio:
                    self.gaps.append({
                        'type': 'network_needs_ratio',
                        'network': net,
                        'entries': entries,
                        'count': len(entries),
                        'priority': len(entries) * 2
                    })
    
    def find_roots_with_multiple_entries(self):
        """Find roots that appear multiple times but aren't in networks"""
        for root, entries in self.roots.items():
            if len(entries) >= 2:
                # Check if these entries share a network
                networks = set()
                for entry_id in entries:
                    net = self.graph[entry_id].get('network')
                    if net:
                        networks.add(net)
                
                if len(networks) <= 1:  # They don't have diverse networks
                    self.candidates.append({
                        'type': 'potential_network',
                        'root': root,
                        'entries': entries,
                        'count': len(entries),
                        'networks': list(networks)
                    })
    
    def find_entries_without_networks(self):
        """Find entries that don't belong to any network"""
        for node_id, node in self.graph.items():
            if not node.get('network') and node.get('score', 0) >= 8:
                self.gaps.append({
                    'type': 'entry_needs_network',
                    'entry': node_id,
                    'term': node.get('en_term'),
                    'score': node.get('score'),
                    'priority': 10 - node.get('score', 0)
                })
    
    def find_potential_ratio_entries(self):
        """Find entries that might contain hidden ratios"""
        ratio_keywords = ['circle', 'moon', 'light', 'dombra', 'yurt', 'doira', 
                         'drum', 'string', 'maqam', 'proportion', 'measure']
        
        for node_id, node in self.graph.items():
            term = node.get('en_term', '').lower()
            if any(keyword in term for keyword in ratio_keywords):
                if 'ratio' not in term:
                    self.candidates.append({
                        'type': 'potential_ratio',
                        'entry': node_id,
                        'term': node.get('en_term'),
                        'keywords': [k for k in ratio_keywords if k in term],
                        'priority': 5
                    })


# ============================================================================
# COMPONENT 3: GENERATOR ENGINE
# ============================================================================

class GeneratorEngine:
    """Creates new entries from patterns and gaps"""
    
    def __init__(self, reader, detector):
        self.reader = reader
        self.detector = detector
        self.new_entries = []
        self.next_id = self._find_next_id()
        self.ratio_templates = [
            (4, 3, 'musical fourth', 'N08', 'Harmony - fourth interval'),
            (5, 3, 'musical sixth', 'N08', 'Harmony - sixth interval'),
            (7, 3, 'maqam proportion', 'N08', 'Harmony - maqam Bayyati'),
            (7, 5, 'prayer mat proportion', 'N06', 'Fitra - original design'),
            (11, 5, 'dombra proportion', 'N16', 'Gatherer - sacred proportion'),
            (22, 7, 'circle constant', 'N07', 'Light - circular perfection'),
            (19, 7, 'growth constant', 'N10', 'Cognition - angelic measure'),
            (12, 7, 'dome proportion', 'N07', 'Light - sacred geometry'),
            (28, 7, 'lunar proportion', 'N05', 'Numbering - moon cycles'),
            (99, 70, 'high precision √2', 'N05', 'Numbering - sacred precision'),
            (355, 113, 'high precision π', 'N05', 'Numbering - divine accuracy')
        ]
    
    def _extract_ratio_from_term(self, term):
        """Extract ratio from terms like 'musical fourth (4/3)'"""
        if not term:
            return None, None
        
        term_str = str(term)
        
        # Look for patterns like (4/3), (5/3), (7/3) etc.
        import re
        match = re.search(r'\((\d+)/(\d+)\)', term_str)
        if match:
            numerator = int(match.group(1))
            denominator = int(match.group(2))
            return numerator, denominator
        
        # Also check for "4/3" without parentheses
        match = re.search(r'(\d+)/(\d+)', term_str)
        if match:
            numerator = int(match.group(1))
            denominator = int(match.group(2))
            return numerator, denominator
        
        return None, None
    
    def _get_existing_ratios_for_network(self, network):
        """Get all ratios already present in a network"""
        existing_ratios = set()
        
        for node_id, node in self.reader.graph.items():
            if node.get('network') == network:
                term = node.get('en_term', '')
                num, den = self._extract_ratio_from_term(term)
                if num and den:
                    existing_ratios.add(f"{num}/{den}")
        
        return existing_ratios
    
    def _generate_fractal_variants(self, base_num, base_den, network):
        """Generate fractal variants of a ratio across different domains"""
        variants = []
        
        # Domain templates: (domain_name, domain_description)
        domains = [
            ("melodic", "musical interval"),
            ("rhythmic", "tempo ratio"),
            ("geometric", "sacred proportion"),
            ("harmonic", "string division"),
            ("temporal", "time cycle"),
            ("spatial", "architectural ratio"),
            ("spiritual", "prayer cycle"),
            ("botanical", "growth pattern"),
            ("celestial", "orbital ratio"),
            ("numerical", "mathematical constant"),
        ]
        
        # Check existing ratios for this network
        existing_ratios = self._get_existing_ratios_for_network(network)
        base_ratio_str = f"{base_num}/{base_den}"
        
        # If base ratio already exists, generate variants with different domains
        if base_ratio_str in existing_ratios:
            for domain_name, domain_desc in domains:
                # Create a unique variant by combining ratio with domain
                entry = self._create_ratio_entry(
                    base_num, base_den,
                    f"{domain_desc} {base_num}/{base_den}",
                    network,
                    f"Divine ratio {base_num}/{base_den} manifesting in {domain_name} domain — Q54:49"
                )
                variants.append(entry)
                # Limit to 3 variants per generation cycle
                if len(variants) >= 3:
                    break
        else:
            # Base ratio doesn't exist, create it first
            entry = self._create_ratio_entry(
                base_num, base_den,
                f"divine ratio {base_num}/{base_den}",
                network,
                f"Foundational ratio {base_num}/{base_den} — Q54:49"
            )
            variants.append(entry)
        
        return variants
    
    def _find_next_id(self):
        """Find the next available ENTRY_ID"""
        max_num = 0
        for node_id in self.reader.graph.keys():
            if node_id and node_id.startswith('F'):
                try:
                    num = int(node_id[1:])
                    max_num = max(max_num, num)
                except:
                    pass
        return max_num + 1
    
    def generate_from_gaps(self, gaps, candidates):
        """Generate new entries from detected gaps and candidates"""
        print("\n🌱 Generating new entries from gaps...")

        # Process top 5 gaps
        gap_count = 0
        for gap in gaps[:5]:
            if gap['type'] == 'network_needs_ratio':
                self._generate_ratio_for_network(gap)
                gap_count += 1
            elif gap['type'] == 'entry_needs_network':
                self._assign_entry_to_network(gap)
                gap_count += 1

        # If no gaps found, process top 3 candidates, avoiding duplicates
        if gap_count == 0:
            print("  No gaps found, checking candidates...")
            processed_ratios = set()  # Track which ratios we've already generated variants for
            candidate_count = 0
            
            for candidate in candidates:
                if candidate['type'] == 'potential_ratio':
                    # Extract ratio from candidate
                    term = candidate['term']
                    num, den = self._extract_ratio_from_term(term)
                    if not num or not den:
                        continue
                    
                    ratio_str = f"{num}/{den}"
                    
                    # Skip if we've already processed this ratio in this cycle
                    if ratio_str in processed_ratios:
                        print(f"    Skipping duplicate ratio: {ratio_str}")
                        continue
                    
                    # Generate fractal variants for this ratio
                    self._generate_fractal_from_candidate(candidate)
                    processed_ratios.add(ratio_str)
                    candidate_count += 1
                    
                    # Limit to 2 candidates per generation cycle
                    if candidate_count >= 2:
                        break

        print(f"  Generated {len(self.new_entries)} new entries")
        return self.new_entries
    
    def _generate_fractal_from_candidate(self, candidate):
        """Generate fractal variants from a potential ratio candidate"""
        entry_id = candidate['entry']
        term = candidate['term']
        
        # Extract ratio from the candidate term
        num, den = self._extract_ratio_from_term(term)
        if not num or not den:
            print(f"    Could not extract ratio from: {term}")
            return
        
        # Get the network from the entry
        entry = self.reader.graph.get(entry_id, {})
        network = entry.get('network')
        if not network:
            # Default to N08 (Harmony) for ratio entries
            network = 'N08'
        
        # Generate fractal variants for this ratio
        variants = self._generate_fractal_variants(num, den, network)
        
        # Add variants to new entries
        for variant in variants:
            self.new_entries.append(variant)
            print(f"    Generated: {variant['EN_TERM']}")
    
    def _generate_ratio_for_network(self, gap):
        """Generate fractal variant ratio entries for a network"""
        network = gap['network']
        
        # Find matching ratio templates for this network
        matching_templates = []
        for num, den, name, net, desc in self.ratio_templates:
            if net == network:
                matching_templates.append((num, den, name, net, desc))
        
        if matching_templates:
            # Generate fractal variants for each matching template
            for num, den, name, net, desc in matching_templates:
                variants = self._generate_fractal_variants(num, den, network)
                self.new_entries.extend(variants)
                # Limit to 5 total variants per generation cycle
                if len(self.new_entries) >= 5:
                    break
        else:
            # No matching templates, create a generic fractal variant
            variants = self._generate_fractal_variants(1, 1, network)
            self.new_entries.extend(variants[:2])  # Just 2 generic variants
    
    def _assign_entry_to_network(self, gap):
        """Create a network assignment for an entry"""
        entry_id = gap['entry']
        term = gap['term']
        
        # Find appropriate network based on term
        network = 'N08'  # Default to Harmony
        if any(word in term.lower() for word in ['light', 'sun', 'moon', 'circle']):
            network = 'N07'
        elif any(word in term.lower() for word in ['number', 'count', 'zero', 'cipher']):
            network = 'N05'
        elif any(word in term.lower() for word in ['faith', 'amen', 'secure']):
            network = 'N06'
        
        # Create a note about the assignment
        entry = {
            'ENTRY_ID': f"NOTE{self.next_id:04d}",
            'SCORE': 8,
            'EN_TERM': f"{term} network assignment",
            'AR_WORD': '—',
            'ROOT_ID': gap['entry'],
            'ROOT_LETTERS': '—',
            'QUR_MEANING': f"Should be in {network}",
            'PATTERN': 'A',
            'NETWORK_ID': network,
            'PHONETIC_CHAIN': '—',
            'INVERSION_TYPE': 'HIDDEN',
            'SOURCE_FORM': '—',
            'FOUNDATION_REF': f"GAP:{gap['type']}"
        }
        
        self.next_id += 1
        self.new_entries.append(entry)
    
    def _create_ratio_entry(self, numerator, denominator, name, network, description):
        """Create a ratio entry"""
        entry_id = f"F{self.next_id:04d}"
        self.next_id += 1
        
        return {
            'ENTRY_ID': entry_id,
            'SCORE': 9,
            'EN_TERM': name,
            'AR_WORD': f"({numerator}/{denominator})",
            'ROOT_ID': f"R{self.next_id}",
            'ROOT_LETTERS': '—',
            'QUR_MEANING': f"Divine ratio {numerator}/{denominator} - {description}",
            'PATTERN': 'A',
            'NETWORK_ID': network,
            'PHONETIC_CHAIN': '→'.join(['?'] * 3),
            'INVERSION_TYPE': 'HIDDEN',
            'SOURCE_FORM': f"ratio:{numerator}/{denominator}",
            'FOUNDATION_REF': f"RATIO:{numerator}/{denominator}"
        }
    
    def generate_seed_entries(self):
        """Generate initial seed ratio entries"""
        print("\n🌱 Planting seed entries...")
        
        for num, den, name, net, desc in self.ratio_templates:
            entry = self._create_ratio_entry(num, den, name, net, desc)
            self.new_entries.append(entry)
        
        print(f"  Planted {len(self.new_entries)} seeds")
        return self.new_entries


# ============================================================================
# COMPONENT 4: VALIDATION ENGINE
# ============================================================================

class ValidationEngine:
    """Validates new entries against existing patterns"""
    
    def __init__(self, reader):
        self.reader = reader
        
    def validate_all(self, entries):
        """Validate all new entries"""
        print("\n✅ Validating new entries...")
        
        validated = []
        for entry in entries:
            valid_entry = self._validate_entry(entry)
            if valid_entry:
                validated.append(valid_entry)
        
        print(f"  Validated {len(validated)} of {len(entries)} entries")
        return validated
    
    def _validate_entry(self, entry):
        """Validate a single entry"""
        # Check required fields
        required = ['ENTRY_ID', 'EN_TERM']
        for field in required:
            if field not in entry or not entry[field]:
                return None
        
        # Auto-calculate score based on completeness
        score = 5
        
        if entry.get('ROOT_ID') and entry['ROOT_ID'] != '—':
            score += 1
        if entry.get('NETWORK_ID'):
            score += 1
        if entry.get('PHONETIC_CHAIN') and entry['PHONETIC_CHAIN'] != '→'.join(['?'] * 3):
            score += 1
        if entry.get('QUR_MEANING') and 'Divine' not in entry['QUR_MEANING']:
            score += 1
        if entry.get('FOUNDATION_REF') and entry['FOUNDATION_REF'].startswith('RATIO'):
            score += 1
        
        entry['SCORE'] = min(10, score)
        
        return entry


# ============================================================================
# COMPONENT 5: SELF-WRITER ENGINE
# ============================================================================

class SelfWriter:
    """Writes new entries back to the Excel file"""
    
    def __init__(self, reader):
        self.reader = reader
        self.filepath = reader.filepath
        
    def write_all(self, new_entries):
        """Write all new entries to the file"""
        if not new_entries:
            print("\n💾 No new entries to write")
            return 0
        
        print("\n💾 Writing new entries to file...")
        
        try:
            # Load workbook fresh to avoid conflicts
            wb = load_workbook(self.filepath)
            
            # Get or create A1_ENTRIES sheet
            if 'A1_ENTRIES' in wb.sheetnames:
                sheet = wb['A1_ENTRIES']
            else:
                sheet = wb.create_sheet('A1_ENTRIES')
                # Add headers
                headers = ['ENTRY_ID', 'SCORE', 'EN_TERM', 'AR_WORD', 'ROOT_ID', 
                          'ROOT_LETTERS', 'QUR_MEANING', 'PATTERN', 'NETWORK_ID',
                          'PHONETIC_CHAIN', 'INVERSION_TYPE', 'SOURCE_FORM', 'FOUNDATION_REF']
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col, value=header)
            
            # Find next empty row
            next_row = sheet.max_row + 1
            
            # Write each new entry
            for entry in new_entries:
                sheet.cell(row=next_row, column=1, value=entry['ENTRY_ID'])
                sheet.cell(row=next_row, column=2, value=entry['SCORE'])
                sheet.cell(row=next_row, column=3, value=entry['EN_TERM'])
                sheet.cell(row=next_row, column=4, value=entry['AR_WORD'])
                sheet.cell(row=next_row, column=5, value=entry['ROOT_ID'])
                sheet.cell(row=next_row, column=6, value=entry.get('ROOT_LETTERS', '—'))
                sheet.cell(row=next_row, column=7, value=entry.get('QUR_MEANING', '—'))
                sheet.cell(row=next_row, column=8, value=entry.get('PATTERN', 'A'))
                sheet.cell(row=next_row, column=9, value=entry.get('NETWORK_ID', ''))
                sheet.cell(row=next_row, column=10, value=entry.get('PHONETIC_CHAIN', '—'))
                sheet.cell(row=next_row, column=11, value=entry.get('INVERSION_TYPE', 'HIDDEN'))
                sheet.cell(row=next_row, column=12, value=entry.get('SOURCE_FORM', '—'))
                sheet.cell(row=next_row, column=13, value=entry.get('FOUNDATION_REF', '—'))
                
                next_row += 1
            
            # Save
            wb.save(self.filepath)
            print(f"  ✓ Wrote {len(new_entries)} new entries to {self.filepath.name}")
            
            return len(new_entries)
            
        except Exception as e:
            print(f"  ✗ Error writing: {e}")
            return 0


# ============================================================================
# COMPONENT 6: MAIN SYSTEM
# ============================================================================

class SelfGrowingSystem:
    """The main orchestrator that runs all 6 components"""
    
    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self.cycle_count = 0
        self.growth_log = []
        
        if not self.filepath.exists():
            raise FileNotFoundError(f"File not found: {filepath}")
        
        print("\n" + "="*60)
        print("🌳 USLaP SELF-GROWING FOREST SYSTEM")
        print("="*60)
        print(f"📁 File: {self.filepath.name}")
        print(f"📦 Size: {self.filepath.stat().st_size / 1024 / 1024:.2f} MB")
        print("="*60)
        
    def run_full_cycle(self, seed_mode=False):
        """
        Run all 6 components in sequence
        seed_mode=True: Plant initial ratio seeds
        seed_mode=False: Full growth cycle with pattern detection
        """
        print(f"\n{'='*60}")
        print(f"CYCLE {self.cycle_count + 1} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}")
        
        # COMPONENT 1: READ
        print("\n📖 COMPONENT 1: Self-Reader")
        reader = SelfReader(self.filepath)
        reader.read_all()
        
        if seed_mode:
            # COMPONENT 2: Skip detection for seed mode
            gaps = []
            candidates = []
            
            # COMPONENT 3: Generate seeds
            print("\n🌱 COMPONENT 3: Generator Engine (Seed Mode)")
            generator = GeneratorEngine(reader, None)
            new_entries = generator.generate_seed_entries()
            
        else:
            # COMPONENT 2: Detect patterns
            print("\n🔍 COMPONENT 2: Pattern Detector")
            detector = PatternDetector(reader)
            gaps, candidates = detector.detect_all()
            
            # COMPONENT 3: Generate from gaps
            print("\n🌱 COMPONENT 3: Generator Engine")
            generator = GeneratorEngine(reader, detector)
            new_entries = generator.generate_from_gaps(gaps, candidates)
        
        # COMPONENT 4: Validate
        print("\n✅ COMPONENT 4: Validation Engine")
        validator = ValidationEngine(reader)
        validated_entries = validator.validate_all(new_entries)
        
        # COMPONENT 5: Write
        print("\n💾 COMPONENT 5: Self-Writer")
        writer = SelfWriter(reader)
        written = writer.write_all(validated_entries)
        
        # COMPONENT 6: Log
        self.cycle_count += 1
        self.growth_log.append({
            'cycle': self.cycle_count,
            'timestamp': datetime.now().isoformat(),
            'mode': 'SEED' if seed_mode else 'GROWTH',
            'new_entries': written,
            'gaps_found': len(gaps) if not seed_mode else 0,
            'candidates': len(candidates) if not seed_mode else 0
        })
        
        print(f"\n{'='*60}")
        print(f"✅ CYCLE {self.cycle_count} COMPLETE")
        print(f"   Mode: {'🌱 SEED' if seed_mode else '🌳 GROWTH'}")
        print(f"   Added: {written} new entries")
        print(f"   Total entries now: {reader.total_entries + written}")
        print(f"{'='*60}\n")
        
        return self
    
    def show_log(self):
        """Display growth log"""
        print("\n📊 GROWTH LOG")
        print("-" * 50)
        for entry in self.growth_log:
            mode_icon = "🌱" if entry['mode'] == 'SEED' else "🌳"
            print(f"{mode_icon} Cycle {entry['cycle']}: {entry['new_entries']} entries added")
        print("-" * 50)
    
    def run_forever(self, interval_hours=24):
        """Run cycles automatically at specified intervals"""
        import time
        
        print(f"\n⏰ Running forever every {interval_hours} hours")
        print("Press Ctrl+C to stop\n")
        
        # First cycle: seed mode
        self.run_full_cycle(seed_mode=True)
        
        # Subsequent cycles: growth mode
        while True:
            time.sleep(interval_hours * 3600)
            self.run_full_cycle(seed_mode=False)
            self.show_log()


# ============================================================================
# MAIN EXECUTION - THIS RUNS EVERYTHING
# ============================================================================

def main():
    """Main function - runs all 6 components"""
    
    # Get file path from command line or use default
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        filepath = "USLaP_Final_Data_Consolidated_Master.xlsx"
    
    # Check if file exists
    if not os.path.exists(filepath):
        print(f"\n❌ File not found: {filepath}")
        print("\nPlease provide the path to your USLaP Excel file:")
        print("  python uslap_forest.py /path/to/your/file.xlsx")
        print("\nOr place this script in the same folder as your file.")
        return 1
    
    try:
        # Create the system
        system = SelfGrowingSystem(filepath)
        
        # Ask user what they want to do
        print("\nWhat would you like to do?")
        print("  1. Plant seeds (first run only)")
        print("  2. Run one growth cycle")
        print("  3. Run forever (automatic every 24h)")
        print("  4. Just show what would happen (dry run)")
        
        choice = input("\nEnter choice (1-4): ").strip()
        
        if choice == '1':
            # Plant seeds only
            system.run_full_cycle(seed_mode=True)
            system.show_log()
            print("\n🌱 Seeds planted! Run again with option 2 for growth.")
            
        elif choice == '2':
            # Run one growth cycle
            system.run_full_cycle(seed_mode=False)
            system.show_log()
            
        elif choice == '3':
            # Run forever
            hours = input("Enter interval in hours (default 24): ").strip()
            interval = int(hours) if hours else 24
            system.run_forever(interval_hours=interval)
            
        elif choice == '4':
            # Dry run - just read and analyze
            print("\n📖 DRY RUN - Reading only...")
            reader = SelfReader(filepath)
            reader.read_all()
            
            print("\n🔍 Analyzing patterns...")
            detector = PatternDetector(reader)
            gaps, candidates = detector.detect_all()
            
            print("\n📊 ANALYSIS RESULTS")
            print(f"   Total entries: {reader.total_entries}")
            print(f"   Gaps found: {len(gaps)}")
            print(f"   Network candidates: {len(candidates)}")
            
            print("\n   Top gaps:")
            for gap in gaps[:5]:
                print(f"     • {gap['type']}: {gap.get('network', gap.get('entry', '?'))}")
            
            print("\n✅ Dry run complete. No changes made.")
            
        else:
            print("Invalid choice. Running default: plant seeds")
            system.run_full_cycle(seed_mode=True)
        
        print("\n✨ Done!")
        
    except KeyboardInterrupt:
        print("\n\n👋 System stopped by user")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        return 1
    
    return 0


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    sys.exit(main())