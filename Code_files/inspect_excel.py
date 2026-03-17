#!/usr/bin/env python3
import openpyxl
import sys

def inspect_excel(filepath):
    print(f"Opening {filepath}...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    print("\n=== Sheet Names ===")
    for sheet_name in wb.sheetnames:
        print(f"- {sheet_name}")
    
    # Check the primary sheets mentioned
    primary_sheets = ["UMD_OPERATIONS", "CHILD_SCHEMA", "DP_REGISTER", "ATT_TERMS", "PHONETIC_REVERSAL", "SESSION_INDEX"]
    print("\n=== Primary Sheets Status ===")
    for sheet in primary_sheets:
        if sheet in wb.sheetnames:
            print(f"✓ {sheet} - exists")
        else:
            print(f"✗ {sheet} - MISSING")
    
    # Inspect first few rows of each primary sheet
    print("\n=== Sample Data (first 5 rows) ===")
    for sheet_name in primary_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- {sheet_name} ---")
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                # Print column headers (first row)
                print("Headers:", rows[0])
                # Print a few data rows
                for i, row in enumerate(rows[1:6], 1):
                    print(f"Row {i}:", row)
            else:
                print("Empty sheet")
    
    wb.close()
    print("\n=== Sheet Count ===")
    print(f"Total sheets: {len(wb.sheetnames)}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        inspect_excel(sys.argv[1])
    else:
        inspect_excel("USLaP_Final_Data_Consolidated_Master_v2.xlsx")