#!/usr/bin/env python3
"""
USLaP Phase 18: al-Jumuʿah — الجُمُعَة (Surah 62)
Excel Gathering/Sync Engine

Q62:9: فَٱسْعَوْا۟ إِلَىٰ ذِكْرِ ٱللَّهِ
"Hasten to the remembrance of Allah"

Syncs ALL DB tables to the master Excel file.
Creates backup before writing.

Usage:
    python3 uslap_jumuah.py status      # Show gaps
    python3 uslap_jumuah.py sync        # Full sync
    python3 uslap_jumuah.py sync --sheet A1_ENTRIES  # Single sheet
"""

import sqlite3
try:
    from uslap_db_connect import connect as _uslap_connect
    _HAS_WRAPPER = True
except ImportError:
    _HAS_WRAPPER = False, sys, os, shutil
from datetime import datetime

DB = os.path.join(os.path.dirname(__file__), 'uslap_database_v3.db')
XLSX = '/Users/mmsetubal/Documents/USLaP workplace/USLaP_Final_Data_Consolidated_Master_v3.xlsx'

# ═══════════════════════════════════════════════════════════
# SYNC MAP: DB table → Excel sheet name + column config
# ═══════════════════════════════════════════════════════════

SYNC_MAP = [
    {
        'db_table': 'a1_entries',
        'sheet': 'A1_ENTRIES',
        'columns': [
            ('entry_id', 'Entry ID'),
            ('score', 'Score'),
            ('en_term', 'English Term'),
            ('aa_word', 'Arabic Word'),
            ('root_id', 'Root ID'),
            ('root_letters', 'Root Letters'),
            ('qur_meaning', 'Quranic Meaning'),
            ('pattern', 'Pattern'),
            ('allah_name_id', 'Allah Name ID'),
            ('network_id', 'Network ID'),
            ('phonetic_chain', 'Phonetic Chain'),
            ('inversion_type', 'Inversion Type'),
            ('source_form', 'Source Form'),
            ('foundation_ref', 'Foundation Ref'),
        ]
    },
    {
        'db_table': 'a1_записи',
        'sheet': 'A1_ЗАПИСИ',
        'columns': [
            ('запись_id', 'Запись ID'),
            ('балл', 'Балл'),
            ('рус_термин', 'Русский Термин'),
            ('ар_слово', 'Арабское Слово'),
            ('корень_id', 'Корень ID'),
            ('корневые_буквы', 'Корневые Буквы'),
            ('коранич_значение', 'Кораническое Значение'),
            ('паттерн', 'Паттерн'),
            ('имя_аллаха_id', 'Имя Аллаха ID'),
            ('сеть_id', 'Сеть ID'),
            ('фонетическая_цепь', 'Фонетическая Цепь'),
            ('тип_инверсии', 'Тип Инверсии'),
            ('исходная_форма', 'Исходная Форма'),
            ('основание', 'Основание'),
        ]
    },
    {
        'db_table': 'persian_a1_mad_khil',
        'sheet': 'PERSIAN_A1_MADĀKHIL',
        'columns': None,  # Use DB column names directly (bilingual)
    },
    {
        'db_table': 'bitig_a1_entries',
        'sheet': 'BITIG_A1_ENTRIES',
        'columns': [
            ('entry_id', 'Entry ID'),
            ('score', 'Score'),
            ('orig2_term', 'ORIG2 Term'),
            ('orig2_script', 'ORIG2 Script'),
            ('root_letters', 'Root Letters'),
            ('kashgari_attestation', 'Kashgari Attestation'),
            ('ibn_sina_attestation', 'Ibn Sina Attestation'),
            ('modern_reflexes', 'Modern Reflexes'),
            ('navoi_attestation', 'Navoi Attestation'),
            ('downstream_forms', 'Downstream Forms'),
            ('phonetic_chain', 'Phonetic Chain'),
            ('semantic_field', 'Semantic Field'),
            ('dispersal_range', 'Dispersal Range'),
            ('status', 'Status'),
            ('notes', 'Notes'),
        ]
    },
    {
        'db_table': 'child_schema',
        'sheet': 'CHILD_SCHEMA',
        'columns': None,
    },
    {
        'db_table': 'a2_names_of_allah',
        'sheet': 'A2_NAMES_OF_ALLAH',
        'columns': None,
    },
    {
        'db_table': 'a3_quran_refs',
        'sheet': 'A3_QURAN_REFS',
        'columns': None,
    },
    {
        'db_table': 'a4_derivatives',
        'sheet': 'A4_DERIVATIVES',
        'columns': None,
    },
    {
        'db_table': 'a5_cross_refs',
        'sheet': 'A5_CROSS_REFS',
        'columns': None,
    },
    {
        'db_table': 'a6_country_names',
        'sheet': 'A6_COUNTRY_NAMES',
        'columns': None,
    },
    {
        'db_table': 'm4_networks',
        'sheet': 'M4_NETWORKS',
        'columns': None,
    },
    {
        'db_table': 'm3_scholars',
        'sheet': 'M3_SCHOLARS',
        'columns': None,
    },
    {
        'db_table': 'chronology',
        'sheet': 'CHRONOLOGY',
        'columns': None,
    },
    {
        'db_table': 'dp_register',
        'sheet': 'DP_REGISTER',
        'columns': None,
    },
    {
        'db_table': 'european_a1_entries',
        'sheet': 'EUROPEAN_A1_ENTRIES',
        'columns': None,
        'new_sheet': True,  # May not exist yet
    },
]


class Jumuah:
    """al-Jumuʿah: Excel Gathering Engine."""

    def __init__(self):
        self.conn = _uslap_connect(DB) if _HAS_WRAPPER else sqlite3.connect(DB)
    conn.execute("PRAGMA foreign_keys = ON")

    def status(self):
        """Show sync gaps between DB and Excel."""
        import openpyxl
        wb = openpyxl.load_workbook(XLSX, read_only=True)
        xl_sheets = {s: wb[s].max_row - 1 for s in wb.sheetnames}  # -1 for header
        wb.close()

        print("═══ al-Jumuʿah: Sync Gap Analysis ═══")
        print(f"{'Sheet':35s} {'Excel':>7s} {'DB':>7s} {'Gap':>7s}")
        print("─" * 60)

        total_gap = 0
        for sm in SYNC_MAP:
            tbl = sm['db_table']
            sheet = sm['sheet']
            try:
                db_cnt = self.conn.execute(f'SELECT COUNT(*) FROM "{tbl}"').fetchone()[0]
            except:
                db_cnt = 0

            xl_cnt = xl_sheets.get(sheet, 0)
            if xl_cnt < 0:
                xl_cnt = 0
            gap = db_cnt - xl_cnt
            total_gap += max(gap, 0)

            marker = '✓' if gap == 0 else f'+{gap}' if gap > 0 else str(gap)
            print(f"  {sheet:33s} {xl_cnt:7d} {db_cnt:7d} {marker:>7s}")

        print("─" * 60)
        print(f"  {'TOTAL NEW ROWS':33s} {'':7s} {'':7s} {'+' + str(total_gap):>7s}")

    def sync(self, target_sheet=None):
        """Sync DB to Excel. Optionally sync a single sheet."""
        import openpyxl

        # Step 1: Backup
        ts = datetime.now().strftime('%Y%m%d_%H%M')
        backup = XLSX.replace('.xlsx', f'_backup_{ts}.xlsx')
        shutil.copy2(XLSX, backup)
        print(f"✓ Backup: {os.path.basename(backup)}")

        # Step 2: Open workbook (preserve everything)
        wb = openpyxl.load_workbook(XLSX)

        synced = 0
        for sm in SYNC_MAP:
            sheet_name = sm['sheet']

            # Filter to single sheet if requested
            if target_sheet and sheet_name != target_sheet:
                continue

            tbl = sm['db_table']

            # Get DB data
            try:
                rows = self.conn.execute(f'SELECT * FROM "{tbl}"').fetchall()
                col_names = [c[1] for c in self.conn.execute(f'PRAGMA table_info("{tbl}")').fetchall()]
            except Exception as e:
                print(f"  ✗ {sheet_name}: DB error — {e}")
                continue

            if not rows:
                print(f"  ⊘ {sheet_name}: empty table, skipping")
                continue

            # Get or create sheet
            if sheet_name in wb.sheetnames:
                # Delete existing sheet and recreate
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)

            # Write headers
            if sm.get('columns'):
                headers = [c[1] for c in sm['columns']]
            else:
                headers = col_names

            for c_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c_idx, value=h)
                cell.font = openpyxl.styles.Font(bold=True)

            # Write data rows
            for r_idx, row in enumerate(rows, 2):
                for c_idx, val in enumerate(row, 1):
                    if c_idx <= len(headers):
                        ws.cell(row=r_idx, column=c_idx, value=val)

            # Auto-width (approximate)
            for c_idx, h in enumerate(headers, 1):
                col_letter = openpyxl.utils.get_column_letter(c_idx)
                ws.column_dimensions[col_letter].width = min(max(len(str(h)) + 2, 12), 50)

            print(f"  ✓ {sheet_name:33s} → {len(rows)} rows ({len(headers)} cols)")
            synced += 1

        # Step 3: Save
        wb.save(XLSX)
        wb.close()

        print(f"\n✓ Synced {synced} sheets to {os.path.basename(XLSX)}")
        print(f"  File size: {os.path.getsize(XLSX) / 1024 / 1024:.1f} MB")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 uslap_jumuah.py [status|sync] [--sheet SHEET_NAME]")
        return

    cmd = sys.argv[1]
    jum = Jumuah()

    if cmd == 'status':
        jum.status()
    elif cmd == 'sync':
        target = None
        if '--sheet' in sys.argv:
            idx = sys.argv.index('--sheet')
            if idx + 1 < len(sys.argv):
                target = sys.argv[idx + 1]
        jum.sync(target_sheet=target)
    else:
        print(f"Unknown command: {cmd}")


if __name__ == '__main__':
    main()
