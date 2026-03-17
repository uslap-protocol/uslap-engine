#!/usr/bin/env python3
"""
Build SQLite database from USLaP_Final_Data_Consolidated_Master_v2.xlsx
Preserves UTF-8 Arabic text exactly as stored.
Creates one table per primary sheet.
Adds cross-reference table linking ENTRY_ID → PARENT_OP → DP_REF.
"""

import sqlite3
import openpyxl
import re
from pathlib import Path

def clean_column_name(col):
    """Convert Excel column header to valid SQLite column name."""
    if col is None:
        return "unknown"
    # Remove non-alphanumeric, replace spaces with underscore
    col = str(col).strip()
    col = re.sub(r'[^\w\s]', '', col)  # Remove punctuation
    col = re.sub(r'\s+', '_', col)     # Replace spaces with underscore
    col = col.lower()
    if not col:
        return "unknown"
    return col

def extract_parent_ops(parent_op_str):
    """Extract individual parent operations from strings like 'UMD-RL1 + UMD-ST1'"""
    if not parent_op_str:
        return []
    # Split by various separators
    ops = re.split(r'[·,\+\s]+', str(parent_op_str))
    # Filter for operation codes (UMD- prefix)
    parent_ops = []
    for op in ops:
        op = op.strip()
        if op and re.match(r'^UMD-', op):
            parent_ops.append(op)
    return parent_ops

def extract_dp_codes(dp_string):
    """Extract individual DP codes from strings like 'DP08 · DP07 · DP11 · DP15'"""
    if not dp_string:
        return []
    # Split by various separators
    codes = re.split(r'[·,\s]+', str(dp_string))
    # Filter for DP codes (start with DP followed by digits or word)
    dp_codes = []
    for code in codes:
        code = code.strip()
        if code and (re.match(r'^DP\d+', code) or re.match(r'^DP-', code)):
            dp_codes.append(code)
    return dp_codes

def create_tables(conn, cursor):
    """Create tables for each primary sheet."""
    
    # Create UMD_OPERATIONS table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS umd_operations (
        op_id TEXT PRIMARY KEY,
        op_name TEXT,
        op_class TEXT,
        qur_primary TEXT,
        qur_secondary TEXT,
        op_structure TEXT,
        founding_instances TEXT,
        dp_always_active TEXT,
        gate_shortcut TEXT,
        np_layer TEXT,
        darvo_active TEXT,
        notes TEXT,
        status TEXT,
        last_updated TEXT
    )
    ''')
    
    # Create CHILD_SCHEMA table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS child_schema (
        entry_id TEXT PRIMARY KEY,
        shell_name TEXT,
        shell_language TEXT,
        orig_class TEXT,
        orig_root TEXT,
        orig_lemma TEXT,
        orig_meaning TEXT,
        operation_role TEXT,
        shell_meaning TEXT,
        inversion_direction TEXT,
        phonetic_chain TEXT,
        qur_anchors TEXT,
        dp_codes TEXT,
        nt_code TEXT,
        pattern TEXT,
        parent_op TEXT,
        gate_status TEXT,
        notes TEXT
    )
    ''')
    
    # Create DP_REGISTER table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS dp_register (
        dp_code TEXT PRIMARY KEY,
        name TEXT,
        class TEXT,
        trigger TEXT,
        mechanism TEXT,
        qur_anchor TEXT,
        example TEXT,
        distinct_from TEXT,
        protocol_note TEXT,
        status TEXT
    )
    ''')
    
    # Create ATT_TERMS table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS att_terms (
        term_id TEXT PRIMARY KEY,
        arabic TEXT,
        transliteration TEXT,
        translation TEXT,
        root TEXT,
        qur_anchor TEXT,
        function_in_lattice TEXT,
        umd_op_ref TEXT,
        dp_ref TEXT,
        inversion_type TEXT,
        notes TEXT
    )
    ''')
    
    # Create PHONETIC_REVERSAL table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS phonetic_reversal (
        shift_code TEXT PRIMARY KEY,
        from_modern TEXT,
        to_orig TEXT,
        class TEXT,
        mechanism TEXT,
        attested_example TEXT,
        entry_ref TEXT,
        reliability TEXT,
        notes TEXT,
        status TEXT
    )
    ''')
    
    # Create SESSION_INDEX table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS session_index (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        item_type TEXT,
        entry_id TEXT,
        description TEXT,
        sheet TEXT,
        status TEXT,
        notes TEXT
    )
    ''')
    
    # Create cross-reference table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS cross_reference (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entry_id TEXT,
        parent_op TEXT,
        dp_ref TEXT,
        FOREIGN KEY (entry_id) REFERENCES child_schema (entry_id),
        FOREIGN KEY (parent_op) REFERENCES umd_operations (op_id)
    )
    ''')
    
    conn.commit()
    print("Created all tables")

def import_sheet_data(conn, cursor, wb, sheet_name, table_name):
    """Import data from an Excel sheet to SQLite table."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        print(f"  No data in {sheet_name}")
        return 0
    
    # Find header row based on sheet-specific patterns
    header_row = None
    data_start = 0
    
    # Special handling for each sheet based on debug output
    if sheet_name == "DP_REGISTER":
        # Row 0: title, Row 1: description, Row 2: headers
        if len(rows) > 2:
            header_row = rows[2]
            data_start = 3
    elif sheet_name == "PHONETIC_REVERSAL":
        # Row 0: title, Row 1: description, Row 2: zone header, Row 3: headers
        if len(rows) > 3:
            header_row = rows[3]
            data_start = 4
    elif sheet_name == "UMD_OPERATIONS":
        # Row 0: title, Row 1: description, Row 2: headers
        if len(rows) > 2:
            header_row = rows[2]
            data_start = 3
    elif sheet_name == "CHILD_SCHEMA":
        # Row 0: title, Row 1: description, Row 2: headers
        if len(rows) > 2:
            header_row = rows[2]
            data_start = 3
    elif sheet_name == "ATT_TERMS":
        # Row 0: title, Row 1: description, Row 2: headers
        if len(rows) > 2:
            header_row = rows[2]
            data_start = 3
    elif sheet_name == "SESSION_INDEX":
        # Row 0: title, Row 1: description, Row 2: headers
        if len(rows) > 2:
            header_row = rows[2]
            data_start = 3
    else:
        # Generic fallback: find row with typical column names
        for i, row in enumerate(rows):
            if row and any(isinstance(cell, str) and ('ID' in cell or 'NAME' in cell or 'CODE' in cell) for cell in row):
                header_row = row
                data_start = i + 1
                break
    
    if header_row is None:
        print(f"  Could not find headers in {sheet_name}")
        return 0
    
    # Clean column names
    col_names = []
    for cell in header_row:
        col_name = clean_column_name(cell)
        # Handle edge case where title row got mistaken as header
        if col_name and 'uslap' in col_name and any(x in col_name for x in ['title', 'header', 'master']):
            # This looks like a title, not a column name
            col_names.append(clean_column_name(str(header_row.index(cell))))
        else:
            col_names.append(col_name)
    
    # Create placeholders for SQL
    placeholders = ', '.join(['?' for _ in col_names])
    col_list = ', '.join(col_names)
    
    # Prepare insert statement
    insert_sql = f"INSERT OR REPLACE INTO {table_name} ({col_list}) VALUES ({placeholders})"
    
    # Insert data rows
    count = 0
    for i in range(data_start, len(rows)):
        row = rows[i]
        # Skip empty rows or rows where first cell is None
        if not row or row[0] is None:
            continue
        
        # Ensure row has same length as columns
        row_data = list(row)
        while len(row_data) < len(col_names):
            row_data.append(None)
        row_data = row_data[:len(col_names)]
        
        try:
            cursor.execute(insert_sql, row_data)
            count += 1
        except Exception as e:
            print(f"  Error inserting row {i} in {sheet_name}: {e}")
            print(f"  Row data: {row_data}")
    
    conn.commit()
    print(f"  Imported {count} rows into {table_name}")
    return count

def build_cross_reference(conn, cursor):
    """Build cross-reference table from CHILD_SCHEMA data."""
    cursor.execute("DELETE FROM cross_reference")
    
    # Get all child schema entries
    cursor.execute("SELECT entry_id, parent_op, dp_codes FROM child_schema")
    rows = cursor.fetchall()
    
    count = 0
    for entry_id, parent_op_str, dp_codes_str in rows:
        if not entry_id or entry_id == 'ENTRY_ID':
            continue
        
        # Extract individual parent operations (e.g., "UMD-RL1 + UMD-ST1" -> ["UMD-RL1", "UMD-ST1"])
        parent_ops = extract_parent_ops(parent_op_str)
        
        # Extract individual DP codes
        dp_codes = extract_dp_codes(dp_codes_str)
        
        # Insert one row per parent_op per dp_code
        for parent_op in parent_ops:
            # Verify parent_op exists in umd_operations
            cursor.execute("SELECT op_id FROM umd_operations WHERE op_id = ?", (parent_op,))
            if not cursor.fetchone():
                print(f"  Warning: Parent operation {parent_op} not found in umd_operations for entry {entry_id}")
                continue
                
            for dp_code in dp_codes:
                cursor.execute(
                    "INSERT INTO cross_reference (entry_id, parent_op, dp_ref) VALUES (?, ?, ?)",
                    (entry_id, parent_op, dp_code)
                )
                count += 1
    
    conn.commit()
    print(f"  Built {count} cross-reference entries")
    return count

def main():
    excel_path = "USLaP_Final_Data_Consolidated_Master_v2.xlsx"
    db_path = "uslap_database.db"
    
    print(f"Opening Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    print(f"Creating SQLite database: {db_path}")
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Enable foreign keys
    cursor.execute("PRAGMA foreign_keys = ON")
    
    # Create tables
    create_tables(conn, cursor)
    
    # Import data from each primary sheet
    primary_sheets = [
        ("UMD_OPERATIONS", "umd_operations"),
        ("CHILD_SCHEMA", "child_schema"),
        ("DP_REGISTER", "dp_register"),
        ("ATT_TERMS", "att_terms"),
        ("PHONETIC_REVERSAL", "phonetic_reversal"),
        ("SESSION_INDEX", "session_index")
    ]
    
    total_rows = 0
    for excel_sheet, db_table in primary_sheets:
        if excel_sheet in wb.sheetnames:
            print(f"\nImporting {excel_sheet} -> {db_table}")
            count = import_sheet_data(conn, cursor, wb, excel_sheet, db_table)
            total_rows += count
        else:
            print(f"\nSheet {excel_sheet} not found in Excel file")
    
    # Build cross-reference table
    print("\nBuilding cross-reference table...")
    cross_ref_count = build_cross_reference(conn, cursor)
    
    # Create indexes for faster searching
    print("\nCreating indexes...")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_child_schema_entry_id ON child_schema(entry_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_child_schema_parent_op ON child_schema(parent_op)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_cross_ref_entry_id ON cross_reference(entry_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_cross_ref_dp_ref ON cross_reference(dp_ref)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_umd_operations_op_id ON umd_operations(op_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_dp_register_dp_code ON dp_register(dp_code)")
    
    conn.commit()
    
    # Print statistics
    print(f"\n=== Import Summary ===")
    print(f"Total rows imported: {total_rows}")
    print(f"Cross-reference entries: {cross_ref_count}")
    
    # Verify data counts
    for table in ["umd_operations", "child_schema", "dp_register", "att_terms", "phonetic_reversal", "session_index", "cross_reference"]:
        cursor.execute(f"SELECT COUNT(*) FROM {table}")
        count = cursor.fetchone()[0]
        print(f"{table}: {count} rows")
    
    wb.close()
    conn.close()
    print(f"\nDatabase created successfully: {db_path}")

if __name__ == "__main__":
    main()