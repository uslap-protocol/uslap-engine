#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
USLaP Autonomous Engine v1.0
Unified Source Language Proof — Linguistic Intelligence Engine

Processes English words / Arabic roots / ratios / phrases through the QUF pipeline.
Produces: (A) Lattice placement in master file, (B) 360-degree HTML report.

Architecture (8 components):
  1. InputRouter       — detects input type, routes to pipeline
  2. PhoneticReversal  — English consonants → ORIG root candidates
  3. QGate             — Qur'anic attestation (binary PASS/FAIL)
  4. UGate             — Phonetic unity verification
  5. FGate             — Foundation layer (DS/DP/network assignment)
  6. ClusterExpander   — root → all English words sharing that root
  7. EntryWriter       — writes to A1_ENTRIES + A4/A5/M4/SESSION_INDEX/ENGINE_QUEUE
  8. ReportGenerator   — 360-degree HTML report from all domains

بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ
"""

import sys
import os
import re
import json
import shutil
import itertools
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict

import openpyxl
from openpyxl import load_workbook

# ─── FILE PATHS ───────────────────────────────────────────────────────────────
MASTER_FILE   = "/Users/mmsetubal/Documents/USLaP workplace/USLaP_Final_Data_Consolidated_Master_v3.xlsx"
QURAN_FILE    = "/Users/mmsetubal/Documents/USLaP Master Folder/Linguistic /USLaP_Quran_Root_Count.xlsx"
REPORTS_DIR   = "/Users/mmsetubal/Documents/USLaP workspace/Reports"
WORKSPACE_DIR = "/Users/mmsetubal/Documents/USLaP workplace"
KASHGARI_FILE = "/Users/mmsetubal/Documents/USLaP Master Folder/Linguistic /Kashgari 1.2.3.txt"

# ─── THRESHOLDS ───────────────────────────────────────────────────────────────
SCORE_AUTO_WRITE  = 9   # score >= 9  → queue as auto-write candidate (CONFIRMED_HIGH)
SCORE_QUEUE       = 7   # score 7-8   → queue for oversight (PENDING_REVIEW)
SCORE_REJECT      = 6   # score <= 6  → reject (log only)
# v2.1: CONFIRMED_HIGH target = 15-25% of batch. If > 33% → scoring inflated.
# Tightening: require Q+U+positional all pass (score>=9 alone is no longer enough)
MAX_CLUSTER_DEPTH = 3   # max recursion depth in ClusterExpander

# ─── v3.4: MODERN TERMINOLOGY (EN→RU direction — exception to RU>EN rule) ────
# These words entered Russian FROM English (modern tech/medicine/finance).
# For these, EN cognate IS authoritative. For all others, RU is closer to
# both originals and EN cognate is confirmatory only.
MODERN_TECH_TERMS = {
    'КОМПЬЮТЕР', 'ИНТЕРНЕТ', 'ТЕЛЕФОН', 'ТЕЛЕВИЗОР', 'ПРИНТЕР', 'СЕРВЕР',
    'ПРОЦЕССОР', 'МОНИТОР', 'ДИСПЛЕЙ', 'МОДЕМ', 'РОУТЕР', 'БРАУЗЕР',
    'МЕНЕДЖЕР', 'МАРКЕТИНГ', 'БИЗНЕС', 'ОФИС', 'ДИЗАЙН', 'БРЕНД',
    'ИНВЕСТОР', 'ДИЛЕР', 'БРОКЕР', 'ЛИЗИНГ', 'ФАКС', 'ИМЕЙЛ',
    'ПЕНИЦИЛЛИН', 'ИНСУЛИН', 'АНТИБИОТИК', 'ВАКЦИНА', 'ЛАЗЕР',
    'РАДАР', 'ПЛАСТИК', 'НЕЙЛОН', 'ТЕФЛОН', 'СИЛИКОН',
}

# ─── SUFFIX LIST (OP_SUFFIX stripping — longest first) ───────────────────────
LATIN_SUFFIXES = sorted([
    'ation', 'ition', 'ment', 'ness', 'ance', 'ence', 'ancy', 'ency',
    'ical', 'ary', 'ory', 'ery', 'ity', 'ous', 'ious', 'ion', 'ism',
    'ist', 'ize', 'ise', 'ify', 'ship', 'hood', 'ward', 'wise', 'ic',
    'ant',          # OP_SUFFIX: Latin -ant (COVENANT→COVEN, PLEASANT→PLEAS, SERVANT→SERV)
                    # Note: -ent NOT added — strips too aggressively (ANCIENT, MOMENT broken)
    'al', 'fy', 'ed', 'ing', 'ive', 'ly', 'er', 'or', 'ar', 'es', 'e', 's'
], key=len, reverse=True)

# ─── FUNCTION WORDS (for phrase parsing) ─────────────────────────────────────
FUNCTION_WORDS = {
    'the','a','an','and','or','but','in','on','at','to','for','of',
    'with','by','from','is','are','was','were','be','been','it','its',
    'this','that','these','those','as','into','not','no','if','then'
}

# ─── KNOWN مَفْعَل PATTERNS (Gate 3e: M-prefix place noun skeletons) ──────────
# When a word starts with M and the remaining consonants match a known مَفْعَل
# skeleton, the مَفْعَل candidate is boosted (same logic as N15 priority for
# C/G/K-R-N).  Each entry: (remaining_consonants_pattern, forced_root, label).
# Sources: 8 existing lattice entries confirmed as مَفْعَل + new patterns.
MAFAL_SKELETONS = {
    # مَرْكَز markaz (center/market) → ر-ك-ز — MARKET #249
    'rkt': 'ر-ك-ز',  'rkz': 'ر-ك-ز',  'rks': 'ر-ك-ز',
    # مَخْزَن makhzan (storehouse/magazine) → خ-ز-ن — MAGAZINE #13
    'khzn': 'خ-ز-ن',  'gzn': 'خ-ز-ن',  'kzn': 'خ-ز-ن',
    # مَسْجِد masjid (mosque) → س-ج-د — MOSQUE #20
    'sjd': 'س-ج-د',  'zgd': 'س-ج-د',
    # مَنْزِل manzil (station/house) → ن-ز-ل — from MINARET #21 family
    'nzl': 'ن-ز-ل',
    # مَدْرَسَة madrasa (school) → د-ر-س — MADRASA #19
    'drs': 'د-ر-س',
    # مَطْرَح maṭraḥ (place of throwing) → ط-ر-ح — MATTRESS #54
    'trh': 'ط-ر-ح',  'trs': 'ط-ر-ح',
    # مَقْبَرَة maqbara (graveyard) → ق-ب-ر — MACABRE #82
    'qbr': 'ق-ب-ر',  'kbr': 'ق-ب-ر',  'cbr': 'ق-ب-ر',
    # مِرْآة mirʾāh (mirror) → ر-أ-ي — MIRROR #130
    'rr': 'ر-أ-ي',
    # مَنَارَة manāra (lighthouse) → ن-و-ر — MINARET #21
    'nrt': 'ن-و-ر',  'nr': 'ن-و-ر',
    # مَوْصِل Mawṣil (junction) → و-ص-ل — MUSLIN #77
    'sl': 'و-ص-ل',  'zl': 'و-ص-ل',
    # مِنْهَاج minhāj (methodology) → ن-ه-ج — MANAGER #143
    'nhj': 'ن-ه-ج',  'ngr': 'ن-ه-ج',
}

# ─── DERIVATIVE CHAINS (Gate 3f: known parent→child word families) ──────────
# When a word is a known derivative of a confirmed lattice entry, route it to
# A4_DERIVATIVES instead of creating a new A1 entry.  Maps EN_TERM → parent.
# Built from: A4_DERIVATIVES (582 entries) + batch root families.
KNOWN_DERIVATIVES = {
    # MARKET family (R211 ر-ك-ز)
    'merchant': 'MARKET', 'merchandise': 'MARKET', 'mercantile': 'MARKET',
    'mercenary': 'MARKET', 'mercy': 'MARKET', 'commerce': 'MARKET',
    'commercial': 'MARKET', 'marketing': 'MARKET', 'marketplace': 'MARKET',
    'supermarket': 'MARKET',
    # HORN/N15 family (R133 ق-ر-ن) — derivatives of existing entries
    'grain': 'HORN', 'corn': 'HORN', 'caravan': 'HORN', 'cornet': 'HORN',
    'corona': 'HORN', 'coronation': 'HORN', 'cornea': 'HORN',
    # GOVERN family (R08 ج-ب-ر)
    'governor': 'GOVERN', 'government': 'GOVERN', 'governance': 'GOVERN',
    # EMPIRE family (R01 أ-م-ر)
    'emperor': 'EMPIRE', 'empress': 'EMPIRE', 'imperial': 'EMPIRE',
    # ALGEBRA family (R08 ج-ب-ر)
    'algebraic': 'ALGEBRA', 'algebraist': 'ALGEBRA',
    # COFFEE family (R168 ق-ه-ر)
    'cafe': 'COFFEE', 'cafeteria': 'COFFEE', 'caffeine': 'COFFEE',
    # SULTAN family (R07 س-ل-ط)
    'sultanate': 'SULTAN',
    # ALCOHOL family (R27 ك-ح-ل)
    'alcoholic': 'ALCOHOL', 'alcoholism': 'ALCOHOL',
    # SUGAR family (R53 ش-ك-ر)
    'sugary': 'SUGAR', 'sugarcane': 'SUGAR',
    # CRIME family (R10 ح-ر-م)
    'criminal': 'CRIME', 'criminology': 'CRIME',
    # COTTON family
    'cottonseed': 'COTTON', 'cottontail': 'COTTON',
    # CALIBRE family (R31 ق-ل-ب)
    'calibrate': 'CALIBRE', 'calibration': 'CALIBRE',
    # PATTERN family (R85 ف-ط-ر)
    'patterning': 'PATTERN', 'patterned': 'PATTERN',
    # SACRIFICE family (R200 ش-ك-ر)
    'sacrificial': 'SACRIFICE', 'sacrificing': 'SACRIFICE',
    # REVOLUTION family (R195 ب-ل-و)
    'revolutionary': 'REVOLUTION', 'revolt': 'REVOLUTION', 'revolve': 'REVOLUTION',
    'rebellion': 'REVOLUTION', 'rebel': 'REVOLUTION',
    # MEDICINE family (R160 م-ي-د)
    'medical': 'MEDICINE', 'medic': 'MEDICINE', 'medication': 'MEDICINE',
    # MILITARY family (R02 م-ل-ك)
    'militia': 'MILITARY', 'militant': 'MILITARY',
    # MORTAL family (R103 م-و-ت)
    'mortality': 'MORTAL', 'immortal': 'MORTAL',
    # SORCERY family (R10 ح-ر-م)
    'sorcerer': 'SORCERY', 'sorceress': 'SORCERY',
    # PHILOSOPHY family
    'philosopher': 'PHILOSOPHY', 'philosophical': 'PHILOSOPHY',
    # PROTOCOL family
    'protocolar': 'PROTOCOL',
}

# ─── COGNATE CROSSREF (v3.3: Russian↔English sibling lookup) ──────────────────
# Maps Russian words to their known English cognates.  When the engine processes
# a Russian word that has an English cousin, it also processes the ENGLISH form
# through the English PhoneticReversal pipeline and compares results.
# Rationale: English preserves root consonants that French→Russian corridor loses.
# Example: ДЕСАНТ lost the К from سَكَنَ (S-K-N-D→DESCENT), but Russian only has
# Д-С-Н-Т.  The English pipeline recovers the root; the Russian pipeline cannot.
# Format: { 'russian_lower': 'ENGLISH_UPPER', ... }
# Bidirectional: the reverse mapping is generated automatically.
COGNATE_CROSSREF_RU_TO_EN = {
    # ── MILITARY + WARFARE ──────────────────────────────────────────────────
    'десант':     'DESCENT',
    'бастион':    'BASTION',
    'батарея':    'BATTERY',
    'гарнизон':   'GARRISON',
    'мушкет':     'MUSKET',
    'арсенал':    'ARSENAL',
    'маршал':     'MARSHAL',
    'генерал':    'GENERAL',
    'адмирал':    'ADMIRAL',
    'артиллерия': 'ARTILLERY',
    'барьер':     'BARRIER',
    # ── TRADE + ECONOMY ─────────────────────────────────────────────────────
    'банк':       'BANK',
    'тариф':      'TARIFF',
    'караван':    'CARAVAN',
    'магазин':    'MAGAZINE',
    'талант':     'TALENT',
    'базар':      'BAZAAR',
    'баланс':     'BALANCE',
    # ── GOVERNANCE + LAW ────────────────────────────────────────────────────
    'султан':     'SULTAN',
    'эмир':       'EMIR',
    'грамота':    'GRAMMAR',
    # ── RELIGION + FAITH ────────────────────────────────────────────────────
    'минарет':    'MINARET',
    'масджид':    'MOSQUE',
    'джихад':     'JIHAD',
    'намаз':      'NAMAZ',
    'халиф':      'CALIPH',
    'муфтий':     'MUFTI',
    'шариат':     'SHARIAT',
    # ── SCIENCE + CRAFT ─────────────────────────────────────────────────────
    'алхимия':    'ALCHEMY',
    'алгебра':    'ALGEBRA',
    'зенит':      'ZENITH',
    'азимут':     'AZIMUTH',
    'алкоголь':   'ALCOHOL',
    'эликсир':    'ELIXIR',
    'бальзам':    'BALSAM',
    'химия':      'CHEMISTRY',
    # ── FOOD + DRINK ────────────────────────────────────────────────────────
    'кофе':       'COFFEE',
    'сахар':      'SUGAR',
    'лимон':      'LEMON',
    'шафран':     'SAFFRON',
    'йогурт':     'YOGURT',
    # ── TEXTILES + HOUSEHOLD ────────────────────────────────────────────────
    'хлопок':     'COTTON',
    'матрас':     'MATTRESS',
    'диван':      'DIVAN',
    'лак':        'LACQUER',
    # ── NATURE + GEOGRAPHY ──────────────────────────────────────────────────
    'муссон':     'MONSOON',
    'тундра':     'TUNDRA',
    # ── BODY + HEALTH ───────────────────────────────────────────────────────
    'массаж':     'MASSAGE',
    # ── BORROWED INTERNATIONAL (Latin/Greek corridor shared) ────────────────
    'крепость':   'FORTRESS',
    'пошлина':    'CUSTOMS',
    'рубль':      'RUBLE',
    'самовар':    'SAMOVAR',     # no English cousin — but included for completeness
    'кинжал':     'DAGGER',      # loose cognate — different corridors
    'шахта':      'SHAFT',
    'табурет':    'TABOURET',
    'лакировка':  'LACQUER',
}

# Auto-build reverse map: EN → RU
COGNATE_CROSSREF_EN_TO_RU = {}
for _ru, _en in COGNATE_CROSSREF_RU_TO_EN.items():
    COGNATE_CROSSREF_EN_TO_RU.setdefault(_en, []).append(_ru)

# ─── BANNED TERMS (auto-scan before write) ───────────────────────────────────
BANNED_TERMS = [
    'semitic', 'loanword', 'loan word', 'borrowed from', 'cognate',
    'proto-indo-european', 'proto indo european', 'pie root',
    'prosthetic vowel', 'pre-greek substrate', 'adoption',
    # Wrapper terms (CLAUDE.md §7 + Turkish≠Turkic rule)
    'islamic origin', 'islamic civilization', 'islamic science',
    'turkish origin', 'turkish language', 'from turkish', 'old turkish',
]


# ═══════════════════════════════════════════════════════════════════════════════
# DATA CLASSES
# ═══════════════════════════════════════════════════════════════════════════════

class GateResult:
    """Result of a QUF gate check — binary PASS or FAIL."""
    def __init__(self, passed: bool, details: dict = None):
        self.passed  = passed
        self.details = details or {}

    def __bool__(self):
        return self.passed

    def __repr__(self):
        return f"GateResult({'PASS' if self.passed else 'FAIL'}, {self.details})"


class RootCandidate:
    """A candidate ORIG1 Arabic root for a given English word."""
    def __init__(self, letters: str, token_count: int = 0, lemma_count: int = 0,
                 ar_word: str = '', operations: list = None):
        self.letters      = letters        # e.g. "ق-ر-ن"
        self.token_count  = token_count
        self.lemma_count  = lemma_count
        self.ar_word      = ar_word        # e.g. "قَرَن"
        self.operations   = operations or []
        self.phonetic_chain = ''
        self.score        = 0
        self.positional_score   = 0.5   # R11: consonant ORDER fidelity (0.0–1.0); 0.5 = neutral
        self.transposition_flag = False  # R11: True = consonant ORDER inverted vs root order
        self._n15_priority      = False  # R09: True = N15 skeleton forced this candidate first
        self.extra_consonants   = 0      # Coverage: word consonants NOT covered by root

    def __repr__(self):
        return f"RootCandidate({self.letters}, tokens={self.token_count}, score={self.score})"


class EntryRecord:
    """A full 14-column lattice entry ready for writing to A1_ENTRIES."""
    def __init__(self):
        self.entry_id      : int = 0
        self.score         : int = 0
        self.en_term       : str = ''
        self.ar_word       : str = ''
        self.root_id       : str = ''
        self.root_letters  : str = ''
        self.qur_meaning   : str = ''
        self.pattern       : str = 'A'
        self.allah_name_id : str = ''
        self.network_id    : str = ''
        self.phonetic_chain: str = ''
        self.inversion_type: str = 'HIDDEN'
        self.source_form   : str = ''
        self.foundation_ref: str = ''

    def to_row(self) -> tuple:
        """Return as 14-tuple matching A1_ENTRIES column order."""
        return (
            self.entry_id, self.score, self.en_term, self.ar_word,
            self.root_id, self.root_letters, self.qur_meaning, self.pattern,
            self.allah_name_id, self.network_id, self.phonetic_chain,
            self.inversion_type, self.source_form, self.foundation_ref
        )


class ProcessResult:
    """Full result of processing one input term through the engine."""
    def __init__(self, input_term: str, input_type: str):
        self.input_term        = input_term
        self.input_type        = input_type
        self.existing_entry_id : Optional[int]       = None
        self.root_candidates   : List[RootCandidate] = []
        self.confirmed_root    : Optional[RootCandidate] = None
        self.q_gate            : Optional[GateResult]    = None
        self.u_gate            : Optional[GateResult]    = None
        self.f_gate            : Optional[GateResult]    = None
        self.entry_record      : Optional[EntryRecord]   = None
        self.cluster_members   : list = []
        self.queue_id          : Optional[str] = None
        self.report_path       : Optional[str] = None
        self.derivative_of     : Optional[tuple] = None   # (parent_name, parent_id) if derivative chain detected
        self.orig2_track       : bool = False              # True if routed through ORIG2/Kashgari track
        self.orig2_details     : Optional[dict] = None     # Kashgari attestation details
        self.log               : list = []

    def add_log(self, msg: str):
        ts = datetime.now().strftime('%H:%M:%S')
        self.log.append(f"[{ts}] {msg}")
        print(f"  {msg}")


# ═══════════════════════════════════════════════════════════════════════════════
# COMPONENT 1 — InputRouter
# ═══════════════════════════════════════════════════════════════════════════════

class InputRouter:
    """Detects input type and routes to appropriate processing pipeline."""

    ARABIC_CHARS = set(
        'ابتثجحخدذرزسشصضطظعغفقكلمنهوي'
        'أإآءةىًٌٍَُِّْ'
    )

    CYRILLIC_CHARS = set(
        'абвгдежзийклмнопрстуфхцчшщъыьэюя'
        'АБВГДЕЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ'
        'ёЁ'
    )

    def detect(self, raw: str) -> tuple:
        """
        Returns: (input_type, cleaned, key_terms)
          input_type: 'english_word' | 'russian_word' | 'arabic_root' | 'ratio'
                    | 'quran_ref' | 'phrase' | 'russian_phrase'
          key_terms:  list of individual terms to process
        """
        s = raw.strip()

        # Qur'anic reference: Q18:83
        if re.match(r'^Q\d+:\d+$', s, re.IGNORECASE):
            return ('quran_ref', s.upper(), [s.upper()])

        # Ratio: 4/3, 19/7
        if re.match(r'^\d+/\d+$', s):
            return ('ratio', s, [s])

        # Arabic root (contains Arabic characters)
        arabic_count = sum(1 for c in s if c in self.ARABIC_CHARS)
        if arabic_count >= 2:
            cleaned = re.sub(r'[\s\-—–]+', '-', s).strip('-')
            return ('arabic_root', cleaned, [cleaned])

        # Russian word (contains Cyrillic characters)
        cyrillic_count = sum(1 for c in s if c in self.CYRILLIC_CHARS)
        if cyrillic_count >= 2:
            words = s.split()
            if len(words) > 1:
                content = [w for w in words if any(ch in self.CYRILLIC_CHARS for ch in w)]
                if len(content) > 1:
                    return ('russian_phrase', s, content)
                if content:
                    return ('russian_word', content[0], [content[0]])
            clean = re.sub(r'[^а-яА-ЯёЁ\-]', '', s)
            return ('russian_word', clean, [clean])

        # Phrase (multiple words)
        words = s.split()
        if len(words) > 1:
            content = [w for w in words if w.lower() not in FUNCTION_WORDS and w.isalpha()]
            if len(content) > 1:
                return ('phrase', s, content)
            if content:
                return ('english_word', content[0], [content[0]])

        # Single English word
        clean = re.sub(r'[^a-zA-Z\-]', '', s)
        return ('english_word', clean, [clean])


# ═══════════════════════════════════════════════════════════════════════════════
# COMPONENT 3 — QGate  (loaded first; PhoneticReversal depends on it)
# ═══════════════════════════════════════════════════════════════════════════════

class QGate:
    """Binary PASS/FAIL against 1,681 Qur'anic roots in ROOT_LIST."""

    def __init__(self, quran_root_file: str):
        self.roots: Dict[str, dict] = {}
        self._load_roots(quran_root_file)
        print(f"  QGate: {len(self.roots)} Qur'anic roots loaded")

    def _load_roots(self, filepath: str):
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb['ROOT_LIST']
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not any(row):
                    continue
                d = dict(zip(headers, row))
                # ROOT_LIST: 'Root' column = bare Arabic string (no hyphens)
                # 'Letters' column = integer count of letters (3 or 4)
                bare_root = str(d.get('Root', '') or '').strip()
                if bare_root and len(bare_root) >= 2:
                    try:
                        tc = int(d.get('Token Count', 0) or 0)
                    except (ValueError, TypeError):
                        tc = 0
                    try:
                        lc = int(d.get('Lemma Count', 0) or 0)
                    except (ValueError, TypeError):
                        lc = 0
                    self.roots[bare_root] = {
                        'token_count': tc,
                        'lemma_count': lc,
                        'ar_word'    : bare_root,
                    }
            wb.close()
        except Exception as e:
            print(f"  QGate load error: {e}")

    def check(self, root_letters: str) -> GateResult:
        normalized = re.sub(r'[\s\-—–]+', '-', root_letters.strip()).strip('-')
        # ROOT_LIST stores bare strings without hyphens — strip them for lookup
        bare_lookup = re.sub(r'[\-\s]', '', normalized)
        bare_lookup = re.sub(r'[ًٌٍَُِّْ]', '', bare_lookup)

        if bare_lookup in self.roots:
            d = self.roots[bare_lookup]
            return GateResult(True, {**d, 'root_letters': normalized})
        # Also try with harakat stripped from normalized (hyphenated) form
        bare_hyph = re.sub(r'[ًٌٍَُِّْ]', '', normalized)
        if bare_hyph in self.roots:
            d = self.roots[bare_hyph]
            return GateResult(True, {**d, 'root_letters': bare_hyph})
        return GateResult(False, {
            'root_letters'   : normalized,
            'reason'         : 'Not in 1,681 Qur\'anic roots — flagged ORIG2 candidate for human review',
            'orig2_candidate': True,   # two-track gate: human must check Kashgari corpus next
        })


# ═══════════════════════════════════════════════════════════════════════════════
# COMPONENT 2b — KashgariIndex + KashgariGate (ORIG2 track)
# ═══════════════════════════════════════════════════════════════════════════════

class KashgariIndex:
    """
    Parses and indexes the Kashgari Dīwān corpus (Dankoff & Kelly, Harvard 1982-1985)
    for ORIG2 attestation.  74K-line OCR'd plain text → searchable by consonant skeleton.

    Three search modes:
      1. skeleton_match: consonant skeleton of English word matches Kashgari entry
      2. translit_match: direct transliteration lookup
      3. meaning_match: English word found in Kashgari definition/gloss
    """

    # Turkic vowels (broader set than English — includes ü, ö, ı, etc.)
    TURKIC_VOWELS = set('aeiouüöıäəāēīōūAEIOUÜÖ')

    def __init__(self, corpus_file: str):
        self.entries: Dict[str, list] = {}         # translit → [entry_dicts]
        self.skeleton_index: Dict[str, list] = {}  # consonant_skeleton → [entry_dicts]
        self._parse(corpus_file)
        self._build_skeleton_index()

    @staticmethod
    def extract_consonants(translit: str) -> str:
        """Strip vowels from transliteration to get consonant skeleton."""
        vowels = KashgariIndex.TURKIC_VOWELS
        return ''.join(c for c in translit.lower()
                       if c.isalpha() and c not in vowels)

    def _parse(self, filepath: str):
        """Parse Kashgari corpus line by line, extracting dictionary entries."""
        if not os.path.exists(filepath):
            print(f"  KashgariIndex: corpus file not found: {filepath}")
            return

        # Pattern: HEADWORD (CAPS, possibly with ' or - or special chars)
        #          + transliteration (mixed case — OCR uses uppercase for č→C, ş→S, etc.)
        #          + definition (in quotes or following text)
        entry_re = re.compile(
            r"""^['"]?                             # optional leading quote
                ([A-Z][A-Z0-9'\-§_\^]+)           # headword in CAPS
                \s+
                ([a-zA-ZüöıçşğÜÖ][a-zA-Z\-üöıçşğÜÖ]*)  # transliteration (mixed case for OCR)
                \s+
                [""\"]?                            # optional opening quote
                (.+?)                              # definition text
                [""\"]?\s*                         # optional closing quote
                (?:0\s*)?$                         # optional entry-end marker '0'
            """, re.VERBOSE)

        count = 0
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            for line_no, line in enumerate(f, 1):
                if line_no < 6300:  # skip header/intro
                    continue
                stripped = line.strip()
                if not stripped or len(stripped) < 5:
                    continue
                # Skip page refs, footnotes, numerals-only lines
                if re.match(r'^\[\w', stripped) or re.match(r'^\d+\.?\s*$', stripped):
                    continue

                m = entry_re.match(stripped)
                if m:
                    headword = m.group(1).strip('-')
                    translit = m.group(2).lower().strip('-')  # normalize to lowercase
                    meaning  = m.group(3).strip(' ."\'')

                    # Skip very short or noise entries
                    if len(translit) < 1 or len(meaning) < 3:
                        continue

                    # Extract Arabic gloss in parentheses
                    ar_m = re.search(r'\(([^)]+)\)', meaning)
                    arabic_gloss = ar_m.group(1) if ar_m else ''

                    entry = {
                        'headword':     headword,
                        'translit':     translit,
                        'meaning':      meaning,
                        'arabic_gloss': arabic_gloss,
                        'line':         line_no,
                    }

                    self.entries.setdefault(translit, []).append(entry)
                    count += 1

        print(f"  KashgariIndex: parsed {count} entries from corpus "
              f"({len(self.entries)} unique transliterations)")

    def _build_skeleton_index(self):
        """Build consonant skeleton → entries index for fast lookup."""
        for translit, entries_list in self.entries.items():
            skel = self.extract_consonants(translit)
            if skel and len(skel) >= 1:
                for e in entries_list:
                    rec = {**e, 'skeleton': skel}
                    self.skeleton_index.setdefault(skel, []).append(rec)
        print(f"  KashgariIndex: {len(self.skeleton_index)} unique consonant skeletons indexed")

    def search_skeleton(self, consonants: str) -> list:
        """Search by consonant skeleton (e.g., 'blq' → balıq = city)."""
        return self.skeleton_index.get(consonants.lower(), [])

    def search_translit(self, term: str) -> list:
        """Search by exact transliteration."""
        return self.entries.get(term.lower(), [])

    def search_english(self, english_word: str) -> list:
        """Search all definitions for an English word (whole-word match only)."""
        results = []
        # Require whole-word match to avoid false positives
        # e.g., "head" should NOT match "thread" or "heading"
        pattern = re.compile(r'\b' + re.escape(english_word.lower()) + r'\b')
        for translit, entries_list in self.entries.items():
            for e in entries_list:
                if pattern.search(e['meaning'].lower()):
                    results.append(e)
        return results


class KashgariGate:
    """
    ORIG2 attestation gate — checks Kashgari corpus for Turkic roots.
    Fires when Q-Gate (ORIG1) FAILS.  Implements B01-B07 phonology checks.

    Protocol (from CLAUDE.md §6 two-track gate):
      ROOT_LIST FAIL → Kashgari search → if attested → ORIG2 entry → BITIG_A1_ENTRIES
      If BOTH fail → entry cannot exceed score 7.
    """

    # ── KNOWN ORIG2 ENTRIES ─────────────────────────────────────────────────
    # v3.2: Manually verified entries from Kashgari corpus research that
    # the KashgariIndex parser misses (thematic sections, OCR issues).
    # Keyed by consonant skeleton → list of entry dicts.
    # Source: verified with page/line citations from Dankoff & Kelly.
    KNOWN_ORIG2_ENTRIES = {
        'çp': [{'translit': 'çap-', 'meaning': 'beat, strike (neck), swim [ḍaraba]',
                'line': 57461, 'headword': 'ÇAP-', 'skeleton': 'çp'}],
        'cp': [{'translit': 'çap-', 'meaning': 'beat, strike (neck), swim [ḍaraba]',
                'line': 57461, 'headword': 'ÇAP-', 'skeleton': 'cp'}],
        'sp': [{'translit': 'sap-', 'meaning': 'thread (needle), bind, mend',
                'line': 61433, 'headword': 'SAP-', 'skeleton': 'sp'}],
        'sb': [{'translit': 'sap-', 'meaning': 'thread (needle), bind, mend [OP_VOICE p→b]',
                'line': 61433, 'headword': 'SAP-', 'skeleton': 'sb'},
               {'translit': 'çap-', 'meaning': 'beat, strike [OP_VOICE ç→s, p→b]',
                'line': 57461, 'headword': 'ÇAP-', 'skeleton': 'sb'}],
        'qlc': [{'translit': 'qılıç', 'meaning': 'sword (sayf)',
                 'line': 19685, 'headword': 'QILIÇ', 'skeleton': 'qlc'}],
        'qlç': [{'translit': 'qılıç', 'meaning': 'sword (sayf)',
                 'line': 19685, 'headword': 'QILIÇ', 'skeleton': 'qlç'}],
        'krt': [{'translit': 'kirit', 'meaning': 'key (miftāḥ) — al-Kashgari: close to iqlīd',
                 'line': 19635, 'headword': 'KIRIT', 'skeleton': 'krt'}],
        'kld': [{'translit': 'kirit', 'meaning': 'key (miftāḥ) — Kashgari: iqlīd with q→k, l→r, d→t',
                 'line': 19635, 'headword': 'KIRIT', 'skeleton': 'kld'}],
        'bl':  [{'translit': 'böl-', 'meaning': 'divide into groups',
                 'line': 57002, 'headword': 'BÖL-', 'skeleton': 'bl'}],
        'tn':  [{'translit': 'ton-', 'meaning': 'freeze, become ice',
                 'line': 19820, 'headword': 'TON-', 'skeleton': 'tn'}],
        'tmn': [{'translit': 'tuman', 'meaning': 'fog, mist (10,000)',
                 'line': 19780, 'headword': 'TUMAN', 'skeleton': 'tmn'}],
        # v3.2: English ch = Turkic ç (same sound, different notation)
        'chp': [{'translit': 'çap-', 'meaning': 'beat, strike (neck), swim [ḍaraba]',
                 'line': 57461, 'headword': 'ÇAP-', 'skeleton': 'chp'}],
        'chb': [{'translit': 'çap-', 'meaning': 'beat, strike [OP_VOICE p→b]',
                 'line': 57461, 'headword': 'ÇAP-', 'skeleton': 'chb'}],
    }

    def __init__(self, kashgari_index: KashgariIndex):
        self.index = kashgari_index

    def check(self, en_word: str, consonants: list) -> GateResult:
        """
        ORIG2 attestation check.
        Args:
            en_word:    the English word being processed
            consonants: extracted consonant list from PhoneticReversal
        Returns:
            GateResult with Kashgari attestation details if found
        """
        skel = ''.join(consonants).lower()

        # 1. Direct skeleton search (parsed index)
        hits = self.index.search_skeleton(skel)

        # 1b. v3.2: Check KNOWN_ORIG2_ENTRIES (manually verified, parser-missed)
        if not hits:
            hits = list(self.KNOWN_ORIG2_ENTRIES.get(skel, []))

        # 2. Try Bitig consonant variants (q↔k↔g, p↔b, etc.)
        if not hits:
            for v in self._bitig_variants(skel):
                v_hits = self.index.search_skeleton(v)
                if not v_hits:
                    v_hits = list(self.KNOWN_ORIG2_ENTRIES.get(v, []))
                if v_hits:
                    hits.extend(v_hits)
                    break  # first successful variant is enough

        # 3. Try shorter skeletons (strip suffixes — B03 agglutinative)
        if not hits and len(skel) >= 3:
            for trim in range(1, min(3, len(skel) - 1)):
                trimmed = skel[:-trim]
                t_hits = self.index.search_skeleton(trimmed)
                if not t_hits:
                    t_hits = list(self.KNOWN_ORIG2_ENTRIES.get(trimmed, []))
                # Also check voicing variants of trimmed skeleton
                if not t_hits:
                    for v in self._bitig_variants(trimmed):
                        v_hits = self.index.search_skeleton(v)
                        if not v_hits:
                            v_hits = list(self.KNOWN_ORIG2_ENTRIES.get(v, []))
                        if v_hits:
                            t_hits.extend(v_hits)
                            break
                if t_hits:
                    hits.extend(t_hits)
                    break

        # 4. Meaning search fallback
        meaning_hits = []
        if not hits:
            meaning_hits = self.index.search_english(en_word)

        # Build result
        if hits:
            best = hits[0]
            warnings = self._phonology_checks(best.get('translit', ''), en_word)
            return GateResult(True, {
                'kashgari_translit': best['translit'],
                'kashgari_meaning':  best['meaning'],
                'kashgari_line':     best['line'],
                'kashgari_headword': best['headword'],
                'skeleton':          best.get('skeleton', skel),
                'all_hits':          len(hits),
                'bitig_warnings':    warnings,
                'attestation_type':  'skeleton_match',
            })

        if meaning_hits:
            best = meaning_hits[0]
            warnings = self._phonology_checks(best.get('translit', ''), en_word)
            return GateResult(True, {
                'kashgari_translit': best['translit'],
                'kashgari_meaning':  best['meaning'],
                'kashgari_line':     best['line'],
                'kashgari_headword': best['headword'],
                'skeleton':          KashgariIndex.extract_consonants(best['translit']),
                'all_hits':          len(meaning_hits),
                'bitig_warnings':    warnings,
                'attestation_type':  'meaning_match',
            })

        return GateResult(False, {
            'reason':           f"Not in Kashgari corpus (skeleton '{skel}' unattested)",
            'skeleton_searched': skel,
        })

    def _bitig_variants(self, skeleton: str) -> list:
        """Generate Bitig consonant equivalences for broader search (B01-B05 informed)."""
        equivs = {
            'p': ['b'],       # B01: /f/→/p/→/b/ in Bitig
            'b': ['p'],
            'k': ['q', 'g'],  # velar variants
            'q': ['k', 'g'],
            'g': ['k', 'q'],
            'c': ['s', 'z', 'j'],  # sibilant variants
            's': ['z', 'c'],
            'z': ['s'],
            't': ['d'],
            'd': ['t'],
            'j': ['c'],
            'f': ['p', 'b'],  # B01: /f/ non-native, closest Bitig equivalents
        }
        variants = set()
        for i, c in enumerate(skeleton):
            for alt in equivs.get(c, []):
                v = skeleton[:i] + alt + skeleton[i+1:]
                if v != skeleton:
                    variants.add(v)
        return list(variants)[:12]

    def _phonology_checks(self, translit: str, en_word: str) -> list:
        """Run B01-B07 automated phonology checks."""
        warnings = []
        tl = translit.lower()

        # B01: No /f/ phoneme in Bitig
        if 'f' in tl:
            warnings.append("B01: /f/ in Turkic form — foreign contamination flag")

        # B05: No /w/ phoneme in old Bitig
        if 'w' in tl:
            warnings.append("B05: /w/ detected — possible ORIG1, not ORIG2")

        # B03: Agglutinative morphology — flag long forms for suffix stripping
        turkic_suffixes = ['lar', 'ler', 'liq', 'lik', 'chi', 'ci',
                           'mak', 'mek', 'gan', 'gen', 'diq', 'dik']
        for sfx in turkic_suffixes:
            if tl.endswith(sfx) and len(tl) > len(sfx) + 2:
                warnings.append(f"B03: suffix -{sfx} detected — strip before root trace")
                break

        return warnings


# ═══════════════════════════════════════════════════════════════════════════════
# COMPONENT 2 — PhoneticReversal Engine
# ═══════════════════════════════════════════════════════════════════════════════

class PhoneticReversal:
    """
    The core missing engine: English word → ranked ORIG1 root candidates.
    Works backward through M1_PHONETIC_SHIFTS.
    """

    def __init__(self, master_file: str, q_gate: QGate):
        self.q_gate       = q_gate
        self.shift_data   : List[dict] = []
        self.forward_map  : Dict[str, tuple] = {}   # AR_letter → (shift_id, [en_chars])
        self.reverse_map  : Dict[str, list]  = {}   # EN_char → [(AR_letter, shift_id)]
        self._load_shifts(master_file)
        self._build_reverse_map()
        print(f"  PhoneticReversal: {len(self.shift_data)} shifts, "
              f"{len(self.reverse_map)} EN patterns in reverse map")

    def _load_shifts(self, filepath: str):
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb['M1_PHONETIC_SHIFTS']
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not any(row):
                    continue
                d = dict(zip(headers, row))
                sid = str(d.get('SHIFT_ID', '') or '').strip()
                ar  = str(d.get('AR_LETTER', '') or '').strip()
                en_raw = str(d.get('EN_OUTPUTS', '') or '').strip()
                if sid and ar:
                    en_list = [x.strip().lower() for x in en_raw.split(',') if x.strip()]
                    self.shift_data.append({'shift_id': sid, 'ar_letter': ar, 'en_outputs': en_list})
                    self.forward_map[ar] = (sid, en_list)
            wb.close()
        except Exception as e:
            print(f"  PhoneticReversal load error: {e}")

    def _build_reverse_map(self):
        for shift in self.shift_data:
            ar, sid = shift['ar_letter'], shift['shift_id']
            for en in shift['en_outputs']:
                key = en.lower()
                self.reverse_map.setdefault(key, [])
                if (ar, sid) not in self.reverse_map[key]:
                    self.reverse_map[key].append((ar, sid))

    # ── string-level helpers ──────────────────────────────────────────────────

    def strip_operations(self, word: str) -> tuple:
        """
        Strip OP_SUFFIX from English word.
        Returns: (stripped_word, operations_list, suffix_removed)

        v2.4: OP_STOP removed from here — now a separate candidate generation
        path in reverse() via _generate_op_stop_variants(). This prevents
        universal ND→N from destroying CALENDAR, CYLINDER, BOUNDARY etc.
        OP_STOP is a HYPOTHESIS about geminated nasals, not a certainty.

        v2.3 FIX — Minimum-consonant guard: if suffix stripping leaves
        fewer than 3 consonants, UNDO the strip. Root letters > suffix.
        """
        w = word.lower().strip()
        ops = []
        suffix_removed = ''

        # OP_SUFFIX — with minimum-consonant guard
        w_before_suffix = w
        for suffix in LATIN_SUFFIXES:
            if w.endswith(suffix) and len(w) - len(suffix) >= 2:
                candidate = w[:-len(suffix)]
                # Count consonants in stripped form
                vowels = set('aeiou')
                cons_count = sum(1 for ch in candidate if ch.isalpha() and ch not in vowels)
                if cons_count >= 3:
                    w = candidate
                    suffix_removed = suffix
                    ops.append(f'OP_SUFFIX(-{suffix})')
                # else: stripping would leave < 3 consonants — skip this suffix
                break

        return w, ops, suffix_removed

    def extract_consonants(self, word: str) -> list:
        """Return ordered consonant skeleton (digraphs counted as one unit).

        Fix v3: Terminal-Y rule — Y at the END of a word is treated as a vowel
        (century, glory, victory, territory, democracy…).  Y at the START or
        MIDDLE of a word remains a consonant (yard, beyond, style).

        Fix v4 (v2.3): Digraph split fallback — when digraph extraction yields
        fewer than 3 consonants, re-try with digraphs split into separate letters.
        Example: FAITH → ['f','th'] (2 cons) → fallback → ['f','t','h'] (3 cons).
        This catches words where TH = ت+ح (two root letters) rather than ث (one).
        """
        result = self._extract_consonants_inner(word, use_digraphs=True)
        if len(result) < 3:
            # Try splitting digraphs — might recover hidden root consonants
            split_result = self._extract_consonants_inner(word, use_digraphs=False)
            if len(split_result) > len(result):
                return split_result
        return result

    def _extract_consonants_inner(self, word: str, use_digraphs: bool = True) -> list:
        """Inner extraction with optional digraph handling."""
        vowels = set('aeiou')
        w = word.lower()
        # Strip terminal-Y before processing (terminal Y = vowel in English)
        if w.endswith('y') and len(w) > 1 and w[-2] not in ('a', 'e', 'i', 'o', 'u'):
            w = w[:-1]   # e.g. century→centur, glory→glor, territory→territor
        DIGRAPHS = ('sh', 'ch', 'gh', 'th', 'ph', 'wh', 'qu')
        result = []
        i = 0
        while i < len(w):
            digraph = w[i:i+2] if i + 1 < len(w) else ''
            if use_digraphs and digraph in DIGRAPHS:
                result.append(digraph)
                i += 2
            elif w[i] not in vowels:
                result.append(w[i])
                i += 1
            else:
                i += 1
        return result

    def map_consonants_to_arabic(self, consonants: list) -> list:
        """Each consonant position → list of (AR_letter, shift_id) pairs."""
        mapped = []
        for c in consonants:
            candidates = self.reverse_map.get(c, [])
            if not candidates and len(c) == 2:
                candidates = self.reverse_map.get(c[0], [])
            mapped.append(candidates)
        return mapped

    def generate_root_permutations(self, mapped: list) -> list:
        """Generate 3-consonant root strings from mapped consonant candidates."""
        positions = len(mapped)
        if positions < 2:
            return []
        ar_per_pos = [list({ar for ar, sid in pos}) for pos in mapped]
        roots = set()
        n = 3 if positions >= 3 else positions
        for pos_combo in itertools.combinations(range(positions), n):
            for combo in itertools.product(*[ar_per_pos[p] for p in pos_combo]):
                if all(combo):
                    roots.add('-'.join(combo))
        return list(roots)

    def reverse(self, en_word: str) -> List[RootCandidate]:
        """
        Main public method: English word → ranked list of Qur'anic root candidates.
        Returns only candidates that pass Q-Gate.

        v2.4 — VOWEL-STRIP-FIRST ARCHITECTURE:
          PRIMARY:   raw consonant skeleton (strip vowels only) — full skeleton
          SECONDARY: suffix-stripped consonants (OP_SUFFIX applied) — Latin/Greek
          OP_STOP:   ND→N / MB→M variants on PRIMARY — separate candidate path
          N15, Gate 3e, R08a all run on PRIMARY consonants.

        This replaces v2.3's dual-path (suffix-first + raw fallback) architecture.
        The user's insight: "strip vowels first → consonants are immediately visible.
        POWER minus O and E leaves PWR. No suffix issue."

        Operations are now CANDIDATE GENERATORS, not destructive pre-processors.
        """
        # ═══ PRIMARY: raw vowel extraction (vowel-strip-first principle) ══════════
        primary_cons = self.extract_consonants(en_word)

        # ═══ SECONDARY: suffix-stripped consonants (OP_SUFFIX only) ═══════════════
        stripped, ops, suffix = self.strip_operations(en_word)
        secondary_cons = self.extract_consonants(stripped)

        # ═══ OP_STOP: generate ND→N / MB→M variants from PRIMARY ═════════════════
        op_stop_variants = self._generate_op_stop_variants(primary_cons)

        passing = []
        seen    = set()

        # ── N15 PRIORITY (R09): C/G/K-R-N skeleton → force ق-ر-ن first ──────────
        if self._check_n15_priority(primary_cons):
            n15 = 'ق-ر-ن'
            qr  = self.q_gate.check(n15)
            if qr.passed and n15 not in seen:
                seen.add(n15)
                c = RootCandidate(
                    letters     = n15,
                    token_count = qr.details.get('token_count', 0),
                    lemma_count = qr.details.get('lemma_count', 0),
                    ar_word     = qr.details.get('ar_word', ''),
                    operations  = []
                )
                c.positional_score   = 1.0    # N15 priority — forced to head
                c.transposition_flag = False
                c._n15_priority      = True   # R09: sentinel — survives sort
                c.extra_consonants   = max(0, len(primary_cons) - 3)
                passing.append(c)

        # ── GATE 3e: مَفْعَل SKELETON PRIORITY ──────────────────────────────────────
        if en_word.lower().startswith('m'):
            m_remaining = self.extract_consonants(
                self.strip_operations(en_word[1:])[0]
            )
            m_skel = ''.join(m_remaining)
            forced_root = MAFAL_SKELETONS.get(m_skel)
            if forced_root:
                mf_key = forced_root + '__MAFAL'
                qr = self.q_gate.check(forced_root)
                if qr.passed and mf_key not in seen:
                    seen.add(mf_key)
                    seen.add(forced_root)   # prevent standard path duplicate
                    mf_c = RootCandidate(
                        letters     = forced_root,
                        token_count = qr.details.get('token_count', 0),
                        lemma_count = qr.details.get('lemma_count', 0),
                        ar_word     = qr.details.get('ar_word', ''),
                        operations  = ['OP_PREFIX(مَفْعَل→m)']
                    )
                    mf_c.positional_score   = 1.0    # forced to head
                    mf_c.transposition_flag = False
                    mf_c._n15_priority      = True   # reuse N15 sentinel for sort
                    mf_c.extra_consonants   = max(0, len(m_remaining) - 3)
                    passing.append(mf_c)

        # ── PRIMARY PATH: all permutations from raw consonants ───────────────────
        mapped_primary = self.map_consonants_to_arabic(primary_cons)
        primary_roots  = self.generate_root_permutations(mapped_primary)

        for rs in primary_roots:
            if rs in seen:
                continue
            seen.add(rs)
            qr = self.q_gate.check(rs)
            if qr.passed:
                pos_s, trans = self._positional_score(primary_cons, rs)
                root_size    = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                c = RootCandidate(
                    letters     = rs,
                    token_count = qr.details.get('token_count', 0),
                    lemma_count = qr.details.get('lemma_count', 0),
                    ar_word     = qr.details.get('ar_word', ''),
                    operations  = []
                )
                c.positional_score   = pos_s
                c.transposition_flag = trans
                c.extra_consonants   = max(0, len(primary_cons) - root_size)
                passing.append(c)

        # ── SECONDARY PATH: suffix-stripped consonants (if different) ─────────────
        # v2.4 FIX: If a root was already found via PRIMARY, the SECONDARY path
        # may have FEWER extra_consonants (suffix stripped → closer to root size).
        # Update the existing candidate's extra_consonants if secondary is better.
        # This prevents the coverage penalty from over-penalizing raw-path roots.
        if secondary_cons != primary_cons and len(secondary_cons) >= 2:
            mapped_sec = self.map_consonants_to_arabic(secondary_cons)
            sec_roots  = self.generate_root_permutations(mapped_sec)
            for rs in sec_roots:
                root_size = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                sec_extra = max(0, len(secondary_cons) - root_size)
                if rs in seen:
                    # Root already found via PRIMARY — update extra_consonants
                    # if suffix-stripped path gives better coverage
                    for existing in passing:
                        if existing.letters == rs and sec_extra < existing.extra_consonants:
                            existing.extra_consonants = sec_extra
                            if ops and not existing.operations:
                                existing.operations = ops
                            break
                    continue
                seen.add(rs)
                qr = self.q_gate.check(rs)
                if qr.passed:
                    pos_s, trans = self._positional_score(secondary_cons, rs)
                    c = RootCandidate(
                        letters     = rs,
                        token_count = qr.details.get('token_count', 0),
                        lemma_count = qr.details.get('lemma_count', 0),
                        ar_word     = qr.details.get('ar_word', ''),
                        operations  = ops
                    )
                    c.positional_score   = pos_s
                    c.transposition_flag = trans
                    c.extra_consonants   = sec_extra
                    passing.append(c)

        # ── OP_STOP PATH: ND→N / MB→M variants ──────────────────────────────────
        for stop_cons, stop_label in op_stop_variants:
            if len(stop_cons) >= 2:
                mapped_stop = self.map_consonants_to_arabic(stop_cons)
                stop_roots  = self.generate_root_permutations(mapped_stop)
                for rs in stop_roots:
                    root_size  = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                    stop_extra = max(0, len(stop_cons) - root_size)
                    if rs in seen:
                        # Update extra_consonants if OP_STOP path is better
                        for existing in passing:
                            if existing.letters == rs and stop_extra < existing.extra_consonants:
                                existing.extra_consonants = stop_extra
                                existing.operations = [stop_label]
                                break
                        continue
                    seen.add(rs)
                    qr = self.q_gate.check(rs)
                    if qr.passed:
                        pos_s, trans = self._positional_score(stop_cons, rs)
                        c = RootCandidate(
                            letters     = rs,
                            token_count = qr.details.get('token_count', 0),
                            lemma_count = qr.details.get('lemma_count', 0),
                            ar_word     = qr.details.get('ar_word', ''),
                            operations  = [stop_label]
                        )
                        c.positional_score   = pos_s
                        c.transposition_flag = trans
                        c.extra_consonants   = stop_extra
                        passing.append(c)

        # ── R08a: M-PREFIX PARALLEL PATH ─────────────────────────────────────────
        # Gate 3d (v2.1): TWO M-prefix patterns tested in parallel:
        #   مُ (mu-) = active participle prefix (مُرْسَل → MIRACLE)
        #   مَ (ma-) = مَفْعَل place noun prefix (مَرْكَز → MARKET)
        if en_word.lower().startswith('m') and not any('OP_PREFIX' in op for op in ops):
            m_stripped, m_ops, m_sfx = self.strip_operations(en_word[1:])
            m_consonants             = self.extract_consonants(m_stripped)
            if len(m_consonants) >= 2:
                m_mapped  = self.map_consonants_to_arabic(m_consonants)
                m_roots   = self.generate_root_permutations(m_mapped)
                prefix_labels = ['OP_PREFIX(مُ→m)', 'OP_PREFIX(مَفْعَل→m)']
                for pfx_label in prefix_labels:
                    for rs in m_roots:
                        seen_key = rs + '__' + pfx_label
                        if seen_key in seen:
                            continue
                        seen.add(seen_key)
                        qr = self.q_gate.check(rs)
                        if qr.passed:
                            pos_s, trans = self._positional_score(m_consonants, rs)
                            m_c = RootCandidate(
                                letters     = rs,
                                token_count = qr.details.get('token_count', 0),
                                lemma_count = qr.details.get('lemma_count', 0),
                                ar_word     = qr.details.get('ar_word', ''),
                                operations  = [pfx_label] + m_ops
                            )
                            m_c.positional_score   = pos_s
                            m_c.transposition_flag = trans
                            m_root_size = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                            m_c.extra_consonants   = max(0, len(m_consonants) - m_root_size)
                            passing.append(m_c)

        # ── SORT: N15 always first (R09), then positional_score DESC, token_count DESC ──
        n15_hits = [c for c in passing if getattr(c, '_n15_priority', False)]
        others   = [c for c in passing if not getattr(c, '_n15_priority', False)]
        others.sort(key=lambda r: (r.positional_score, r.token_count), reverse=True)
        # v2.5: Increase candidate pool to 15 (was 10) for multi-candidate scoring.
        # Also guarantee TIER DIVERSITY: include the best candidate from each
        # extra_consonants tier, even if it would otherwise be cut by [:15].
        top_n = (n15_hits + others)[:15]
        # Collect tiers already represented
        tiers_present = set(getattr(c, 'extra_consonants', 99) for c in top_n)
        # Add best-from-missing-tiers from the full 'others' list
        for cand in others[15:]:
            tier = getattr(cand, 'extra_consonants', 99)
            if tier not in tiers_present:
                top_n.append(cand)
                tiers_present.add(tier)
        return top_n

    def _positional_score(self, consonants: list, root_letters: str) -> tuple:
        """
        R11 — Transposition as Semantic-First Diagnostic.
        Scores how closely the consonant ORDER in the English word matches the root order.
        If an engine assigned the wrong root because of semantic pull, the consonant
        positions will be OUT of order (transposed) — this catches that failure.

        Algorithm:
          For each root letter (in order), find the FIRST English consonant that
          could map to it (via forward_map).  Collect the position indices.
          If positions are monotonically increasing → correct order → score 1.0.
          If strictly reversed → transposition detected → score 0.1, flag True.
          Partial disorder → score 0.4, flag True.

        Returns: (positional_score: float, transposition_flag: bool)
        """
        root_list = [l.strip() for l in re.split(r'[\-\s]+', root_letters) if l.strip()]
        if not root_list or not consonants:
            return 0.5, False

        match_positions = []
        for ar in root_list:
            info = self.forward_map.get(ar)
            if not info:
                continue
            _, en_outputs = info
            found_pos = None
            for i, c in enumerate(consonants):
                if c in en_outputs:
                    found_pos = i
                    break
                # Partial: single-char match against first char of digraph outputs
                if len(c) == 1 and any(c == eo[0] for eo in en_outputs if eo):
                    found_pos = i
                    break
            if found_pos is not None:
                match_positions.append(found_pos)

        if len(match_positions) < 2:
            return 0.5, False   # Not enough data — neutral

        # Monotonically increasing = correct order
        in_order = all(match_positions[i] < match_positions[i + 1]
                       for i in range(len(match_positions) - 1))
        if in_order:
            return 1.0, False

        # Strictly reversed = mirror transposition (strongest R11 signal)
        reversed_order = all(match_positions[i] > match_positions[i + 1]
                             for i in range(len(match_positions) - 1))
        if reversed_order:
            return 0.1, True

        # Partial disorder
        return 0.4, True

    def _check_n15_priority(self, consonants: list) -> bool:
        """
        R09: Check if consonant skeleton matches N15 pattern → force ق-ر-ن first.
        N15 network = القَرْن DERIVATIVE FAMILY.  Triggered if ANY valid triple
        (ci, ri, ni) exists where ci < ri < ni — handles words like CONCERN where
        a nasal appears early before 'r' but another 'n' follows (c-n-c-R-N).
        """
        ck_set    = {'c', 'k', 'g', 'q'}
        ck_pos    = [i for i, c in enumerate(consonants) if c in ck_set]
        r_pos     = [i for i, c in enumerate(consonants) if c == 'r']
        n_pos     = [i for i, c in enumerate(consonants) if c == 'n']
        if not (ck_pos and r_pos and n_pos):
            return False
        # Check if ANY valid triple exists with ci < ri < ni
        for ci in ck_pos:
            for ri in r_pos:
                if ri <= ci:
                    continue
                for ni in n_pos:
                    if ni > ri:
                        return True
        return False

    def _generate_op_stop_variants(self, consonants: list) -> list:
        """
        v2.4: OP_STOP as candidate generator, not destructive pre-processor.
        Finds consecutive N-D or M-B in consonant skeleton and generates
        variants with the stop removed (ND→N, MB→M).

        This is the correct architecture: OP_STOP is a HYPOTHESIS about the
        word's history, not a certainty. CALENDAR has ND but it's NOT from NN
        gemination. TANDOOR has ND from NN (تَنُّور). By generating BOTH
        variants (with and without OP_STOP), the Q-gate and positional score
        determine which is correct.

        Returns: list of (modified_consonants, op_label) tuples
        """
        variants = []
        # ND→N: find consecutive ['n', 'd'] and remove 'd'
        for i in range(len(consonants) - 1):
            if consonants[i] == 'n' and consonants[i + 1] == 'd':
                new_cons = consonants[:i + 1] + consonants[i + 2:]
                variants.append((new_cons, 'OP_STOP(ND→N)'))
                break  # only first occurrence
        # MB→M: find consecutive ['m', 'b'] and remove 'b'
        for i in range(len(consonants) - 1):
            if consonants[i] == 'm' and consonants[i + 1] == 'b':
                new_cons = consonants[:i + 1] + consonants[i + 2:]
                variants.append((new_cons, 'OP_STOP(MB→M)'))
                break
        return variants


# ═══════════════════════════════════════════════════════════════════════════════
# COMPONENT 2b — RussianPhoneticReversal (ORIG1 + ORIG2 dual track)
# ═══════════════════════════════════════════════════════════════════════════════

# Russian suffixes for stripping (longest first) — grammatical + derivational
RUSSIAN_SUFFIXES = sorted([
    'ность', 'ство', 'ение', 'ание', 'ация', 'ация',
    'тель', 'ский', 'ская', 'ское', 'ские',
    'ость', 'ник', 'чик', 'щик', 'ция',
    'ный', 'ная', 'ное', 'ные', 'ной', 'ной',
    'ить', 'ать', 'ять', 'еть', 'ова',
    'ка', 'ок', 'ик', 'ек', 'ёк',
    'ый', 'ая', 'ое', 'ые', 'ий', 'ой',
    'ь',   # soft sign at word end — strip
], key=len, reverse=True)


class RussianPhoneticReversal:
    """
    Russian word → ranked ORIG1/ORIG2 root candidates.
    Works backward through M1_ФОНЕТИЧЕСКИЕ_СДВИГИ (Russian shift table).

    Key differences from English PhoneticReversal:
      - Cyrillic consonant extraction (no digraphs — each letter = one phoneme)
      - Russian vowels: а,е,ё,и,о,у,ы,э,ю,я
      - Russian soft/hard signs (ь,ъ) treated as modifiers, not consonants
      - Loads from M1_ФОНЕТИЧЕСКИЕ_СДВИГИ sheet (Russian column names)
      - >50% Bitig (ORIG2) influence — dual-track processing
      v3.1 additions:
      - Compound word detection (САМ+О+ВАР, ПАРО+ВОЗ patterns)
      - Palatalization stripping (Д↔Ж, Т↔Ч, С↔Ш, К↔Ч, Г↔Ж, СТ↔Щ)
      - Latin-to-Cyrillic transliteration (user has no Cyrillic keyboard)
    """

    CYRILLIC_VOWELS     = set('аеёиоуыэюя')
    CYRILLIC_MODIFIERS  = set('ьъ')   # soft/hard signs
    CYRILLIC_CONSONANTS = set('бвгджзклмнпрстфхцчшщ')

    # ── LATIN → CYRILLIC TRANSLITERATION TABLE ──────────────────────────────
    # User types Latin script → engine converts to Cyrillic before processing.
    # Digraphs FIRST (longest match), then single chars.
    LATIN_TO_CYRILLIC_DIGRAPHS = [
        ('shch', 'щ'), ('sch', 'щ'),
        ('zh', 'ж'), ('kh', 'х'), ('ch', 'ч'), ('sh', 'ш'),
        ('ts', 'ц'), ('yu', 'ю'), ('ya', 'я'), ('yo', 'ё'),
    ]
    LATIN_TO_CYRILLIC_SINGLE = {
        'a': 'а', 'b': 'б', 'v': 'в', 'g': 'г', 'd': 'д',
        'e': 'е', 'z': 'з', 'i': 'и', 'j': 'й',
        'k': 'к', 'l': 'л', 'm': 'м', 'n': 'н', 'o': 'о',
        'p': 'п', 'r': 'р', 's': 'с', 't': 'т', 'u': 'у',
        'f': 'ф', 'h': 'х', 'c': 'ц', 'w': 'в',
        'x': 'кс', 'y': 'ы',
    }

    # ── COMPOUND WORD PREFIXES ──────────────────────────────────────────────
    # Russian (like German) merges roots using О or Е as bridge vowels.
    # Pattern: PREFIX + О/Е + ROOT. Engine splits at bridge, processes each part.
    COMPOUND_PREFIXES = {
        # prefix_cyrillic: (meaning, strip_length_including_bridge_vowel)
        'само': ('self/auto', True),      # самовар, самолёт, самосвал
        'сам':  ('self/auto', True),       # when bridge vowel is already next char
        'паро': ('steam', True),           # паровоз, пароход
        'пар':  ('steam', True),
        'водо': ('water', True),           # водопровод, водопад
        'вод':  ('water', True),
        'полу': ('half', False),           # полуостров — no bridge vowel
        'обще': ('common', False),         # общежитие
        'ледо': ('ice', True),             # ледокол
        'звуко': ('sound', True),          # звукозапись
        'земле': ('earth', True),          # землетрясение
        'тепло': ('warmth', True),         # теплоход
        'хлебо': ('bread', True),          # хлебозавод
        'нефте': ('oil', True),            # нефтепровод
        'верто': ('spin', True),           # вертолёт
    }

    # ── PALATALIZATION MAP ──────────────────────────────────────────────────
    # Russian morphological alternations — these are NOT separate consonants.
    # The palatalized form must be UN-palatalized to recover the true root.
    # Direction: palatalized → base (what the engine should trace).
    DEPALATALIZE = {
        'ж': ['д', 'г', 'з'],    # водить→вождь, бег→бежать, возить→вожу
        'ч': ['т', 'к'],          # крутить→кручение, рука→ручной
        'ш': ['с', 'х'],          # писать→пишу, тихий→тишина
        'щ': ['ст', 'ск', 'т'],   # простить→прощение, искать→ищу, светить→свещение
    }
    # Reverse: which consonants CAN palatalize
    CAN_PALATALIZE = {'д', 'г', 'з', 'т', 'к', 'с', 'х', 'ст', 'ск'}

    # ── CYRILLIC → LATIN CONVERSION (for Kashgari ORIG2 search) ──────────────
    # v3.2: Russian consonants must be converted to Latin equivalents before
    # searching Kashgari corpus (which uses Latin transliteration).
    # Single-char mapping for skeleton matching.
    CYRILLIC_TO_LATIN_SIMPLE = {
        'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'ж': 'j',
        'з': 'z', 'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n',
        'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'ф': 'f',
        'х': 'h', 'ц': 'c', 'ч': 'c', 'ш': 's', 'щ': 's',
        'дж': 'j',
    }
    # Russian → Turkic voicing equivalences (for broader Kashgari search)
    # These pairs represent systematic voicing alternations in the
    # Turkic→Russian borrowing corridor.
    RU_TURKIC_VOICING = {
        'б': ['п'],       # б↔p voicing
        'п': ['б'],
        'г': ['к', 'q'],  # г↔k velar voicing
        'к': ['г', 'q'],
        'д': ['т'],       # д↔t voicing
        'т': ['д'],
        'ж': ['ч', 'ш'],  # affricate/sibilant alternation
        'ч': ['ж', 'ц'],
        'з': ['с'],       # з↔s voicing
        'с': ['з'],
    }

    def to_latin_skeleton(self, cyrillic_consonants: list) -> str:
        """Convert Cyrillic consonant list to Latin skeleton for Kashgari search."""
        return ''.join(self.CYRILLIC_TO_LATIN_SIMPLE.get(c, c) for c in cyrillic_consonants)

    def to_latin_skeleton_variants(self, cyrillic_consonants: list) -> list:
        """Generate Latin skeleton + voicing variants for Kashgari search.

        Returns list of Latin skeleton strings (main + voicing alternations).
        Also generates suffix-stripped variants (B03 agglutinative morphology).
        """
        main = self.to_latin_skeleton(cyrillic_consonants)
        variants = [main]

        # Voicing variants (swap one consonant at a time)
        for i, cyr_c in enumerate(cyrillic_consonants):
            for alt_cyr in self.RU_TURKIC_VOICING.get(cyr_c, []):
                alt_lat = self.CYRILLIC_TO_LATIN_SIMPLE.get(alt_cyr, alt_cyr)
                v = main[:i] + alt_lat + main[i+1:]
                if v != main and v not in variants:
                    variants.append(v)

        # Suffix-stripped variants (Russian endings that aren't root consonants)
        # Common Russian noun/adj endings: -ля, -ка, -ня, -ра etc.
        if len(main) >= 3:
            stripped_1 = main[:-1]    # drop last consonant
            if stripped_1 not in variants:
                variants.append(stripped_1)
            # Also voicing variants of stripped
            for i, cyr_c in enumerate(cyrillic_consonants[:-1]):
                for alt_cyr in self.RU_TURKIC_VOICING.get(cyr_c, []):
                    alt_lat = self.CYRILLIC_TO_LATIN_SIMPLE.get(alt_cyr, alt_cyr)
                    v = stripped_1[:i] + alt_lat + stripped_1[i+1:]
                    if v != stripped_1 and v not in variants:
                        variants.append(v)

        return variants[:20]  # cap to prevent explosion

    def __init__(self, master_file: str, q_gate: 'QGate'):
        self.q_gate      = q_gate
        self.shift_data  : List[dict] = []
        self.forward_map : Dict[str, tuple] = {}   # AR_letter → (shift_id, [ru_chars])
        self.reverse_map : Dict[str, list]  = {}   # RU_char → [(AR_letter, shift_id)]
        self._load_russian_shifts(master_file)
        self._build_reverse_map()
        # Extra mappings not in shift table (gap fill)
        self._add_gap_mappings()
        print(f"  RussianPhoneticReversal: {len(self.shift_data)} shifts, "
              f"{len(self.reverse_map)} RU patterns in reverse map")

    def _load_russian_shifts(self, filepath: str):
        """Load from M1_ФОНЕТИЧЕСКИЕ_СДВИГИ sheet."""
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb['M1_ФОНЕТИЧЕСКИЕ_СДВИГИ']
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not any(row):
                    continue
                d = dict(zip(headers, row))
                sid    = str(d.get('СДВИГ_ID', '') or '').strip()
                ar     = str(d.get('АР_БУКВА', '') or '').strip()
                ru_raw = str(d.get('РУС_ВЫХОДЫ', '') or '').strip()
                if sid and ar:
                    # Parse Russian outputs — handle "(выпадает)" = drops
                    ru_list = []
                    for x in ru_raw.split(','):
                        x = x.strip().lower()
                        if x and x not in ('(выпадает)', ''):
                            ru_list.append(x)
                    self.shift_data.append({
                        'shift_id': sid, 'ar_letter': ar, 'ru_outputs': ru_list
                    })
                    self.forward_map[ar] = (sid, ru_list)
            wb.close()
        except Exception as e:
            print(f"  RussianPhoneticReversal load error: {e}")

    def _build_reverse_map(self):
        """Build reverse map: Russian char → [(Arabic letter, shift_id)]."""
        for shift in self.shift_data:
            ar, sid = shift['ar_letter'], shift['shift_id']
            for ru in shift['ru_outputs']:
                key = ru.lower()
                self.reverse_map.setdefault(key, [])
                if (ar, sid) not in self.reverse_map[key]:
                    self.reverse_map[key].append((ar, sid))

    def _add_gap_mappings(self):
        """Add consonant mappings not explicitly in the shift table."""
        # ч (ch) — maps to ج (S02) or ش (S05) or ت+ش compound
        # Common in Turkic loanwords: чай (tea), чулан (closet)
        if 'ч' not in self.reverse_map:
            self.reverse_map['ч'] = [('ج', 'S02'), ('ش', 'S05')]
        # Ensure дж is mapped
        if 'дж' not in self.reverse_map:
            self.reverse_map['дж'] = [('ج', 'S02')]

    # ── v3.1: LATIN → CYRILLIC TRANSLITERATION ──────────────────────────────

    def transliterate_latin(self, text: str) -> str:
        """
        Convert Latin-script Russian to Cyrillic.
        User has no Cyrillic keyboard — types: samovar, moloko, vodka, etc.
        Returns Cyrillic string.

        Handles digraphs first (longest match): zh→ж, kh→х, ch→ч, sh→ш, ts→ц
        Then single chars: a→а, b→б, etc.
        """
        t = text.lower().strip()
        result = []
        i = 0
        while i < len(t):
            matched = False
            # Try digraphs (longest first — shch before sh)
            for lat, cyr in self.LATIN_TO_CYRILLIC_DIGRAPHS:
                if t[i:i+len(lat)] == lat:
                    result.append(cyr)
                    i += len(lat)
                    matched = True
                    break
            if not matched:
                ch = t[i]
                if ch in self.LATIN_TO_CYRILLIC_SINGLE:
                    result.append(self.LATIN_TO_CYRILLIC_SINGLE[ch])
                else:
                    result.append(ch)  # spaces, hyphens, digits pass through
                i += 1
        return ''.join(result)

    def _is_latin_russian(self, text: str) -> bool:
        """
        Detect if a string is Latin-script Russian (not English).
        Heuristic: contains common Russian transliteration patterns
        OR matches a known Russian word transliteration.
        """
        t = text.lower().strip()
        # If it has any Cyrillic already → not Latin-Russian
        if any(c in self.CYRILLIC_VOWELS or c in self.CYRILLIC_CONSONANTS
               or c in self.CYRILLIC_MODIFIERS for c in t):
            return False
        # Check for Russian transliteration digraph markers
        ru_digraphs = ['zh', 'kh', 'shch', 'ya', 'yu', 'yo', 'ts']
        if any(d in t for d in ru_digraphs):
            return True
        # Check for Russian word-ending patterns
        ru_endings = ['ov', 'ev', 'aya', 'iya', 'ost', 'nik', 'tel',
                      'stvo', 'ok', 'ka', 'ko', 'da', 'lo']
        if any(t.endswith(e) for e in ru_endings):
            # Also check it's NOT a common English word
            english_words = {'book', 'look', 'cook', 'hook', 'took',
                            'like', 'make', 'take', 'wake', 'bake',
                            'also', 'into', 'onto', 'undo', 'solo'}
            if t not in english_words:
                return True
        return False

    # ── v3.1: COMPOUND WORD DETECTION ────────────────────────────────────────

    def detect_compound(self, word: str) -> tuple:
        """
        Detect Russian compound words with О/Е bridge vowels.

        Russian and German both merge roots: PREFIX + О/Е + ROOT.
        Examples:
          САМОВАР  = САМ + О + ВАР  (self + cook/boil)
          САМОЛЁТ  = САМ + О + ЛЁТ  (self + fly)
          ПАРОВОЗ  = ПАР + О + ВОЗ  (steam + carry)
          ВОДОПАД  = ВОД + О + ПАД  (water + fall)
          ВЕРТОЛЁТ = ВЕРТ + О + ЛЁТ (spin + fly)
          ЛЕДОКОЛ  = ЛЕД + О + КОЛ  (ice + split)

        Returns: (is_compound, prefix_str, root_str, bridge_vowel, compound_label)
                 or (False, None, None, None, None) if not compound.
        """
        w = word.lower().strip()

        # Sort compound prefixes by length (longest first) to avoid partial matches
        sorted_prefixes = sorted(self.COMPOUND_PREFIXES.keys(), key=len, reverse=True)

        for prefix in sorted_prefixes:
            if not w.startswith(prefix):
                continue
            meaning, expects_bridge = self.COMPOUND_PREFIXES[prefix]
            remainder = w[len(prefix):]

            if expects_bridge:
                # Check for О/Е bridge vowel after prefix
                if remainder and remainder[0] in ('о', 'е'):
                    bridge = remainder[0]
                    root_part = remainder[1:]
                    if len(root_part) >= 2:  # root must have at least 2 chars
                        label = f"COMPOUND({prefix.upper()}+{bridge}+{root_part.upper()})"
                        return (True, prefix, root_part, bridge, label)
                # Also check: prefix already ends with the vowel (like САМО, ПАРО, ВОДО)
                # In this case the bridge is already included in the prefix
                elif len(remainder) >= 2:
                    label = f"COMPOUND({prefix.upper()}+{remainder.upper()})"
                    return (True, prefix, remainder, '', label)
            else:
                # No bridge vowel expected (ПОЛУ, ОБЩЕ)
                if len(remainder) >= 2:
                    label = f"COMPOUND({prefix.upper()}+{remainder.upper()})"
                    return (True, prefix, remainder, '', label)

        # Also detect non-prefix compounds: ROOT+О/Е+ROOT pattern
        # CONSERVATIVE: only fire when BOTH parts have >= 3 consonants each
        # and the word is long enough (>= 8 chars) to avoid false positives
        # like ХЛОПОК, МОЛОКО, ПОЛОСА which are NOT compounds.
        if len(w) >= 8:
            for i in range(3, len(w) - 3):
                if w[i] in ('о', 'е'):
                    # Check: consonant immediately before AND after bridge
                    if (w[i-1] in self.CYRILLIC_CONSONANTS and
                        w[i+1] in self.CYRILLIC_CONSONANTS):
                        left  = w[:i]
                        right = w[i+1:]
                        left_cons  = sum(1 for c in left  if c in self.CYRILLIC_CONSONANTS)
                        right_cons = sum(1 for c in right if c in self.CYRILLIC_CONSONANTS)
                        if left_cons >= 3 and right_cons >= 3:
                            label = f"COMPOUND({left.upper()}+{w[i]}+{right.upper()})"
                            return (True, left, right, w[i], label)

        return (False, None, None, None, None)

    # ── v3.1: PALATALIZATION STRIPPING ───────────────────────────────────────

    def depalatalize(self, consonants: list) -> list:
        """
        Generate de-palatalized consonant variants.

        Russian has systematic morphological alternations:
          Д → Ж  (водить → вождь)
          Т → Ч  (крутить → кручение)
          С → Ш  (писать → пишу)
          К → Ч  (рука → ручной)
          Г → Ж  (бег → бежать)
          З → Ж  (возить → вожу)
          СТ → Щ (простить → прощение)
          СК → Щ (искать → ищу)

        These are NOT separate consonants — they are surface alternations
        of the SAME underlying root consonant.

        Returns: list of (new_consonants, op_label) tuples.
                 Each tuple represents one possible de-palatalization.
        """
        variants = []

        for i, c in enumerate(consonants):
            if c in self.DEPALATALIZE:
                for base in self.DEPALATALIZE[c]:
                    if len(base) == 1:
                        # Single consonant replacement: ж→д, ч→т, etc.
                        new_cons = consonants[:i] + [base] + consonants[i+1:]
                        label = f'OP_DEPALATAL({c.upper()}→{base.upper()})'
                        variants.append((new_cons, label))
                    elif len(base) == 2:
                        # Cluster replacement: щ→ст, щ→ск — one consonant expands to two
                        new_cons = consonants[:i] + list(base) + consonants[i+1:]
                        label = f'OP_DEPALATAL({c.upper()}→{base.upper()})'
                        variants.append((new_cons, label))

        return variants

    # ── consonant extraction ──────────────────────────────────────────────────

    def extract_consonants(self, word: str) -> list:
        """
        Extract ordered consonant skeleton from Russian word.
        Each Cyrillic letter = one phoneme (no digraphs like English TH/SH).
        Exception: дж = one phoneme (affricate).
        """
        w = word.lower().strip()
        result = []
        i = 0
        while i < len(w):
            # Check for дж digraph
            if i + 1 < len(w) and w[i:i+2] == 'дж':
                result.append('дж')
                i += 2
            elif w[i] in self.CYRILLIC_CONSONANTS:
                result.append(w[i])
                i += 1
            else:
                # vowels, modifiers, spaces — skip
                i += 1
        return result

    def strip_operations(self, word: str) -> tuple:
        """
        Strip Russian suffixes (OP_SUFFIX equivalent).
        Returns: (stripped_word, operations_list, suffix_removed)
        Minimum-consonant guard: if stripping leaves < 3 consonants, undo.
        """
        w = word.lower().strip()
        ops = []
        suffix_removed = ''

        for suffix in RUSSIAN_SUFFIXES:
            if w.endswith(suffix) and len(w) - len(suffix) >= 2:
                candidate = w[:-len(suffix)]
                # Count consonants in stripped form
                cons_count = sum(1 for ch in candidate if ch in self.CYRILLIC_CONSONANTS)
                if cons_count >= 3:
                    w = candidate
                    suffix_removed = suffix
                    ops.append(f'OP_SUFFIX(-{suffix})')
                break
        return w, ops, suffix_removed

    def map_consonants_to_arabic(self, consonants: list) -> list:
        """Each consonant position → list of (AR_letter, shift_id) pairs."""
        mapped = []
        for c in consonants:
            candidates = self.reverse_map.get(c, [])
            mapped.append(candidates)
        return mapped

    def generate_root_permutations(self, mapped: list) -> list:
        """Generate 3-consonant root strings from mapped consonant candidates."""
        positions = len(mapped)
        if positions < 2:
            return []
        ar_per_pos = [list({ar for ar, sid in pos}) for pos in mapped]
        roots = set()
        n = 3 if positions >= 3 else positions
        for pos_combo in itertools.combinations(range(positions), n):
            for combo in itertools.product(*[ar_per_pos[p] for p in pos_combo]):
                if all(combo):
                    roots.add('-'.join(combo))
        return list(roots)

    def _positional_score(self, consonants: list, root_letters: str) -> tuple:
        """
        R11 — Positional score (same logic as English).
        Scores how closely consonant ORDER in Russian word matches root order.
        """
        root_list = [l.strip() for l in re.split(r'[\-\s]+', root_letters) if l.strip()]
        if not root_list or not consonants:
            return 0.5, False

        match_positions = []
        for ar in root_list:
            info = self.forward_map.get(ar)
            if not info:
                continue
            _, ru_outputs = info
            found_pos = None
            for i, c in enumerate(consonants):
                if c in ru_outputs:
                    found_pos = i
                    break
            if found_pos is not None:
                match_positions.append(found_pos)

        if len(match_positions) < 2:
            return 0.5, False

        monotone = all(match_positions[i] <= match_positions[i + 1]
                       for i in range(len(match_positions) - 1))
        if monotone:
            return 1.0, False

        reversed_check = all(match_positions[i] >= match_positions[i + 1]
                             for i in range(len(match_positions) - 1))
        if reversed_check:
            return 0.1, True
        return 0.4, True

    def _generate_op_stop_variants(self, consonants: list) -> list:
        """OP_STOP: НД→Н / МБ→М variants (same logic as English)."""
        variants = []
        # НД→Н: find consecutive ['н', 'д'] and remove 'д'
        for i in range(len(consonants) - 1):
            if consonants[i] == 'н' and consonants[i + 1] == 'д':
                new_cons = consonants[:i + 1] + consonants[i + 2:]
                variants.append((new_cons, 'OP_STOP(НД→Н)'))
                break
        # МБ→М: find consecutive ['м', 'б'] and remove 'б'
        for i in range(len(consonants) - 1):
            if consonants[i] == 'м' and consonants[i + 1] == 'б':
                new_cons = consonants[:i + 1] + consonants[i + 2:]
                variants.append((new_cons, 'OP_STOP(МБ→М)'))
                break
        return variants

    def reverse(self, ru_word: str) -> List[RootCandidate]:
        """
        Main public method: Russian word → ranked list of Qur'anic root candidates.
        Returns only candidates that pass Q-Gate.

        Architecture v3.1:
          COMPOUND:  detect САМ+О+ВАР type → split + process each part
          PRIMARY:   raw vowel extraction (vowel-strip-first)
          SECONDARY: suffix-stripped consonants
          DEPALATAL: undo Д↔Ж, Т↔Ч, С↔Ш alternations → re-run
          OP_STOP:   НД→Н / МБ→М variants
          M-PREFIX:  М-prefix parallel path (same as English R08a)
        """
        # ═══ v3.1: COMPOUND DETECTION ════════════════════════════════════════════
        is_compound, prefix_part, root_part, bridge, compound_label = self.detect_compound(ru_word)
        if is_compound:
            # Process only the ROOT part through the pipeline
            # The prefix is a known morpheme (САМ=self, ПАРО=steam, etc.)
            # Mark the result with the compound label
            root_candidates = self._reverse_inner(root_part, compound_label)
            # Also try the full word (some compounds have fused so much
            # that the root has its own Q-gate entry)
            full_candidates = self._reverse_inner(ru_word, None)
            # Merge: compound-rooted candidates get priority
            seen_letters = {c.letters for c in root_candidates}
            for fc in full_candidates:
                if fc.letters not in seen_letters:
                    root_candidates.append(fc)
            return root_candidates

        return self._reverse_inner(ru_word, None)

    def _reverse_inner(self, ru_word: str, compound_label: str = None) -> List[RootCandidate]:
        """
        Inner reverse logic — processes a single word (or compound root part).
        Separated from reverse() to allow compound detection to call this
        on just the root portion.
        """
        # ═══ PRIMARY: raw consonant extraction ═════════════════════════════════════
        primary_cons = self.extract_consonants(ru_word)

        # ═══ SECONDARY: suffix-stripped ═════════════════════════════════════════════
        stripped, ops, suffix = self.strip_operations(ru_word)
        secondary_cons = self.extract_consonants(stripped)

        # ═══ OP_STOP: НД→Н / МБ→М variants ════════════════════════════════════════
        op_stop_variants = self._generate_op_stop_variants(primary_cons)

        passing = []
        seen    = set()

        # ── PRIMARY PATH ───────────────────────────────────────────────────────────
        mapped_primary = self.map_consonants_to_arabic(primary_cons)
        primary_roots  = self.generate_root_permutations(mapped_primary)

        for rs in primary_roots:
            if rs in seen:
                continue
            seen.add(rs)
            qr = self.q_gate.check(rs)
            if qr.passed:
                pos_s, trans = self._positional_score(primary_cons, rs)
                root_size    = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                c = RootCandidate(
                    letters     = rs,
                    token_count = qr.details.get('token_count', 0),
                    lemma_count = qr.details.get('lemma_count', 0),
                    ar_word     = qr.details.get('ar_word', ''),
                    operations  = []
                )
                c.positional_score   = pos_s
                c.transposition_flag = trans
                c.extra_consonants   = max(0, len(primary_cons) - root_size)
                passing.append(c)

        # ── SECONDARY PATH ─────────────────────────────────────────────────────────
        if secondary_cons != primary_cons and len(secondary_cons) >= 2:
            mapped_sec = self.map_consonants_to_arabic(secondary_cons)
            sec_roots  = self.generate_root_permutations(mapped_sec)
            for rs in sec_roots:
                root_size = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                sec_extra = max(0, len(secondary_cons) - root_size)
                if rs in seen:
                    for existing in passing:
                        if existing.letters == rs and sec_extra < existing.extra_consonants:
                            existing.extra_consonants = sec_extra
                            if ops and not existing.operations:
                                existing.operations = ops
                            break
                    continue
                seen.add(rs)
                qr = self.q_gate.check(rs)
                if qr.passed:
                    pos_s, trans = self._positional_score(secondary_cons, rs)
                    c = RootCandidate(
                        letters     = rs,
                        token_count = qr.details.get('token_count', 0),
                        lemma_count = qr.details.get('lemma_count', 0),
                        ar_word     = qr.details.get('ar_word', ''),
                        operations  = ops
                    )
                    c.positional_score   = pos_s
                    c.transposition_flag = trans
                    c.extra_consonants   = sec_extra
                    passing.append(c)

        # ── OP_STOP PATH ──────────────────────────────────────────────────────────
        for stop_cons, stop_label in op_stop_variants:
            if len(stop_cons) >= 2:
                mapped_stop = self.map_consonants_to_arabic(stop_cons)
                stop_roots  = self.generate_root_permutations(mapped_stop)
                for rs in stop_roots:
                    root_size  = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                    stop_extra = max(0, len(stop_cons) - root_size)
                    if rs in seen:
                        for existing in passing:
                            if existing.letters == rs and stop_extra < existing.extra_consonants:
                                existing.extra_consonants = stop_extra
                                existing.operations = [stop_label]
                                break
                        continue
                    seen.add(rs)
                    qr = self.q_gate.check(rs)
                    if qr.passed:
                        pos_s, trans = self._positional_score(stop_cons, rs)
                        c = RootCandidate(
                            letters     = rs,
                            token_count = qr.details.get('token_count', 0),
                            lemma_count = qr.details.get('lemma_count', 0),
                            ar_word     = qr.details.get('ar_word', ''),
                            operations  = [stop_label]
                        )
                        c.positional_score   = pos_s
                        c.transposition_flag = trans
                        c.extra_consonants   = stop_extra
                        passing.append(c)

        # ── v3.1: DEPALATALIZATION PATH ──────────────────────────────────────────
        # Russian Д↔Ж, Т↔Ч, С↔Ш, К↔Ч, Г↔Ж are morphological alternations,
        # NOT separate consonants. Try un-palatalizing and re-running.
        depal_variants = self.depalatalize(primary_cons)
        for depal_cons, depal_label in depal_variants:
            if len(depal_cons) >= 2:
                mapped_depal = self.map_consonants_to_arabic(depal_cons)
                depal_roots  = self.generate_root_permutations(mapped_depal)
                for rs in depal_roots:
                    root_size  = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                    depal_extra = max(0, len(depal_cons) - root_size)
                    depal_key = rs + '__' + depal_label
                    if depal_key in seen:
                        continue
                    # Also skip if same root already found without depal
                    if rs in seen:
                        for existing in passing:
                            if existing.letters == rs and depal_extra < existing.extra_consonants:
                                existing.extra_consonants = depal_extra
                                if depal_label not in existing.operations:
                                    existing.operations.append(depal_label)
                                break
                        continue
                    seen.add(depal_key)
                    seen.add(rs)
                    qr = self.q_gate.check(rs)
                    if qr.passed:
                        pos_s, trans = self._positional_score(depal_cons, rs)
                        c = RootCandidate(
                            letters     = rs,
                            token_count = qr.details.get('token_count', 0),
                            lemma_count = qr.details.get('lemma_count', 0),
                            ar_word     = qr.details.get('ar_word', ''),
                            operations  = [depal_label]
                        )
                        c.positional_score   = pos_s
                        c.transposition_flag = trans
                        c.extra_consonants   = depal_extra
                        passing.append(c)

        # ── M-PREFIX PARALLEL PATH (R08a) ──────────────────────────────────────────
        if ru_word.lower().startswith('м'):
            m_stripped, m_ops, m_sfx = self.strip_operations(ru_word[1:])
            m_consonants             = self.extract_consonants(m_stripped)
            if len(m_consonants) >= 2:
                m_mapped = self.map_consonants_to_arabic(m_consonants)
                m_roots  = self.generate_root_permutations(m_mapped)
                for rs in m_roots:
                    seen_key = rs + '__OP_PREFIX(مُ→م)'
                    if seen_key in seen:
                        continue
                    seen.add(seen_key)
                    qr = self.q_gate.check(rs)
                    if qr.passed:
                        pos_s, trans = self._positional_score(m_consonants, rs)
                        m_c = RootCandidate(
                            letters     = rs,
                            token_count = qr.details.get('token_count', 0),
                            lemma_count = qr.details.get('lemma_count', 0),
                            ar_word     = qr.details.get('ar_word', ''),
                            operations  = ['OP_PREFIX(مُ→м)'] + m_ops
                        )
                        m_c.positional_score   = pos_s
                        m_c.transposition_flag = trans
                        m_root_size = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                        m_c.extra_consonants   = max(0, len(m_consonants) - m_root_size)
                        passing.append(m_c)

        # ── v3.4b: OP_RU_PREFIX — Russian grammatical prefix stripping ──────
        # Russian is FULL of prefixes (ДО-, ПО-, НА-, ПРИ-, ПРО-, ЗА-, etc.)
        # These are NOT root consonants — strip before tracing.
        # Same principle as OP_SUFFIX for Latin/Greek but at the FRONT.
        # Example: ДОГОВОР: strip ДО- → ГОВОР → Г-В-Р → ج-ب-ر (21 tokens)
        #          ЗАГОВОР: strip ЗА- → ГОВОР → Г-В-Р → ج-ب-ر (same root)
        RU_PREFIXES = [
            ('пере', 4), ('рас', 3), ('раз', 3), ('вос', 3), ('воз', 3),
            ('при', 3), ('пре', 3), ('про', 3), ('под', 3), ('над', 3),
            ('по', 2), ('на', 2), ('за', 2), ('до', 2), ('от', 2),
            ('из', 2), ('вы', 2), ('об', 2),
            ('у', 1), ('с', 1),
        ]
        ru_lower = ru_word.lower()
        for pfx, pfx_len in RU_PREFIXES:
            if ru_lower.startswith(pfx) and len(ru_lower) > pfx_len + 2:
                remainder = ru_lower[pfx_len:]
                rem_cons = self.extract_consonants(remainder)
                if len(rem_cons) >= 2:
                    rem_mapped = self.map_consonants_to_arabic(rem_cons)
                    rem_roots  = self.generate_root_permutations(rem_mapped)
                    for rs in rem_roots:
                        seen_key = rs + f'__OP_RU_PREFIX({pfx.upper()}-)'
                        if seen_key in seen:
                            continue
                        if rs in seen:
                            # Same root found without prefix — check if prefix
                            # version has fewer extra consonants
                            for existing in passing:
                                root_size = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                                pfx_extra = max(0, len(rem_cons) - root_size)
                                if existing.letters == rs and pfx_extra < existing.extra_consonants:
                                    existing.extra_consonants = pfx_extra
                                    op_label = f'OP_RU_PREFIX({pfx.upper()}-)'
                                    if op_label not in existing.operations:
                                        existing.operations.append(op_label)
                                    break
                            continue
                        seen.add(seen_key)
                        seen.add(rs)
                        qr = self.q_gate.check(rs)
                        if qr.passed:
                            pos_s, trans = self._positional_score(rem_cons, rs)
                            root_size = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                            c = RootCandidate(
                                letters     = rs,
                                token_count = qr.details.get('token_count', 0),
                                lemma_count = qr.details.get('lemma_count', 0),
                                ar_word     = qr.details.get('ar_word', ''),
                                operations  = [f'OP_RU_PREFIX({pfx.upper()}-)', f'remainder={remainder}']
                            )
                            c.positional_score   = pos_s
                            c.transposition_flag = trans
                            c.extra_consonants   = max(0, len(rem_cons) - root_size)
                            passing.append(c)

        # ── v3.4: INITIAL STRIP — fallback for imperfect coverage ────────────
        # When NO candidate has perfect consonant coverage (extra=0), try
        # stripping the initial consonant.  This catches words like ТАЛАНТ/
        # ГАЛАНТ where the initial T/G varies — the root is in the shared
        # ending (-ЛАНТ → Л-Н-Т).
        # Fires when: (a) no candidates at all, OR (b) all candidates have
        # extra_consonants >= 1 (no clean trilateral match found).
        no_clean_match = (not passing or
                          all(getattr(c, 'extra_consonants', 99) >= 1
                              for c in passing))
        if no_clean_match and len(primary_cons) >= 3:
            init_stripped = primary_cons[1:]  # drop first consonant
            if len(init_stripped) >= 2:
                init_mapped = self.map_consonants_to_arabic(init_stripped)
                init_roots  = self.generate_root_permutations(init_mapped)
                for rs in init_roots:
                    if rs in seen:
                        continue
                    seen.add(rs)
                    qr = self.q_gate.check(rs)
                    if qr.passed:
                        pos_s, trans = self._positional_score(init_stripped, rs)
                        root_size = len([l for l in re.split(r'[\-\s]+', rs) if l.strip()])
                        c = RootCandidate(
                            letters     = rs,
                            token_count = qr.details.get('token_count', 0),
                            lemma_count = qr.details.get('lemma_count', 0),
                            ar_word     = qr.details.get('ar_word', ''),
                            operations  = [f'OP_INITIAL_STRIP({primary_cons[0].upper()}-)']
                        )
                        c.positional_score   = pos_s
                        c.transposition_flag = trans
                        c.extra_consonants   = max(0, len(init_stripped) - root_size)
                        passing.append(c)

        # ── v3.1: COMPOUND LABEL PROPAGATION ─────────────────────────────────────
        if compound_label:
            for cand in passing:
                if compound_label not in cand.operations:
                    cand.operations.insert(0, compound_label)

        # ── SORT + TIER DIVERSITY (same as English v2.5) ──────────────────────────
        passing.sort(key=lambda r: (r.positional_score, r.token_count), reverse=True)
        top_n = passing[:15]
        tiers_present = set(getattr(c, 'extra_consonants', 99) for c in top_n)
        for cand in passing[15:]:
            tier = getattr(cand, 'extra_consonants', 99)
            if tier not in tiers_present:
                top_n.append(cand)
                tiers_present.add(tier)
        return top_n


# ═══════════════════════════════════════════════════════════════════════════════
# COMPONENT 4 — UGate
# ═══════════════════════════════════════════════════════════════════════════════

class UGate:
    """Phonetic unity gate — every consonant accounted for via M1 shifts."""

    # OP_VOICE pairs: voicing/devoicing equivalences (documented phonological process)
    # Maps each consonant to ALL possible voicing/devoicing partners.
    VOICE_PAIRS_EN = {
        'z': ['t', 's'],       # ز→t (markaz→market), ز→s (sibilant)
        't': ['d', 'z'],       # ت→d (DEBT), ت→z
        's': ['z'],            # sibilant voicing
        'd': ['t'],            # dental devoicing
        'p': ['b'],            # bilabial voicing
        'b': ['p'],            # bilabial devoicing
        'f': ['v'],            # labiodental voicing
        'v': ['f'],            # labiodental devoicing
        'k': ['g'],            # velar voicing
        'g': ['k'],            # velar devoicing
    }

    # Russian Cyrillic OP_VOICE pairs (v3.0)
    VOICE_PAIRS_RU = {
        'з': ['т', 'с'],       # з↔т, з↔с (sibilant)
        'т': ['д', 'з'],       # т↔д, т↔з
        'с': ['з'],            # sibilant voicing
        'д': ['т'],            # dental devoicing
        'п': ['б'],            # bilabial voicing
        'б': ['п'],            # bilabial devoicing
        'ф': ['в'],            # labiodental voicing
        'в': ['ф'],            # labiodental devoicing
        'к': ['г'],            # velar voicing
        'г': ['к'],            # velar devoicing
        'ш': ['ж'],            # шипящие (sibilant voicing)
        'ж': ['ш'],            # шипящие (sibilant devoicing)
    }

    def __init__(self, reversal):
        self.reversal = reversal
        # Select voice pairs based on reversal type
        self.VOICE_PAIRS = (self.VOICE_PAIRS_RU
                            if isinstance(reversal, RussianPhoneticReversal)
                            else self.VOICE_PAIRS_EN)

    def verify(self, en_word: str, root_letters: str,
               operations: list = None) -> GateResult:
        stripped, ops_applied, suffix = self.reversal.strip_operations(en_word)
        consonants = self.reversal.extract_consonants(stripped)
        root_list  = [l.strip() for l in re.split(r'[\-\s]+', root_letters) if l.strip()]

        if not root_list:
            return GateResult(False, {'reason': 'Could not parse root letters'})

        chain_parts = []
        unmapped    = []

        for ar in root_list:
            info = self.reversal.forward_map.get(ar)
            if not info:
                unmapped.append(ar)
                continue
            sid, en_outputs = info
            matched = None
            voice_match = False
            for en_out in en_outputs:
                if en_out in consonants or en_out in stripped.lower():
                    matched = en_out
                    break
            # OP_VOICE fallback: if direct match failed, check voicing pairs
            if not matched:
                for en_out in en_outputs:
                    partners = self.VOICE_PAIRS.get(en_out, [])
                    for voiced in partners:
                        if voiced in consonants or voiced in stripped.lower():
                            matched = voiced
                            voice_match = True
                            break
                    if voice_match:
                        break
            if matched:
                if voice_match:
                    chain_parts.append(f"{ar}→{matched}({sid}+OP_VOICE)")
                else:
                    chain_parts.append(f"{ar}→{matched}({sid})")
            else:
                unmapped.append(ar)

        chain = ', '.join(chain_parts)
        all_ops = (operations or []) + ops_applied
        ops_str = ' | '.join(all_ops) if all_ops else ''

        if unmapped:
            # v2.4: RAW CONSONANT FALLBACK — try matching against full word
            # before declaring U-gate failure. Catches CALENDAR (OP_SUFFIX strips
            # -ar removing ر→r) and similar cases where suffix stripping removes
            # a root consonant that IS present in the original word.
            raw_cons = self.reversal.extract_consonants(en_word)
            if raw_cons != consonants:
                raw_chain_parts = []
                raw_unmapped    = []
                for ar in root_list:
                    info = self.reversal.forward_map.get(ar)
                    if not info:
                        raw_unmapped.append(ar)
                        continue
                    sid, en_outputs = info
                    matched = None
                    voice_match = False
                    for en_out in en_outputs:
                        if en_out in raw_cons or en_out in en_word.lower():
                            matched = en_out
                            break
                    if not matched:
                        for en_out in en_outputs:
                            partners = self.VOICE_PAIRS.get(en_out, [])
                            for voiced in partners:
                                if voiced in raw_cons or voiced in en_word.lower():
                                    matched = voiced
                                    voice_match = True
                                    break
                            if voice_match:
                                break
                    if matched:
                        if voice_match:
                            raw_chain_parts.append(f"{ar}→{matched}({sid}+OP_VOICE)")
                        else:
                            raw_chain_parts.append(f"{ar}→{matched}({sid})")
                    else:
                        raw_unmapped.append(ar)

                if not raw_unmapped:
                    # Raw consonants pass — use raw chain
                    raw_chain = ', '.join(raw_chain_parts)
                    all_ops_raw = (operations or []) + ops_applied
                    ops_str_raw = ' | '.join(all_ops_raw) if all_ops_raw else ''
                    return GateResult(True, {
                        'phonetic_chain': raw_chain,
                        'operations': ops_str_raw,
                        'consonant_skeleton': ''.join(raw_cons),
                        'stripped_word': en_word.lower(),
                        'note': 'Verified via raw consonants (vowel-strip-first)'
                    })

            return GateResult(False, {
                'phonetic_chain': chain,
                'unmapped': unmapped,
                'reason': f"Unmapped root letters: {', '.join(unmapped)}"
            })
        return GateResult(True, {
            'phonetic_chain': chain,
            'operations': ops_str,
            'consonant_skeleton': ''.join(consonants),
            'stripped_word': stripped
        })


# ═══════════════════════════════════════════════════════════════════════════════
# COMPONENT 5 — FGate
# ═══════════════════════════════════════════════════════════════════════════════

class FGate:
    """Foundation layer assignment — DS corridor, DP codes, network membership."""

    DP08_TRIGGERS = {
        'philosophy','medicine','algebra','algorithm','chemistry',
        'geometry','astronomy','physics','biology','science'
    }

    def __init__(self, master_file: str):
        self.networks        : Dict[str, dict] = {}
        self.root_to_network : Dict[str, str]  = {}
        self._load_networks(master_file)
        self._load_entry_networks(master_file)

    def _load_networks(self, filepath: str):
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb['M4_NETWORKS']
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not any(row):
                    continue
                d = dict(zip(headers, row))
                nid = str(d.get('NETWORK_ID', '') or '').strip()
                if nid:
                    self.networks[nid] = {
                        'title'     : str(d.get('TITLE', '') or '').strip(),
                        'link_verse': str(d.get('LINK_VERSE', '') or '').strip(),
                        'entry_ids' : str(d.get('ENTRY_IDS', '') or '').strip()
                    }
            wb.close()
        except Exception as e:
            print(f"  FGate networks load error: {e}")

    def _load_entry_networks(self, filepath: str):
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb['A1_ENTRIES']
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not any(row):
                    continue
                d  = dict(zip(headers, row))
                rl = str(d.get('ROOT_LETTERS', '') or '').strip()
                nid = str(d.get('NETWORK_ID', '') or '').strip()
                if rl and nid:
                    self.root_to_network[rl] = nid
            wb.close()
        except Exception as e:
            print(f"  FGate entries load error: {e}")

    def assign(self, en_word: str, root_letters: str,
               phonetic_chain: str = '') -> GateResult:
        ds_code   = self._detect_corridor(en_word)
        dp_codes  = self._detect_dp(en_word)
        network   = self.root_to_network.get(root_letters.strip(), '')

        parts = [f"F2: {ds_code}→AL"]
        if dp_codes:
            parts.append(' | '.join(dp_codes))
        if network:
            parts.append(network)

        return GateResult(True, {
            'ds_code'      : ds_code,
            'dp_codes'     : dp_codes,
            'network_id'   : network,
            'foundation_ref': ' | '.join(parts)
        })

    def _detect_corridor(self, word: str) -> str:
        w = word.lower()
        if any(x in w for x in ('ph','th','ys','ps','mn')):
            return 'DS04→DS05'
        if any(w.endswith(s) for s in ('tion','ment','ance','ence','ity','ous')):
            return 'DS05'
        if any(x in w for x in ('sch','tz','gh','wh')):
            return 'DS06'
        return 'DS05→AL'

    def _detect_dp(self, word: str) -> list:
        w = word.lower()
        dp = []
        if w in self.DP08_TRIGGERS or any(w.endswith(s) for s in ('ology','ics','phy')):
            dp.append('DP08')
        return dp


# ═══════════════════════════════════════════════════════════════════════════════
# SCORING ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class Scorer:
    """10-point scoring for a candidate entry."""

    def score(self, candidate: RootCandidate, en_word: str,
              q_result: GateResult, u_result: GateResult,
              f_result: GateResult) -> tuple:
        """
        Returns (score: int, breakdown: dict).

        v2 scoring — corrected per USLaP_BATCH_ENGINE_PROTOCOL:
          Token count capped at 1 point (was 3 — caused semantic-first bias).
          Positional fidelity (R11) contributes 2 points — replaces token inflation.
          Transposition penalty (-2) fires when R11 detects inverted consonant order.
        Max score breakdown: Q(2) + tokens(1) + positional(2) + U(2) + F(1) + chain(1) + network(1) = 10
        """
        s = 0
        b = {}

        # Q-gate pass (+2)
        if q_result and q_result.passed:
            s += 2; b['q_gate_pass'] = 2

        # Token count — capped at 1 point (was 3 — semantic-first bias eliminated)
        tokens = q_result.details.get('token_count', 0) if q_result else 0
        if tokens >= 100:
            s += 1; b['token_count_100'] = 1
        # Removed: >=20 (+1) and >=5 (+1) tiers — both caused semantic-first root selection

        # Positional fidelity — R11 (replaces raw token count as primary ranking signal)
        pos_score  = getattr(candidate, 'positional_score',   0.5)
        trans_flag = getattr(candidate, 'transposition_flag', False)
        if trans_flag:
            # R11 fires: consonant ORDER is inverted → deduct 2 (can go negative, min 0)
            s = max(0, s - 2); b['r11_transposition_penalty'] = -2
        elif pos_score >= 0.8:
            s += 2; b['positional_fidelity_high'] = 2
        elif pos_score >= 0.5:
            s += 1; b['positional_fidelity_ok'] = 1

        # U-gate pass (+2)
        if u_result and u_result.passed:
            s += 2; b['u_gate_pass'] = 2

        # F-gate pass (+1)
        if f_result and f_result.passed:
            s += 1; b['f_gate_pass'] = 1

        # Clean chain (+1) — no unmapped consonants
        if u_result and u_result.passed:
            chain = u_result.details.get('phonetic_chain', '')
            if chain and not u_result.details.get('unmapped'):
                s += 1; b['clean_chain'] = 1

        # Network membership found (+1)
        if f_result and f_result.details.get('network_id'):
            s += 1; b['network_found'] = 1

        # ── COVERAGE PENALTY: penalize if root covers too few word consonants ────
        # extra_consonants = word_consonants_stripped - root_size
        # Allows 1 extra for OP_NASAL or OP_STOP.  2+ = suspicious (unmapped letters).
        extra = getattr(candidate, 'extra_consonants', 0)
        if extra >= 3:
            s = max(0, s - 3); b['excess_consonants_major'] = -3
        elif extra == 2:
            s = max(0, s - 2); b['excess_consonants_minor'] = -2

        return min(s, 10), b


# ═══════════════════════════════════════════════════════════════════════════════
# COMPONENT 6 — ClusterExpander
# ═══════════════════════════════════════════════════════════════════════════════

class ClusterExpander:
    """
    Given a confirmed root, discovers all English words sharing that root.
    Searches /usr/share/dict/words (macOS) against forward-mapped consonant patterns.
    Places confirmed entries → A1_ENTRIES queue; ambiguous → ENGINE_QUEUE.
    """

    WORDLIST_PATH = '/usr/share/dict/words'

    def __init__(self, reversal: PhoneticReversal, existing_terms: dict):
        self.reversal       = reversal
        self.existing_terms = existing_terms   # en_term.upper() → entry_id
        self._wordlist      : Optional[List[str]] = None
        self._load_wordlist()

    def _load_wordlist(self):
        try:
            with open(self.WORDLIST_PATH, 'r', encoding='utf-8', errors='ignore') as f:
                self._wordlist = [line.strip().lower() for line in f
                                  if line.strip().isalpha() and len(line.strip()) >= 4]
            print(f"  ClusterExpander: {len(self._wordlist):,} words in wordlist")
        except FileNotFoundError:
            print(f"  ClusterExpander: wordlist not found at {self.WORDLIST_PATH} — cluster expansion limited")
            self._wordlist = []

    def _build_consonant_patterns(self, root_letters: str) -> list:
        """Forward-map root consonants to all EN character combinations."""
        root_list = [l.strip() for l in re.split(r'[\-\s]+', root_letters) if l.strip()]
        patterns  = []
        per_root  = []
        for ar in root_list:
            info = self.reversal.forward_map.get(ar)
            if info:
                _, en_outputs = info
                per_root.append([e for e in en_outputs if 1 <= len(e) <= 2])  # exclude empty strings
            else:
                per_root.append([])
        # Generate consonant skeletons: all combinations of en outputs
        for combo in itertools.product(*per_root):
            skeleton = ''.join(combo)
            if skeleton:
                patterns.append(skeleton)
        return list(set(patterns))

    def _word_matches_pattern(self, word: str, patterns: list) -> bool:
        """
        Check if word's consonant skeleton contains a root pattern with sufficient coverage.
        v2 fixes:
          - Require pattern length >= 3 (prevents spurious 1-2 char matches like 'bc').
          - Require coverage >= 40% (prevents matching long words on a tiny 3-char skeleton).
        """
        consonants = ''.join(self.reversal.extract_consonants(word))
        if not consonants:
            return False
        for pat in patterns:
            if len(pat) < 3:                          # Skip trivially short patterns
                continue
            if pat in consonants:
                # Coverage: pattern must cover >= 40% of word's consonants.
                # This blocks CONTROL (5 cons) matching a 2-char skeleton pattern.
                if len(pat) / len(consonants) >= 0.40:
                    return True
        return False

    def expand(self, root_letters: str, source_en_term: str,
               depth: int = 0) -> list:
        """
        Find all English words sharing root_letters.
        Returns list of candidate words (excluding source_en_term and existing entries).
        """
        if depth >= MAX_CLUSTER_DEPTH or not self._wordlist:
            return []

        patterns  = self._build_consonant_patterns(root_letters)
        if not patterns:
            return []

        candidates = []
        for word in self._wordlist:
            if word.upper() == source_en_term.upper():
                continue
            if word.upper() in self.existing_terms:
                continue
            if self._word_matches_pattern(word, patterns):
                candidates.append(word)

        # Cap at 15 candidates per expansion cycle (was 50 — produced dictionary dumps)
        return candidates[:15]


# ═══════════════════════════════════════════════════════════════════════════════
# COMPONENT 7 — EntryWriter
# ═══════════════════════════════════════════════════════════════════════════════

class EntryWriter:
    """
    Writes confirmed entries to master file.
    Updates: A1_ENTRIES (14 cols), A4_DERIVATIVES, SESSION_INDEX, ENGINE_QUEUE.
    Uses backup-before-write pattern.
    """

    def __init__(self, master_file: str):
        self.master_file = master_file
        self.backup_dir  = str(Path(master_file).parent / 'backups')
        os.makedirs(self.backup_dir, exist_ok=True)

    def _backup(self) -> str:
        ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
        dest = os.path.join(self.backup_dir, f"Master_backup_{ts}.xlsx")
        shutil.copy2(self.master_file, dest)
        return dest

    def _next_entry_id(self, ws) -> int:
        max_id = 248   # updated baseline — last confirmed entry is NORM #248
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and isinstance(row[0], (int, float)):
                max_id = max(max_id, int(row[0]))
        return max_id + 1

    def _next_empty_row(self, ws) -> int:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(c for c in row if c is not None):
                return i
        return ws.max_row + 1

    # ── public methods ────────────────────────────────────────────────────────

    def write_entry(self, entry: EntryRecord) -> int:
        """Write to A1_ENTRIES. Returns assigned ENTRY_ID."""
        # Self-audit: scan for banned terms
        self._self_audit(entry)

        backup = self._backup()
        try:
            wb = load_workbook(self.master_file)
            ws = wb['A1_ENTRIES']

            entry.entry_id = self._next_entry_id(ws)
            target         = self._next_empty_row(ws)

            for col, val in enumerate(entry.to_row(), start=1):
                ws.cell(row=target, column=col, value=val)

            self._log_session(wb, entry)
            wb.save(self.master_file)
            wb.close()
            print(f"  ✓ Written: #{entry.entry_id} {entry.en_term} → {entry.root_letters} (row {target})")
            return entry.entry_id
        except Exception as e:
            print(f"  ✗ Write failed: {e} — restoring backup")
            shutil.copy2(backup, self.master_file)
            raise

    def queue_for_oversight(self, entry: EntryRecord, flag_reason: str,
                            q_pass: bool, u_pass: bool, f_pass: bool) -> str:
        """Add entry to ENGINE_QUEUE. Returns QUEUE_ID."""
        try:
            wb = load_workbook(self.master_file)
            if 'ENGINE_QUEUE' not in wb.sheetnames:
                ws = wb.create_sheet('ENGINE_QUEUE')
                headers = [
                    'QUEUE_ID','STATUS','INPUT_TERM','ENTRY_CLASS','CANDIDATE_ROOT',
                    'Q_GATE','U_GATE','F_GATE','PHONETIC_CHAIN','DRAFT_ENTRY_ID',
                    'SCORE','FLAG_REASON','DISCOVERED_VIA','TIMESTAMP',
                    'USER_DECISION','DECISION_TIMESTAMP'
                ]
                for ci, h in enumerate(headers, 1):
                    ws.cell(row=1, column=ci, value=h)
            else:
                ws = wb['ENGINE_QUEUE']

            nxt      = ws.max_row + 1
            queue_id = f"Q{nxt - 1:04d}"
            ts       = datetime.now().strftime('%Y-%m-%d %H:%M')

            row_vals = [
                queue_id, 'PENDING', entry.en_term, 'LINGUISTIC', entry.root_letters,
                'PASS' if q_pass else 'FAIL',
                'PASS' if u_pass else 'FAIL',
                'PASS' if f_pass else 'FAIL',
                entry.phonetic_chain, None,
                entry.score, flag_reason, 'ENGINE_AUTO', ts, None, None
            ]
            for ci, v in enumerate(row_vals, 1):
                ws.cell(row=nxt, column=ci, value=v)

            wb.save(self.master_file)
            wb.close()
            print(f"  → Queued: {queue_id} ({entry.en_term}) — {flag_reason}")
            return queue_id
        except Exception as e:
            print(f"  ENGINE_QUEUE write error: {e}")
            return ''

    def export_queue_json(self, output_dir: str):
        """Export PENDING ENGINE_QUEUE rows to JSON for Oversight Dashboard."""
        try:
            wb = load_workbook(self.master_file, read_only=True, data_only=True)
            if 'ENGINE_QUEUE' not in wb.sheetnames:
                wb.close()
                return
            ws      = wb['ENGINE_QUEUE']
            headers = None
            rows    = []
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = list(row)
                    continue
                if not any(row):
                    continue
                d = dict(zip(headers, row))
                if str(d.get('STATUS', '')).upper() == 'PENDING':
                    rows.append({k: (str(v) if v is not None else '') for k, v in d.items()})
            wb.close()

            os.makedirs(output_dir, exist_ok=True)
            out = os.path.join(output_dir, 'engine_queue_export.json')
            with open(out, 'w', encoding='utf-8') as f:
                json.dump({'pending_count': len(rows), 'entries': rows, 'exported': datetime.now().isoformat()}, f,
                          ensure_ascii=False, indent=2)
            print(f"  Queue exported: {len(rows)} PENDING → {out}")
        except Exception as e:
            print(f"  Queue export error: {e}")

    # ── internal helpers ──────────────────────────────────────────────────────

    def _self_audit(self, entry: EntryRecord):
        """Scan entry fields for banned terms before writing."""
        fields = [entry.qur_meaning, entry.foundation_ref, entry.phonetic_chain]
        text   = ' '.join(f for f in fields if f).lower()
        for term in BANNED_TERMS:
            if term in text:
                print(f"  ⚠ SELF-AUDIT: banned term '{term}' detected in entry {entry.en_term} — please review")

    def _log_session(self, wb, entry: EntryRecord):
        """Append gate closure to SESSION_INDEX."""
        try:
            ws  = wb['SESSION_INDEX']
            nxt = ws.max_row + 1
            ts  = datetime.now().strftime('%Y-%m-%d %H:%M')
            vals = [
                'GATE CLOSURE', entry.entry_id,
                f"{entry.en_term} → {entry.root_letters} | Q+U+F | Score {entry.score}/10",
                'A1_ENTRIES', 'CONFIRMED', f"Engine auto-write {ts}"
            ]
            for ci, v in enumerate(vals, 1):
                ws.cell(row=nxt, column=ci, value=v)
        except Exception as e:
            print(f"  SESSION_INDEX log error: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# COMPONENT 8 — ReportGenerator (360-degree HTML output)
# ═══════════════════════════════════════════════════════════════════════════════

DARK_GOLD_CSS = """
:root{--gold:#C9A84C;--dark:#1a1a1a;--panel:#242424;--border:#3a3a2a;
  --text:#e8e0d0;--accent:#8B6914;--pass:#4a7c4e;--fail:#7c4a4a;}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:var(--dark);color:var(--text);font-family:Georgia,serif;
  line-height:1.7;padding:2rem;}
.report-header{border:2px solid var(--gold);padding:1.5rem;margin-bottom:2rem;
  background:var(--panel);}
.report-title{color:var(--gold);font-size:1.8rem;font-weight:bold;}
.report-meta{color:#999;font-size:.85rem;margin-top:.5rem;}
.section{border-left:3px solid var(--gold);margin-bottom:1.5rem;
  padding:1rem 1.5rem;background:#1e1e1e;}
.section-title{color:var(--gold);font-size:1.1rem;font-weight:bold;
  margin-bottom:.75rem;text-transform:uppercase;letter-spacing:.1em;
  border-bottom:1px solid var(--border);padding-bottom:.5rem;}
.att{color:#d4c4a0;font-style:italic;margin:.3rem 0;}
.arabic{font-size:1.3rem;direction:rtl;color:var(--gold);}
.chain{font-family:monospace;background:#2a2a1a;padding:.4rem .8rem;
  border-radius:3px;color:#d4c080;margin:.3rem 0;display:block;}
.gate-pass{color:#6fbf73;font-weight:bold;}
.gate-fail{color:#f44336;font-weight:bold;}
.score-badge{display:inline-block;background:var(--gold);color:var(--dark);
  padding:.2rem .6rem;border-radius:12px;font-weight:bold;font-size:.9rem;}
.dp-tag{display:inline-block;background:#3a1a1a;border:1px solid #7c4a4a;
  color:#f08080;padding:.1rem .4rem;border-radius:3px;font-size:.8rem;margin:.1rem;}
.network-tag{display:inline-block;background:#1a2a3a;border:1px solid var(--gold);
  color:var(--gold);padding:.2rem .5rem;border-radius:3px;font-size:.85rem;}
.decay-arrow{color:#888;}
table{width:100%;border-collapse:collapse;margin-top:.5rem;}
th{background:var(--accent);color:var(--dark);padding:.4rem .8rem;text-align:left;}
td{padding:.35rem .8rem;border-bottom:1px solid var(--border);}
tr:nth-child(even) td{background:#222;}
.quf-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:1rem;margin:.75rem 0;}
.gate-box{padding:.75rem;border:1px solid var(--border);text-align:center;border-radius:4px;}
.gate-box.pass{border-color:var(--pass);background:#1a2a1a;}
.gate-box.fail{border-color:var(--fail);background:#2a1a1a;}
.no-data{color:#666;font-style:italic;}
"""


class ReportGenerator:
    """
    Generates comprehensive 360-degree HTML reports from all lattice domains.
    8 sections: Linguistic | Qur'anic | Cluster | Degradation |
                Intelligence | Mathematical | Current vs Original | Open Investigations
    """

    def __init__(self, master_file: str, reports_dir: str):
        self.master_file = master_file
        self.reports_dir = reports_dir
        os.makedirs(reports_dir, exist_ok=True)
        self._cache: Dict[str, list] = {}

    # ── sheet loader ──────────────────────────────────────────────────────────

    def _load_sheet(self, name: str) -> list:
        if name in self._cache:
            return self._cache[name]
        try:
            wb = load_workbook(self.master_file, read_only=True, data_only=True)
            if name not in wb.sheetnames:
                wb.close()
                return []
            ws = wb[name]
            headers = None
            rows    = []
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(h).strip() if h else f'c{i}' for i, h in enumerate(row)]
                    continue
                if not any(row):
                    continue
                rows.append(dict(zip(headers, row)))
            wb.close()
            self._cache[name] = rows
            return rows
        except Exception as e:
            print(f"  Report: cannot load {name}: {e}")
            return []

    def _find_entry(self, en_term: str) -> dict:
        for e in self._load_sheet('A1_ENTRIES'):
            if str(e.get('EN_TERM', '')).upper() == en_term.upper():
                return e
        return {}

    def _get_derivatives(self, root_id: str) -> list:
        if not root_id:
            return []
        return [d for d in self._load_sheet('A4_DERIVATIVES')
                if str(d.get('ROOT_ID', '')).strip() == root_id.strip()]

    def _get_network_members(self, network_id: str) -> list:
        if not network_id:
            return []
        return [e for e in self._load_sheet('A1_ENTRIES')
                if str(e.get('NETWORK_ID', '')).strip() == network_id.strip()]

    def _search_consolidated(self, term: str) -> list:
        term_lower = term.lower()
        matches = []
        for row in self._load_sheet('EXCEL_DATA_CONSOLIDATED'):
            text = ' '.join(str(v) for v in row.values() if v).lower()
            if term_lower in text:
                matches.append(row)
            if len(matches) >= 8:
                break
        return matches

    # ── section builders ──────────────────────────────────────────────────────

    def _s1_linguistic(self, en_term: str, entry: dict, result: ProcessResult) -> str:
        if entry:
            ar_word    = entry.get('AR_WORD', '—')
            root_let   = entry.get('ROOT_LETTERS', '—')
            root_id    = entry.get('ROOT_ID', '—')
            score      = entry.get('SCORE', '—')
            pattern    = str(entry.get('PATTERN', 'A')).split('+')[0]
            chain      = entry.get('PHONETIC_CHAIN', '—')
            inv        = entry.get('INVERSION_TYPE', 'HIDDEN')
            qur_mean   = entry.get('QUR_MEANING', '—')
            found_ref  = entry.get('FOUNDATION_REF', '—')
            q_cls = u_cls = f_cls = 'pass'
            q_lbl = u_lbl = f_lbl = 'PASS'
        else:
            rc       = result.confirmed_root
            ar_word  = rc.ar_word if rc else '—'
            root_let = rc.letters if rc else '—'
            root_id  = '—'
            score    = result.entry_record.score if result.entry_record else '—'
            pattern  = 'A'
            chain    = rc.phonetic_chain if rc else '—'
            inv      = 'HIDDEN'
            qur_mean = '—'
            found_ref = result.f_gate.details.get('foundation_ref', '—') if result.f_gate else '—'
            q_cls = 'pass' if result.q_gate and result.q_gate.passed else 'fail'
            u_cls = 'pass' if result.u_gate and result.u_gate.passed else 'fail'
            f_cls = 'pass' if result.f_gate and result.f_gate.passed else 'fail'
            q_lbl = 'PASS' if q_cls == 'pass' else 'FAIL'
            u_lbl = 'PASS' if u_cls == 'pass' else 'FAIL'
            f_lbl = 'PASS' if f_cls == 'pass' else 'FAIL'

        pat_desc = {
            'A': "Hidden — Allah's Arabic origin invisible to English speaker",
            'B': "Weaponised — term deployed against its source population",
            'C': "Confessional — English word confesses its Qur'anic origin",
            'D': "Jāhilīan — Qur'anic weight stripped within the Arabic-speaking community"
        }.get(pattern, pattern)

        return f"""
  <div class="section">
    <div class="section-title">Section 1 — Linguistic Core</div>
    <p class="arabic">{ar_word}</p>
    <p>Root: <strong>{root_let}</strong> &nbsp; ID: {root_id} &nbsp;
       Score: <span class="score-badge">{score}/10</span></p>
    <span class="chain">{chain}</span>
    <p>Pattern: <strong>{pattern}</strong> — {pat_desc}</p>
    <p>Inversion: {inv}</p>
    <p style="color:#aaa;margin-top:.5rem;">{qur_mean}</p>
    <div class="quf-grid">
      <div class="gate-box {q_cls}"><strong>Q-Gate</strong><br>
        <span class="gate-{q_cls}">{q_lbl}</span><br>Qur'anic attestation</div>
      <div class="gate-box {u_cls}"><strong>U-Gate</strong><br>
        <span class="gate-{u_cls}">{u_lbl}</span><br>Phonetic unity</div>
      <div class="gate-box {f_cls}"><strong>F-Gate</strong><br>
        <span class="gate-{f_cls}">{f_lbl}</span><br>Foundation layer</div>
    </div>
    <p style="color:#888;font-size:.85rem;margin-top:.5rem;">{found_ref}</p>
  </div>"""

    def _s2_quranic(self, entry: dict) -> str:
        root_id = str(entry.get('ROOT_ID', '') or '').strip() if entry else ''
        refs    = [r for r in self._load_sheet('A3_QURAN_REFS')
                   if str(r.get('ROOT_ID', '')).strip() == root_id] if root_id else []

        allah_id   = str(entry.get('ALLAH_NAME_ID', '') or '').strip() if entry else ''
        allah_html = ''
        if allah_id:
            for n in self._load_sheet('A2_NAMES_OF_ALLAH'):
                if str(n.get('ALLAH_NAME_ID', '')).strip() == allah_id:
                    allah_html = (f'<p>&#127775; Name of Allah: '
                                  f'<strong>{n.get("ALLAH_NAME","")}'
                                  f' / {n.get("TRANSLITERATION","")}'
                                  f' / {n.get("MEANING","")}</strong> ({allah_id})</p>')
                    break

        if refs:
            rows_html = ''.join(
                f'<tr><td>{r.get("VERSE_REF","")}</td>'
                f'<td class="arabic" style="font-size:1.1rem;">{r.get("AR_TEXT","")}</td>'
                f'<td><em>{r.get("TRANSLITERATION","")}</em></td>'
                f'<td>{r.get("MEANING","")}</td></tr>'
                for r in refs[:10]
            )
        else:
            rows_html = '<tr><td colspan="4" class="no-data">No Qur\'anic references indexed for this root yet</td></tr>'

        return f"""
  <div class="section">
    <div class="section-title">Section 2 — Qur'anic Context</div>
    {allah_html}
    <table>
      <tr><th>Verse</th><th>Arabic</th><th>Transliteration</th><th>Meaning</th></tr>
      {rows_html}
    </table>
  </div>"""

    def _s3_cluster(self, entry: dict) -> str:
        net_id    = str(entry.get('NETWORK_ID', '') or '').strip() if entry else ''
        root_id   = str(entry.get('ROOT_ID', '') or '').strip() if entry else ''
        net_title = ''

        if net_id:
            for n in self._load_sheet('M4_NETWORKS'):
                if str(n.get('NETWORK_ID', '')).strip() == net_id:
                    net_title = f"{net_id}: {n.get('TITLE','')} — {n.get('LINK_VERSE','')}"
                    break

        members = self._get_network_members(net_id)
        mem_html = ''.join(
            f'<tr><td>#{m.get("ENTRY_ID","")}</td><td>{m.get("EN_TERM","")}</td>'
            f'<td class="arabic">{m.get("AR_WORD","")}</td><td>{m.get("ROOT_LETTERS","")}</td></tr>'
            for m in members
        ) or '<tr><td colspan="4" class="no-data">No network membership — new network candidate</td></tr>'

        derivs     = self._get_derivatives(root_id)
        deriv_list = ', '.join(str(d.get('EN_DERIVATIVE', '')) for d in derivs[:25])
        deriv_ct   = len(derivs)

        return f"""
  <div class="section">
    <div class="section-title">Section 3 — Cluster / Network</div>
    <p><span class="network-tag">{net_title or 'No network assigned'}</span></p>
    <table style="margin-top:.75rem;">
      <tr><th>#</th><th>Term</th><th>Arabic</th><th>Root</th></tr>
      {mem_html}
    </table>
    <p style="margin-top:.75rem;"><strong>Derivatives ({deriv_ct}+):</strong>
      {deriv_list or '<span class="no-data">None indexed yet</span>'}</p>
  </div>"""

    def _s4_degradation(self, entry: dict, result: ProcessResult) -> str:
        ar_word   = entry.get('AR_WORD', '—') if entry else '—'
        found_ref = entry.get('FOUNDATION_REF', '') if entry else (
            result.f_gate.details.get('foundation_ref', '') if result.f_gate else '')
        ds_match  = re.search(r'DS\d+', str(found_ref))
        ds_code   = ds_match.group(0) if ds_match else 'DS05'
        ds_desc   = {
            'DS04': '"Greek" — Stage 3 decay',
            'DS05': '"Latin" — Stage 3-4 decay',
            'DS06': 'Germanic — Stage 4-5 decay',
            'DS04→DS05': '"Greek" → "Latin" corridor',
        }.get(ds_code, f'{ds_code} — downstream corridor')

        return f"""
  <div class="section">
    <div class="section-title">Section 4 — Degradation Trail</div>
    <p>
      <span class="arabic">{ar_word}</span> (Allah's Arabic — ORIG1)
      <span class="decay-arrow"> &#8594; </span>{ds_desc}
      <span class="decay-arrow"> &#8594; </span>{result.input_term.upper()} (current downstream form)
    </p>
    <p style="color:#888;font-size:.85rem;margin-top:.5rem;">{found_ref}</p>
    <p style="margin-top:.5rem;color:#aaa;">
      Direction of flow: Allah's Arabic is ALWAYS the source.
      All downstream forms are degradations — never origins.
    </p>
  </div>"""

    def _s5_intelligence(self, en_term: str, root_letters: str) -> str:
        matches = self._search_consolidated(en_term)
        if not matches:
            matches = self._search_consolidated(root_letters.replace('-', ''))

        if matches:
            rows_html = ''.join(
                f'<tr><td>{ " | ".join(str(v)[:100] for v in list(m.values())[:4] if v) }</td></tr>'
                for m in matches[:5]
            )
        else:
            rows_html = '<tr><td class="no-data">No mentions found in EXCEL_DATA_CONSOLIDATED for this term</td></tr>'

        return f"""
  <div class="section">
    <div class="section-title">Section 5 — Historical &amp; Intelligence</div>
    <table>
      <tr><th>EXCEL_DATA_CONSOLIDATED search (top 5 matches)</th></tr>
      {rows_html}
    </table>
    <p style="color:#666;font-size:.8rem;margin-top:.5rem;">
      For full operator network data check: Historic Lattice / Updated Intelligence /
    </p>
  </div>"""

    def _s6_mathematical(self, root_letters: str) -> str:
        return f"""
  <div class="section">
    <div class="section-title">Section 6 — Mathematical / Ratio</div>
    <p class="no-data">
      Ratio/formula connections for root {root_letters or '—'} —
      check A1_ENTRIES for F-series entries sharing this root, or
      submit a ratio query (e.g. "4/3") for formula domain analysis.
    </p>
  </div>"""

    def _s7_current_vs_original(self, en_term: str, entry: dict) -> str:
        pattern  = str(entry.get('PATTERN', 'A')).split('+')[0] if entry else 'A'
        qur_mean = entry.get('QUR_MEANING', '—') if entry else '—'
        gap_desc = {
            'A': 'HIDDEN — complete etymological amnesia in the downstream speaker',
            'B': 'WEAPONISED — term deployed as attack tool against the source civilisation',
            'C': 'CONFESSIONAL — the downstream form confesses its own Qur\'anic origin in its sounds',
            'D': 'JĀHILĪAN — Qur\'anic weight stripped within the Arabic-speaking community itself'
        }.get(pattern, pattern)

        return f"""
  <div class="section">
    <div class="section-title">Section 7 — Current Usage vs. Original Meaning</div>
    <table>
      <tr><th>Dimension</th><th>Content</th></tr>
      <tr><td>Current English usage</td><td>{en_term.lower()}</td></tr>
      <tr><td>Qur'anic original meaning</td><td>{qur_mean}</td></tr>
      <tr><td>Classification</td><td>{gap_desc}</td></tr>
      <tr><td>Decay summary</td>
          <td>The downstream form preserves the consonant skeleton but has lost the Qur'anic semantic weight</td></tr>
    </table>
  </div>"""

    def _s8_open_investigations(self, root_letters: str) -> str:
        pending = [q for q in self._load_sheet('ENGINE_QUEUE')
                   if str(q.get('STATUS', '')).upper() == 'PENDING'
                   and str(q.get('CANDIDATE_ROOT', '')).strip() == root_letters.strip()]

        if pending:
            rows_html = ''.join(
                f'<tr><td>{p.get("QUEUE_ID","")}</td><td>{p.get("INPUT_TERM","")}</td>'
                f'<td>{p.get("FLAG_REASON","")}</td><td>PENDING</td></tr>'
                for p in pending
            )
        else:
            rows_html = '<tr><td colspan="4" class="no-data">No pending queue entries for this root</td></tr>'

        return f"""
  <div class="section">
    <div class="section-title">Section 8 — Open Investigations</div>
    <table>
      <tr><th>Queue ID</th><th>Term</th><th>Flag Reason</th><th>Status</th></tr>
      {rows_html}
    </table>
    <p style="color:#666;font-size:.8rem;margin-top:.5rem;">
      Open USLaP_Oversight_Dashboard.html to approve/reject pending entries.
    </p>
  </div>"""

    # ── main generate method ──────────────────────────────────────────────────

    def generate(self, result: ProcessResult) -> str:
        """Generate full 360-degree HTML report. Returns file path."""
        en_term = result.input_term.upper()
        entry   = self._find_entry(en_term)
        root_letters = (entry.get('ROOT_LETTERS', '') if entry
                        else (result.confirmed_root.letters if result.confirmed_root else ''))

        s1 = self._s1_linguistic(en_term, entry, result)
        s2 = self._s2_quranic(entry)
        s3 = self._s3_cluster(entry)
        s4 = self._s4_degradation(entry, result)
        s5 = self._s5_intelligence(en_term, root_letters)
        s6 = self._s6_mathematical(root_letters)
        s7 = self._s7_current_vs_original(en_term, entry)
        s8 = self._s8_open_investigations(root_letters)

        date_str  = datetime.now().strftime('%Y-%m-%d')
        entry_id  = entry.get('ENTRY_ID', '—') if entry else '—'

        html = f"""<!DOCTYPE html>
<html lang="en" dir="ltr">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>USLaP 360 Report — {en_term}</title>
  <style>{DARK_GOLD_CSS}</style>
</head>
<body>
  <div class="report-header">
    <div class="report-title">USLaP 360&#176; REPORT — {en_term}</div>
    <div class="report-meta">
      Generated: {date_str} &nbsp;|&nbsp;
      Entry: #{entry_id} &nbsp;|&nbsp;
      Root: {root_letters or '—'} &nbsp;|&nbsp;
      Query type: {result.input_type}
    </div>
  </div>
  {s1}{s2}{s3}{s4}{s5}{s6}{s7}{s8}
  <div style="margin-top:2rem;color:#555;font-size:.75rem;border-top:1px solid #333;padding-top:1rem;">
    USLaP Unified Linguistic Lattice — Q14:24: &#x623;&#x635;&#x644;&#x647;&#x627; &#x62B;&#x627;&#x628;&#x62A; &#x648;&#x641;&#x631;&#x639;&#x647;&#x627; &#x641;&#x64A; &#x627;&#x644;&#x633;&#x645;&#x627;&#x621;
  </div>
</body>
</html>"""

        # Preserve Arabic chars in filename; replace only non-word chars
        fname = re.sub(r'[^\w]', '_', en_term, flags=re.UNICODE).strip('_') or 'REPORT'
        fpath = os.path.join(self.reports_dir, f"{fname}_360_Report_{date_str.replace('-','')}.html")
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"  Report saved: {fpath}")
        return fpath


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN ORCHESTRATOR — USLaPEngine
# ═══════════════════════════════════════════════════════════════════════════════

class USLaPEngine:
    """
    Coordinates all 8 components.
    Single public entry point: engine.process(input_term)
    """

    def __init__(self, master_file: str = MASTER_FILE,
                 quran_file: str = QURAN_FILE,
                 reports_dir: str = REPORTS_DIR,
                 skip_reports: bool = False):
        print("\n" + "═" * 62)
        print("  USLaP Autonomous Engine v3.3 (cognate-crossref+parallel-ORIG2+depal-competition)")
        print("  بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ")
        print("═" * 62)
        self.skip_reports = skip_reports

        for path in (master_file, quran_file):
            if not os.path.exists(path):
                raise FileNotFoundError(f"Required file not found: {path}")

        print("\nLoading components...")
        self.master_file = master_file
        self.router      = InputRouter()
        self.q_gate      = QGate(quran_file)
        # English components
        self.reversal    = PhoneticReversal(master_file, self.q_gate)
        self.u_gate      = UGate(self.reversal)
        # Russian components (v3.0)
        self.ru_reversal = RussianPhoneticReversal(master_file, self.q_gate)
        self.ru_u_gate   = UGate(self.ru_reversal)   # UGate parameterised by reversal
        # Shared components
        self.f_gate      = FGate(master_file)
        self.scorer      = Scorer()
        self.writer      = EntryWriter(master_file)
        self.reporter    = ReportGenerator(master_file, reports_dir)
        # ORIG2 track (v2.2): Kashgari corpus indexer + gate
        kashgari_path    = KASHGARI_FILE
        if os.path.exists(kashgari_path):
            self.kashgari_index = KashgariIndex(kashgari_path)
            self.kashgari_gate  = KashgariGate(self.kashgari_index)
        else:
            print(f"  WARNING: Kashgari corpus not found at {kashgari_path} — ORIG2 track disabled")
            self.kashgari_index = None
            self.kashgari_gate  = None
        self.existing_terms: Dict[str, int] = {}
        self.existing_ru_terms: Dict[str, int] = {}
        self._build_dedup_cache()
        self.expander    = ClusterExpander(self.reversal, self.existing_terms)
        print("\n✓ Engine ready (EN + RU).\n" + "═" * 62)

    def _build_dedup_cache(self):
        try:
            wb = load_workbook(self.master_file, read_only=True, data_only=True)
            # English entries
            ws = wb['A1_ENTRIES']
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not any(row):
                    continue
                d   = dict(zip(headers, row))
                et  = str(d.get('EN_TERM', '') or '').strip().upper()
                eid = d.get('ENTRY_ID')
                if et and eid:
                    self.existing_terms[et] = int(eid)
            # Russian entries (v3.0)
            ws_ru = wb['A1_ЗАПИСИ']
            ru_headers = None
            for row in ws_ru.iter_rows(values_only=True):
                if ru_headers is None:
                    ru_headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not any(row):
                    continue
                d = dict(zip(ru_headers, row))
                rt = str(d.get('РУС_ТЕРМИН', '') or '').strip().upper()
                rid = d.get('ЗАПИСЬ_ID')
                if rt and rid:
                    self.existing_ru_terms[rt] = int(rid)
            wb.close()
            print(f"  Dedup cache: {len(self.existing_terms)} EN terms, "
                  f"{len(self.existing_ru_terms)} RU terms")
        except Exception as e:
            print(f"  Dedup cache error: {e}")

    # ── public process method ─────────────────────────────────────────────────

    def process(self, raw_input: str, dry_run: bool = False) -> ProcessResult:
        """Full pipeline. Returns ProcessResult."""
        input_type, cleaned, key_terms = self.router.detect(raw_input)

        # v3.1: Latin-script Russian detection
        # If detected as English but looks like transliterated Russian, convert
        if input_type == 'english_word' and self.ru_reversal._is_latin_russian(raw_input):
            cyrillic = self.ru_reversal.transliterate_latin(raw_input)
            print(f"  [v3.1] Latin-Russian detected: '{raw_input}' → '{cyrillic}'")
            input_type = 'russian_word'
            cleaned = cyrillic
            key_terms = [cyrillic]

        print(f"\n{'─'*62}")
        print(f"  Processing: '{raw_input}'")
        print(f"  Type detected: {input_type}")
        print(f"{'─'*62}")

        if input_type == 'phrase':
            return self._process_phrase(raw_input, key_terms, dry_run)
        if input_type == 'russian_phrase':
            # Process each Russian word individually
            results = []
            for t in key_terms:
                results.append(self._process_single_term(t, 'russian_word', dry_run))
            return results[-1] if results else ProcessResult(raw_input, 'russian_phrase')
        if input_type == 'ratio':
            r = ProcessResult(raw_input, 'ratio')
            r.add_log("Ratio input — hand off to formula domain (run forest_v3 GeneratorEngine)")
            return r
        if input_type == 'quran_ref':
            return self._process_quran_ref(cleaned)

        term = key_terms[0] if key_terms else cleaned
        return self._process_single_term(term, input_type, dry_run)

    # ── single-term pipeline ──────────────────────────────────────────────────

    def _process_single_term(self, term: str, input_type: str,
                             dry_run: bool) -> ProcessResult:
        result = ProcessResult(term, input_type)

        # ── LANGUAGE ROUTING (v3.0) ────────────────────────────────────────────
        is_russian = input_type == 'russian_word'
        active_reversal = self.ru_reversal if is_russian else self.reversal
        active_u_gate   = self.ru_u_gate   if is_russian else self.u_gate
        lang_label      = 'RU' if is_russian else 'EN'

        # DEDUP CHECK
        if input_type == 'english_word':
            existing_id = self.existing_terms.get(term.upper())
            if existing_id:
                result.add_log(f"EXISTING ENTRY: {term.upper()} = #{existing_id} (already in lattice)")
                result.existing_entry_id = existing_id
                if not self.skip_reports:
                    result.report_path = self.reporter.generate(result)
                return result
        elif is_russian:
            existing_id = self.existing_ru_terms.get(term.upper())
            if existing_id:
                result.add_log(f"EXISTING RU ENTRY: {term.upper()} = #{existing_id} (already in lattice)")
                result.existing_entry_id = existing_id
                if not self.skip_reports:
                    result.report_path = self.reporter.generate(result)
                return result

        # DERIVATIVE CHAIN CHECK (Gate 3f) — English only for now
        if input_type == 'english_word':
            parent = KNOWN_DERIVATIVES.get(term.lower())
            if parent:
                parent_id = self.existing_terms.get(parent.upper())
                if parent_id:
                    result.add_log(
                        f"DERIVATIVE CHAIN: '{term}' is derivative of {parent} (#{parent_id}). "
                        f"Route to A4_DERIVATIVES, not A1_ENTRIES."
                    )
                    result.derivative_of = (parent, parent_id)
                else:
                    result.add_log(
                        f"DERIVATIVE CHAIN: '{term}' maps to parent '{parent}' but parent "
                        f"not yet in lattice. Processing as independent candidate."
                    )

        # ═══ v3.3: COGNATE CROSS-REFERENCE ═══════════════════════════════════
        # Sibling Database Principle automated: when processing a Russian word,
        # check if there's an English cognate.  Process the ENGLISH form through
        # the English pipeline — it often preserves root consonants better.
        # Example: ДЕСАНТ lost the K from سَكَنَ, but DESCENT preserves it as
        # S-C-N-D → root س-ك-ن at score 9+.
        cognate_result = None
        if input_type in ('english_word', 'russian_word'):
            cognate_result = self._try_cognate_crossref(term, is_russian)
            if cognate_result:
                src = cognate_result.get('source', '')
                if src == 'LATTICE_ENTRY':
                    result.add_log(
                        f"v3.3 COGNATE: {cognate_result.get('note', '')}"
                    )
                elif src == 'EN_PIPELINE':
                    result.add_log(
                        f"v3.3 COGNATE: English cousin '{cognate_result['en_cousin']}' "
                        f"→ root {cognate_result['root_letters']} "
                        f"({cognate_result['token_count']} tok, "
                        f"EN score={cognate_result['score']}/10, "
                        f"chain: {cognate_result.get('phonetic_chain', '—')})"
                    )
                elif src == 'ERROR':
                    result.add_log(
                        f"v3.3 COGNATE: English cousin '{cognate_result['en_cousin']}' "
                        f"pipeline error: {cognate_result.get('error', '?')}"
                    )
                    cognate_result = None  # discard errors
            result.cognate_crossref = cognate_result

        # REVERSAL (word → root candidates) — uses language-specific reversal
        if input_type in ('english_word', 'russian_word'):
            result.add_log(f"Running {lang_label} phonetic reversal for '{term}'...")
            candidates = active_reversal.reverse(term)
            result.root_candidates = candidates
            if not candidates:
                # ══ ORIG2 TRACK (v2.2) ══════════════════════════════════════════
                # Q-Gate failed for all candidates → try Kashgari/ORIG2 track
                # Protocol: ROOT_LIST FAIL → Kashgari search → BITIG scoring
                result.add_log("No ORIG1 root found — routing to ORIG2 (Kashgari) track...")
                # v3.3: If cognate found a root, try it before giving up
                if cognate_result and cognate_result.get('source') == 'EN_PIPELINE':
                    result.add_log(
                        f"v3.3 COGNATE RESCUE: No {lang_label} candidates, but "
                        f"English cousin '{cognate_result['en_cousin']}' found "
                        f"root {cognate_result['root_letters']} — using cognate root."
                    )
                    # Inject the cognate root as the primary candidate
                    cog_cand = RootCandidate(
                        letters=cognate_result['root_letters'],
                        token_count=cognate_result['token_count'],
                        ar_word=cognate_result.get('ar_word', '')
                    )
                    cog_cand.phonetic_chain = cognate_result.get('phonetic_chain', '')
                    cog_cand.positional_score = cognate_result.get('positional', 0.5)
                    cog_cand.operations = cognate_result.get('operations', [])
                    candidates = [cog_cand]
                    result.root_candidates = candidates
                    # Continue to normal scoring below
                else:
                    orig2_result = self._try_orig2_track(term, result, dry_run, is_russian=is_russian)
                    if orig2_result is not None:
                        return orig2_result
                    # Both ORIG1 and ORIG2 failed
                    result.add_log("Neither ORIG1 nor ORIG2 attested — entry cannot enter lattice")
                    return result
                # ═════════════════════════════════════════════════════════════════
            # ═══ v2.5 MULTI-CANDIDATE SCORING ═══════════════════════════════
            # Score top N candidates (up to 3) through full Q→U→F→Scorer pipeline.
            # Pick the candidate with the HIGHEST final score.
            # This prevents the coverage penalty from penalising a root that
            # a lower-ranked candidate (with fewer extra_consonants) would
            # score higher on.  COMMAND: ح-م-د (extra=2, score=6) vs
            # س-م-ن via OP_STOP (extra=0, score=8) — the OP_STOP candidate
            # is in the list but was never scored in v2.4.
            # ════════════════════════════════════════════════════════════════
            # v2.5: Score top N candidates. With tier-diverse candidate pools,
            # we may have >3 candidates representing different extra_consonants
            # tiers.  Score up to 5 to ensure each tier gets evaluated.
            MULTI_N = min(5, len(candidates))
            best_score  = -1
            best_top    = candidates[0]
            best_q      = None
            best_u      = None
            best_f      = None
            best_bdown  = {}

            for ci in range(MULTI_N):
                cand = candidates[ci]
                cq = self.q_gate.check(cand.letters)
                if not cq.passed:
                    continue
                cu = active_u_gate.verify(term, cand.letters, cand.operations)
                cf = self.f_gate.assign(term, cand.letters,
                                        cu.details.get('phonetic_chain', ''))
                cs, cb = self.scorer.score(cand, term, cq, cu, cf)
                if ci == 0:
                    result.add_log(
                        f"Candidate #{ci+1}: {cand.letters} "
                        f"({cand.token_count} tok, extra={getattr(cand,'extra_consonants',0)}) "
                        f"→ score={cs}")
                if cs > best_score:
                    best_score = cs
                    best_top   = cand
                    best_q     = cq
                    best_u     = cu
                    best_f     = cf
                    best_bdown = cb
                    if ci > 0:
                        result.add_log(
                            f"Candidate #{ci+1}: {cand.letters} "
                            f"({cand.token_count} tok, extra={getattr(cand,'extra_consonants',0)}) "
                            f"→ score={cs} ← BETTER than #{1}")

            top      = best_top
            q_result = best_q
            u_result = best_u
            f_result = best_f
            score    = best_score
            breakdown = best_bdown
            result.add_log(f"Selected: {top.letters} ({top.token_count} tokens, score={score})")

            # ═══ v3.2: PARALLEL ORIG2 CHECK ═══════════════════════════════════
            # Run Kashgari check ALONGSIDE ORIG1 (not just as fallback).
            # If ORIG2 finds a hit while ORIG1 also passed → COMPETING TRACKS.
            # Critical for Russian where >50% of vocabulary is Bitig-corridor.
            if self.kashgari_gate is not None and score >= SCORE_QUEUE:
                orig2_par = self._try_orig2_parallel(term, is_russian)
                if orig2_par and orig2_par.get('passed'):
                    o2 = orig2_par
                    result.add_log(
                        f"v3.2 PARALLEL ORIG2 HIT: Kashgari '{o2['kashgari_translit']}' "
                        f"= \"{o2['kashgari_meaning']}\" "
                        f"(line {o2['kashgari_line']}, {o2['attestation_type']}, "
                        f"ORIG2 score={o2['orig2_score']}/10, "
                        f"searched='{o2.get('search_skeleton', '?')}')"
                    )
                    result.orig2_parallel = orig2_par

            # ═══ v3.2: DEPALATALIZATION COMPETITION ═══════════════════════════
            # If both palatalized AND depalatalized candidates exist with
            # different roots and both pass Q-gate → flag for human review.
            # Example: ВОЖДЬ → و-ج-د (via ж→ج) vs و-د-د (via ж→д depal)
            depal_competitor = None
            if is_russian and score >= SCORE_QUEUE:
                for ci in range(min(5, len(candidates))):
                    cand = candidates[ci]
                    cand_ops = getattr(cand, 'operations', [])
                    if (any('OP_DEPALATAL' in str(op) for op in cand_ops)
                            and cand.letters != top.letters):
                        depal_competitor = cand
                        result.add_log(
                            f"v3.2 DEPAL COMPETITION: primary {top.letters} "
                            f"({top.token_count} tok) vs depalatalized "
                            f"{cand.letters} ({cand.token_count} tok, "
                            f"ops={cand_ops})"
                        )
                        result.competing_depal = cand
                        break

            # ═══ v3.3/v3.2: LOW SCORE → COGNATE RESCUE then ORIG2 FALLBACK ═══
            # If ORIG1 candidates scored ≤ SCORE_REJECT:
            #   1. Check if cognate found a better root (v3.3)
            #   2. If not, try Kashgari ORIG2 (v3.2)
            if score <= SCORE_REJECT:
                # v3.3: COGNATE RESCUE — if cognate scored higher, use its root
                if (cognate_result and cognate_result.get('source') == 'EN_PIPELINE'
                        and cognate_result.get('score', 0) > score):
                    cog = cognate_result
                    result.add_log(
                        f"v3.3 COGNATE RESCUE: {lang_label} score {score} too low, "
                        f"but EN cousin '{cog['en_cousin']}' → {cog['root_letters']} "
                        f"(EN score {cog['score']}/10, {cog['token_count']} tok). "
                        f"Using cognate root."
                    )
                    # Replace the top candidate with the cognate root
                    cog_cand = RootCandidate(
                        letters=cog['root_letters'],
                        token_count=cog['token_count'],
                        ar_word=cog.get('ar_word', '')
                    )
                    cog_cand.phonetic_chain = cog.get('phonetic_chain', '')
                    cog_cand.positional_score = cog.get('positional', 0.5)
                    cog_cand.operations = cog.get('operations', [])
                    # Re-score through local Q→U gates
                    cq = self.q_gate.check(cog_cand.letters)
                    if cq.passed:
                        cu = active_u_gate.verify(term, cog_cand.letters,
                                                  cog_cand.operations)
                        cf = self.f_gate.assign(term, cog_cand.letters,
                                                cu.details.get('phonetic_chain', ''))
                        cs, cb = self.scorer.score(cog_cand, term, cq, cu, cf)
                        # Accept if cognate root scores better locally too
                        if cs > score:
                            top = cog_cand
                            q_result = cq
                            u_result = cu
                            f_result = cf
                            score = cs
                            breakdown = cb
                            result.add_log(
                                f"v3.3 COGNATE RESCUE SUCCESS: local re-score "
                                f"{cog_cand.letters} = {cs}/10 (was {best_score})")
                        else:
                            result.add_log(
                                f"v3.3 COGNATE: EN root {cog_cand.letters} scored "
                                f"{cs} locally — not better than {score}. "
                                f"Keeping original, falling to ORIG2.")

                # v3.2: ORIG2 FALLBACK — if still low, try Kashgari
                if score <= SCORE_REJECT and self.kashgari_gate is not None:
                    result.add_log(
                        f"ORIG1 score {score} <= {SCORE_REJECT} — trying ORIG2 fallback...")
                    orig2_fallback = self._try_orig2_track(
                        term, result, dry_run, is_russian=is_russian)
                    if orig2_fallback is not None:
                        return orig2_fallback
                    result.add_log("ORIG2 fallback: no Kashgari match either")

            # ═════════════════════════════════════════════════════════════════
        else:
            # Direct Arabic root
            result.add_log(f"Direct root input: {term}")
            qr = self.q_gate.check(term)
            if not qr.passed:
                result.add_log(f"Q-Gate FAIL: {term} not in Qur'anic root list")
                return result
            top = RootCandidate(letters=term,
                                token_count=qr.details.get('token_count', 0),
                                ar_word=qr.details.get('ar_word', ''))
            q_result = qr
            u_result = GateResult(True, {'phonetic_chain': '', 'note': 'Direct root — U-Gate N/A'})
            f_result = self.f_gate.assign(term, top.letters, '')
            score, breakdown = self.scorer.score(top, term, q_result, u_result, f_result)

        result.confirmed_root = top

        # Store gate results
        result.q_gate = q_result
        result.add_log(f"Q-Gate: {'PASS' if q_result.passed else 'FAIL'} | tokens={q_result.details.get('token_count',0)}")

        result.u_gate = u_result
        top.phonetic_chain = u_result.details.get('phonetic_chain', '')
        result.add_log(f"U-Gate: {'PASS' if u_result.passed else 'FAIL'} | chain: {top.phonetic_chain}")

        result.f_gate = f_result
        result.add_log(f"F-Gate: PASS | DS: {f_result.details.get('ds_code','—')} "
                       f"| Network: {f_result.details.get('network_id','none')}")

        # Score already computed in multi-candidate loop (or direct root path)
        top.score = score
        result.add_log(f"Score: {score}/10 | {breakdown}")

        # ═══ v3.4: COMPOUND PARTS ANALYSIS ═══════════════════════════════════════
        # For Russian compound words (САМО+ВАР, ПАРО+ВОЗ), trace both parts
        # INDEPENDENTLY for dual-root reporting.  Diagnostic only — the primary
        # result (above) stands unchanged.  Human reviewer sees both root traces.
        if is_russian and self.ru_reversal is not None:
            is_cmpd, pfx_part, root_part, bridge, cmpd_label = \
                self.ru_reversal.detect_compound(term)
            if is_cmpd:
                result.add_log(f"v3.4 COMPOUND: {cmpd_label}")
                compound_analysis = {
                    'label': cmpd_label, 'prefix': None,
                    'root': None, 'bridge': bridge
                }

                # ── Trace PREFIX part (e.g. САМО) ────────────────────────
                # v3.4b: САМО/САМА = pronoun "self/auto" — NOT a root.
                # Do NOT trace. Only trace the ROOT part.
                PRONOUN_PREFIXES = {'само', 'сам', 'сама'}
                if pfx_part.lower() in PRONOUN_PREFIXES:
                    compound_analysis['prefix'] = {
                        'part': pfx_part.upper(),
                        'root': None,
                        'token_count': 0,
                        'chain': 'PRONOUN (self/auto) — not traced',
                        'is_pronoun': True,
                    }
                    result.add_log(
                        f"  PREFIX '{pfx_part.upper()}' = pronoun 'self/auto' "
                        f"(R13: not a root, not traced)")
                else:
                    pfx_cands = self.ru_reversal._reverse_inner(pfx_part)
                    for pc in pfx_cands[:3]:
                        pq = self.q_gate.check(pc.letters)
                        if pq.passed:
                            pu = self.ru_u_gate.verify(
                                pfx_part, pc.letters, pc.operations or [])
                            compound_analysis['prefix'] = {
                                'part': pfx_part.upper(),
                                'root': pc.letters,
                                'token_count': pc.token_count,
                                'chain': pu.details.get('phonetic_chain', ''),
                            }
                            result.add_log(
                                f"  PREFIX '{pfx_part.upper()}' → {pc.letters} "
                                f"({pc.token_count} tok)")
                            break
                    if not compound_analysis.get('prefix'):
                        result.add_log(
                            f"  PREFIX '{pfx_part.upper()}' → no ORIG1 trilateral root "
                            f"(2-consonant prefix, or may be ORIG2)")

                # ── Trace ROOT part (e.g. ВАР) ──────────────────────────
                root_cands = self.ru_reversal._reverse_inner(root_part)
                for rc in root_cands[:3]:
                    rq = self.q_gate.check(rc.letters)
                    if rq.passed:
                        ru_check = self.ru_u_gate.verify(
                            root_part, rc.letters, rc.operations or [])
                        compound_analysis['root'] = {
                            'part': root_part.upper(),
                            'root': rc.letters,
                            'token_count': rc.token_count,
                            'chain': ru_check.details.get('phonetic_chain', ''),
                        }
                        result.add_log(
                            f"  ROOT '{root_part.upper()}' → {rc.letters} "
                            f"({rc.token_count} tok)")
                        break
                if not compound_analysis['root']:
                    # Try ORIG2 for root part
                    if self.kashgari_gate is not None:
                        root_cons = self.ru_reversal.extract_consonants(root_part)
                        if root_cons:
                            latin_skel = self.ru_reversal.to_latin_skeleton(root_cons)
                            k_result_cmp = self.kashgari_gate.check(
                                root_part, list(latin_skel))
                            if k_result_cmp.passed:
                                kd_cmp = k_result_cmp.details
                                compound_analysis['root'] = {
                                    'part': root_part.upper(),
                                    'root': kd_cmp.get('skeleton', latin_skel),
                                    'token_count': 0,
                                    'chain': f"ORIG2: {kd_cmp.get('kashgari_translit', '')}",
                                    'is_orig2': True,
                                    'kashgari_meaning': kd_cmp.get('kashgari_meaning', ''),
                                }
                                result.add_log(
                                    f"  ROOT '{root_part.upper()}' → ORIG2 Kashgari: "
                                    f"'{kd_cmp.get('kashgari_translit', '')}' "
                                    f"= \"{kd_cmp.get('kashgari_meaning', '')}\"")
                    if not compound_analysis.get('root'):
                        result.add_log(
                            f"  ROOT '{root_part.upper()}' → no ORIG1/ORIG2 root found")

                result.compound_parts = compound_analysis

        # ═══ v3.4: SEMANTIC REVIEW FLAG ═══════════════════════════════════════════
        # Engine checks phonetics only — NEVER semantics.  Flag ALL entries for
        # semantic review.  Like DEPAL flag: informational, does NOT change score.
        # Prevents false confidence from high phonetic scores with no semantic path
        # (e.g. САБЛЯ score=9 but سَبِيل=way/path ≠ saber/sword).
        result.sem_review = True

        # BUILD ENTRY RECORD
        entry = EntryRecord()
        entry.en_term        = term.upper()
        entry.ar_word        = top.ar_word or top.letters
        entry.root_letters   = top.letters
        entry.phonetic_chain = top.phonetic_chain
        entry.score          = score
        entry.pattern        = 'A'
        entry.network_id     = f_result.details.get('network_id', '')
        entry.foundation_ref = f_result.details.get('foundation_ref', '')
        entry.inversion_type = 'HIDDEN'
        result.entry_record  = entry

        # CLUSTER EXPANSION (for direct root input or high-score words)
        if input_type == 'arabic_root' or score >= SCORE_AUTO_WRITE:
            result.add_log("Running cluster expansion...")
            cluster_candidates = self.expander.expand(top.letters, term)
            result.cluster_members = cluster_candidates
            result.add_log(f"Cluster: {len(cluster_candidates)} candidate(s) found")

        # WRITE DECISION
        # v2.1 tightened: CONFIRMED_HIGH requires score>=9 AND Q-pass AND U-pass
        # AND positional_score >= 0.8 (no transposition). Previously score>=9 alone
        # was sufficient — this produced 33.5% CONFIRMED_HIGH vs target 15-25%.
        # v3.2: competition flags (ORIG2 parallel, depal) force PENDING_REVIEW.
        # v3.3: cognate cross-ref competition flag added.
        pos_ok = getattr(top, 'positional_score', 0.5) >= 0.8
        is_derivative = hasattr(result, 'derivative_of') and result.derivative_of
        has_competing_orig2 = hasattr(result, 'orig2_parallel') and result.orig2_parallel
        has_competing_depal = hasattr(result, 'competing_depal') and result.competing_depal

        # v3.3/v3.4: COGNATE CROSS-REFERENCE EVALUATION
        # v3.4 DIRECTION OF TRANSMISSION:
        #   Russian is closer to BOTH originals than English.
        #   EN cognate agreement = CONFIRMATION (both degraded forms converge).
        #   EN cognate disagreement = NOTE only (English is more degraded).
        #   Exception: MODERN_TECH_TERMS where EN→RU direction applies.
        cognate_ref = getattr(result, 'cognate_crossref', None)
        has_cognate_competition = False
        cognate_agrees = False
        cognate_note_only = False  # v3.4: downgraded competition
        if cognate_ref and cognate_ref.get('source') == 'EN_PIPELINE':
            cog_root = cognate_ref.get('root_letters', '')
            local_root = top.letters if top else ''
            if cog_root and local_root and cog_root != local_root:
                cog_score = cognate_ref.get('score', 0)
                if cog_score >= SCORE_QUEUE:
                    # v3.4: Check direction of transmission
                    is_modern_term = (is_russian and
                                     term.upper() in MODERN_TECH_TERMS)
                    if is_russian and not is_modern_term:
                        # RU is closer to originals — EN disagreement is
                        # informational only, does NOT force PENDING_REVIEW
                        cognate_note_only = True
                        result.add_log(
                            f"v3.4 COGNATE NOTE (not competition): "
                            f"{lang_label} → {local_root} (score {score}) vs "
                            f"EN cousin '{cognate_ref['en_cousin']}' → {cog_root} "
                            f"(EN score {cog_score}). EN is more degraded — "
                            f"RU pipeline authoritative."
                        )
                    else:
                        # EN word or modern tech term — full competition
                        has_cognate_competition = True
                        result.add_log(
                            f"v3.3 COGNATE COMPETITION: {lang_label} pipeline "
                            f"→ {local_root} (score {score}) vs EN cousin "
                            f"'{cognate_ref['en_cousin']}' → {cog_root} "
                            f"(score {cog_score})"
                        )
            elif cog_root and local_root and cog_root == local_root:
                cognate_agrees = True
                result.add_log(
                    f"v3.3 COGNATE AGREES: both {lang_label} and EN cousin "
                    f"'{cognate_ref['en_cousin']}' → {cog_root} ✓"
                )

        if not dry_run:
            # ── v3.4: Build common flag suffixes ─────────────────────────
            # SEM_REVIEW: appended to ALL queue entries (engine never checks
            # semantics — every entry needs semantic verification by human)
            sem_suffix = " | SEM_REVIEW: semantic pathway not verified"

            # COMPOUND suffix: if compound detected, show both root traces
            compound_suffix = ''
            cp = getattr(result, 'compound_parts', None)
            if cp:
                cp_parts = []
                if cp.get('prefix'):
                    cp_parts.append(
                        f"PREFIX '{cp['prefix']['part']}'→{cp['prefix']['root']}")
                if cp.get('root'):
                    rt = cp['root']
                    rtag = 'ORIG2' if rt.get('is_orig2') else 'ORIG1'
                    cp_parts.append(
                        f"ROOT '{rt['part']}'→{rt['root']}({rtag})")
                if cp_parts:
                    compound_suffix = (
                        f" | COMPOUND: {cp['label']} — "
                        + ' + '.join(cp_parts))

            # COGNATE NOTE suffix: v3.4 downgraded competition for RU words
            cognate_note_suffix = ''
            if cognate_note_only and cognate_ref:
                cognate_note_suffix = (
                    f" | COGNATE NOTE (EN more degraded): "
                    f"'{cognate_ref['en_cousin']}' → "
                    f"{cognate_ref['root_letters']} "
                    f"(EN score {cognate_ref.get('score', '?')})")

            if is_derivative:
                parent_name, parent_id = result.derivative_of
                result.add_log(
                    f"DERIVATIVE: route to A4_DERIVATIVES of {parent_name} (#{parent_id}), "
                    f"NOT A1_ENTRIES"
                )
                qid = self.writer.queue_for_oversight(
                    entry,
                    f"DERIVATIVE of {parent_name} #{parent_id} — write to A4_DERIVATIVES"
                    + sem_suffix + compound_suffix,
                    q_result.passed, u_result.passed, f_result.passed
                )
                result.queue_id = qid

            elif has_competing_orig2:
                # v3.2: COMPETING TRACKS — ORIG1 passed but ORIG2 also found.
                # ALWAYS force PENDING_REVIEW for human decision.
                o2 = result.orig2_parallel
                flag = (
                    f"COMPETING TRACKS [{lang_label}]: "
                    f"ORIG1 ({top.letters}, {top.token_count} tok, score {score}/10) vs "
                    f"ORIG2 (Kashgari '{o2['kashgari_translit']}' "
                    f"= \"{o2['kashgari_meaning']}\", line {o2['kashgari_line']}, "
                    f"ORIG2 score {o2['orig2_score']}/10). "
                    f"Human review required — determine correct origin track."
                    + sem_suffix + compound_suffix + cognate_note_suffix
                )
                result.add_log(f"COMPETING TRACKS → forced PENDING [{lang_label}]")
                qid = self.writer.queue_for_oversight(
                    entry, flag,
                    q_result.passed, u_result.passed, f_result.passed
                )
                result.queue_id = qid

            elif has_competing_depal:
                # v3.2: DEPALATALIZATION COMPETITION — primary root differs from
                # depalatalized root. Force PENDING_REVIEW.
                dc = result.competing_depal
                dc_ops = getattr(dc, 'operations', [])
                flag = (
                    f"DEPAL COMPETITION [{lang_label}]: "
                    f"primary {top.letters} ({top.token_count} tok, score {score}/10) vs "
                    f"depalatalized {dc.letters} ({dc.token_count} tok, ops={dc_ops}). "
                    f"Palatalized consonant may mask true root. Human review required."
                    + sem_suffix + compound_suffix + cognate_note_suffix
                )
                result.add_log(f"DEPAL COMPETITION → forced PENDING [{lang_label}]")
                qid = self.writer.queue_for_oversight(
                    entry, flag,
                    q_result.passed, u_result.passed, f_result.passed
                )
                result.queue_id = qid

            elif has_cognate_competition:
                # v3.3: COGNATE COMPETITION — local pipeline root disagrees with
                # English cognate root.  Force PENDING_REVIEW for human decision.
                # (v3.4: only fires for EN words or MODERN_TECH_TERMS now)
                cog = cognate_ref
                cog_root = cog['root_letters']
                cog_score = cog['score']
                flag = (
                    f"COGNATE COMPETITION [{lang_label}]: "
                    f"{lang_label} pipeline → {top.letters} ({top.token_count} tok, "
                    f"score {score}/10) vs EN cousin '{cog['en_cousin']}' "
                    f"→ {cog_root} ({cog['token_count']} tok, "
                    f"EN score {cog_score}/10, "
                    f"chain: {cog.get('phonetic_chain', '—')}). "
                    f"Sibling root disagreement — human review required."
                    + sem_suffix + compound_suffix
                )
                result.add_log(f"COGNATE COMPETITION → forced PENDING [{lang_label}]")
                qid = self.writer.queue_for_oversight(
                    entry, flag,
                    q_result.passed, u_result.passed, f_result.passed
                )
                result.queue_id = qid

            elif (score >= SCORE_AUTO_WRITE and q_result.passed
                  and u_result.passed and pos_ok):
                result.add_log(f"Score {score}/10 >= {SCORE_AUTO_WRITE} + positional OK — queued as AUTO-WRITE candidate [{lang_label}]")
                aw_flag = f"Score {score}/10 [{lang_label}] — auto-write candidate. Confirm ROOT_ID and QUR_MEANING before final write."
                if cognate_agrees and cognate_ref:
                    aw_flag += f" | COGNATE CONFIRMS: EN cousin '{cognate_ref['en_cousin']}' agrees → {cognate_ref['root_letters']}"
                aw_flag += sem_suffix + compound_suffix + cognate_note_suffix
                qid = self.writer.queue_for_oversight(
                    entry, aw_flag,
                    q_result.passed, u_result.passed, f_result.passed
                )
                result.queue_id = qid
                if is_russian:
                    self.existing_ru_terms[term.upper()] = -1
                else:
                    self.existing_terms[term.upper()] = -1  # prevent re-queuing
            elif SCORE_QUEUE <= score:
                result.add_log(f"Score {score}/10 — queued for oversight")
                pq_flag = f"Score {score}/10 — review required (positional={getattr(top, 'positional_score', '?')})"
                if cognate_agrees and cognate_ref:
                    pq_flag += f" | COGNATE CONFIRMS: EN cousin '{cognate_ref['en_cousin']}' agrees → {cognate_ref['root_letters']}"
                elif cognate_ref and cognate_ref.get('source') == 'EN_PIPELINE':
                    pq_flag += (f" | COGNATE NOTE: EN cousin '{cognate_ref['en_cousin']}' "
                                f"→ {cognate_ref['root_letters']} (EN score {cognate_ref.get('score','?')})")
                pq_flag += sem_suffix + compound_suffix + cognate_note_suffix
                qid = self.writer.queue_for_oversight(
                    entry, pq_flag,
                    q_result.passed, u_result.passed, f_result.passed
                )
                result.queue_id = qid
            else:
                result.add_log(f"Score {score}/10 <= {SCORE_REJECT} — rejected (not queued)")
        else:
            result.add_log("[DRY RUN] No writes performed")

        # 360-DEGREE REPORT
        if not self.skip_reports:
            result.report_path = self.reporter.generate(result)

        # EXPORT QUEUE JSON for Oversight Dashboard
        self.writer.export_queue_json(WORKSPACE_DIR)

        return result

    # ── phrase pipeline ───────────────────────────────────────────────────────

    def _process_phrase(self, raw: str, key_terms: list, dry_run: bool) -> ProcessResult:
        result = ProcessResult(raw, 'phrase')
        result.add_log(f"Phrase: {len(key_terms)} content words → {key_terms}")
        sub_results = []
        for term in key_terms:
            result.add_log(f"\n  → Processing term: {term}")
            sub = self._process_single_term(term, 'english_word', dry_run)
            sub_results.append(sub)
        result.cluster_members = sub_results
        result.add_log(f"\nPhrase complete. {len(sub_results)} terms processed.")
        return result

    # ── Qur'anic ref lookup ───────────────────────────────────────────────────

    def _process_quran_ref(self, ref: str) -> ProcessResult:
        result = ProcessResult(ref, 'quran_ref')
        result.add_log(f"Looking up entries for {ref}...")
        try:
            wb = load_workbook(self.master_file, read_only=True, data_only=True)
            ws = wb['A3_QURAN_REFS']
            headers = None
            found   = []
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not any(row):
                    continue
                d = dict(zip(headers, row))
                vr = str(d.get('VERSE_REF', '') or '').strip()
                if ref in vr or vr in ref:
                    found.append(d)
            wb.close()
            result.add_log(f"Found {len(found)} entries linked to {ref}")
            for fd in found:
                result.add_log(f"  #{fd.get('ENTRY_ID','?')} {fd.get('EN_TERM','?')} "
                               f"(root: {fd.get('ROOT_LETTERS','?')})")
        except Exception as e:
            result.add_log(f"Qur'an ref lookup error: {e}")
        return result

    # ── ORIG2 / Kashgari track (v2.2) ──────────────────────────────────────────

    def _try_orig2_track(self, term: str, result: ProcessResult,
                         dry_run: bool,
                         is_russian: bool = False) -> Optional[ProcessResult]:
        """
        Attempt ORIG2 (Kashgari/Bitig) attestation after ORIG1 (Q-Gate) fails.

        Protocol (CLAUDE.md §6 two-track gate):
          ROOT_LIST FAIL → Kashgari search → if attested → ORIG2 entry
          Score under Bitig protocol.  Route to ENGINE_QUEUE with ORIG2 flag.

        v3.2: language-aware — uses Russian extraction + Latin conversion
        when is_russian=True.

        Returns ProcessResult if ORIG2 found, None if both tracks fail.
        """
        if self.kashgari_gate is None:
            result.add_log("ORIG2 track disabled (Kashgari corpus not loaded)")
            return None

        # v3.2: language-aware consonant extraction for ORIG2 search
        if is_russian and self.ru_reversal is not None:
            cyrillic_cons = self.ru_reversal.extract_consonants(term)
            if not cyrillic_cons:
                result.add_log("ORIG2: no consonants to search (Russian)")
                return None
            # Convert Cyrillic → Latin for Kashgari search
            latin_skel = self.ru_reversal.to_latin_skeleton(cyrillic_cons)
            consonants = list(latin_skel)
            _, ops, suffix = self.ru_reversal.strip_operations(term)
        else:
            consonants = self.reversal.extract_consonants(term)
            _, ops, suffix = self.reversal.strip_operations(term)  # ops for metadata only

        if not consonants:
            result.add_log("ORIG2: no consonants to search")
            return None

        result.add_log(f"ORIG2: searching Kashgari for skeleton '{(''.join(consonants))}'...")
        k_result = self.kashgari_gate.check(term, consonants)

        if not k_result.passed:
            result.add_log(f"ORIG2 FAIL: {k_result.details.get('reason', 'not in Kashgari')}")
            return None

        # ── ORIG2 PASS ──────────────────────────────────────────────────────
        kd = k_result.details
        result.add_log(
            f"ORIG2 PASS: Kashgari attests '{kd['kashgari_translit']}' "
            f"= \"{kd['kashgari_meaning']}\" "
            f"(line {kd['kashgari_line']}, {kd['attestation_type']})"
        )

        # Log B01-B07 warnings
        for w in kd.get('bitig_warnings', []):
            result.add_log(f"  ⚠ {w}")

        # Create ORIG2 root candidate
        orig2_root = RootCandidate(
            letters     = kd.get('skeleton', ''.join(consonants)),
            token_count = 0,   # no Qur'anic tokens — ORIG2
            ar_word     = kd.get('kashgari_translit', ''),
            operations  = ['ORIG2_BITIG'] + ops,
        )
        orig2_root.positional_score = 0.5  # neutral for ORIG2

        result.confirmed_root  = orig2_root
        result.orig2_track     = True
        result.orig2_details   = kd

        # Score ORIG2 entry
        score = self._score_orig2(k_result, consonants, term)
        orig2_root.score = score
        result.add_log(f"ORIG2 Score: {score}/10")

        # Build entry record (BITIG format — still uses EntryRecord for queue)
        entry = EntryRecord()
        entry.en_term        = term.upper()
        entry.ar_word        = kd.get('kashgari_translit', '')
        entry.root_letters   = kd.get('skeleton', '')
        entry.score          = score
        entry.pattern        = 'A'      # Hidden — English speaker unaware of Bitig origin
        entry.inversion_type = 'HIDDEN'
        entry.source_form    = kd.get('kashgari_translit', '')
        entry.phonetic_chain = f"ORIG2: {kd.get('kashgari_headword', '')} → {term}"
        entry.foundation_ref = (
            f"ORIG2 BITIG: Kashgari Dīwān line {kd.get('kashgari_line', '?')} | "
            f"Meaning: {kd.get('kashgari_meaning', '?')} | "
            f"Attestation: {kd.get('attestation_type', '?')} | "
            f"Hits: {kd.get('all_hits', 0)}"
        )
        result.entry_record = entry

        # Queue for oversight — ORIG2 ALWAYS needs human review
        if not dry_run:
            bitig_warns = kd.get('bitig_warnings', [])
            flag = (
                f"ORIG2 candidate — Kashgari '{kd.get('kashgari_translit', '')}' "
                f"({kd.get('attestation_type', '')}, score {score}/10)"
            )
            if bitig_warns:
                flag += f" | Warnings: {'; '.join(bitig_warns)}"

            # v3.3b: Append cognate intelligence if available
            cog_ref = getattr(result, 'cognate_crossref', None)
            if cog_ref and cog_ref.get('source') == 'EN_PIPELINE':
                flag += (
                    f" | COGNATE: EN cousin '{cog_ref['en_cousin']}' → "
                    f"{cog_ref['root_letters']} ({cog_ref.get('token_count', 0)} tok, "
                    f"EN score {cog_ref.get('score', '?')}/10, "
                    f"chain: {cog_ref.get('phonetic_chain', '—')})"
                )

            qid = self.writer.queue_for_oversight(
                entry, flag,
                False,   # Q-gate (ORIG1) = FAIL
                False,   # U-gate N/A for ORIG2
                True     # F-gate = PASS (auto)
            )
            result.queue_id = qid
            result.add_log(f"ORIG2 queued for oversight: {qid}")

        # Generate report
        if not self.skip_reports:
            result.report_path = self.reporter.generate(result)
        self.writer.export_queue_json(WORKSPACE_DIR)
        return result

    def _score_orig2(self, kashgari_result: GateResult,
                     consonants: list, en_word: str) -> int:
        """
        Score ORIG2 entry.  Parallel to Scorer but Kashgari-based.

        v2.3 FIX — meaning_match phonetic verification:
        meaning_match alone (no consonant overlap) is R10 violation.
        WIFE→'kis' scores LOW because W-F ≠ K-S (zero consonant overlap).
        meaning_match now REQUIRES partial skeleton overlap to score above 5.

        Max score breakdown:
          Kashgari attestation (+3)
          Match quality: skeleton(+2) or meaning_with_overlap(+1) or meaning_only(+0)
          Multiple attestations (+1)
          No B01-B07 warnings (+2)
          Consonant coverage (+2)
        Total possible: 10
        """
        score = 0
        kd = kashgari_result.details

        # Kashgari attestation (+3) — equivalent to Q-gate for ORIG1
        score += 3

        # Match quality (+2/+1/+0) — v2.3: meaning_match now verified
        att_type = kd.get('attestation_type', '')
        if att_type == 'skeleton_match':
            score += 2
        elif att_type == 'meaning_match':
            # v2.3: check if ANY consonants overlap between English word
            # and the Kashgari transliteration. No overlap = R10 violation.
            k_translit = kd.get('kashgari_translit', '')
            en_cons = set(c for c in en_word.lower() if c.isalpha() and c not in 'aeiou')
            # Bitig consonant equivalences for overlap check
            equiv = {'q': 'kgc', 'k': 'qgc', 'g': 'qk', 'p': 'b', 'b': 'p',
                     'c': 'sjz', 's': 'cjz', 'z': 'cs', 'j': 'csz',
                     't': 'd', 'd': 't', 'f': 'pv', 'v': 'fpw', 'w': 'v'}
            k_cons = set(c for c in k_translit.lower() if c.isalpha() and c not in 'aeiouüöıäəāēīōū')
            # Check direct or equivalent overlap
            overlap = 0
            for ec in en_cons:
                if ec in k_cons:
                    overlap += 1
                elif any(eq in k_cons for eq in equiv.get(ec, '')):
                    overlap += 1
            if overlap >= 2:
                score += 1  # meaning_match WITH phonetic overlap — acceptable
            else:
                score += 0  # meaning_match WITHOUT overlap — R10 violation, no bonus
                kd['meaning_only_warning'] = (
                    f"meaning_match but only {overlap} consonant overlap "
                    f"(EN={sorted(en_cons)}, Kashgari={sorted(k_cons)})"
                )

        # Multiple attestations (+1)
        if kd.get('all_hits', 0) > 1:
            score += 1

        # No B01-B07 warnings (+2)
        warnings = kd.get('bitig_warnings', [])
        if not warnings:
            score += 2
        elif len(warnings) == 1:
            score += 1

        # Consonant coverage (+2)
        skel_len = len(kd.get('skeleton', ''))
        word_cons = len([c for c in en_word.lower()
                         if c.isalpha() and c not in 'aeiou'])
        if att_type == 'skeleton_match' and skel_len > 0 and abs(word_cons - skel_len) <= 1:
            score += 2
        elif att_type == 'skeleton_match' and skel_len > 0 and abs(word_cons - skel_len) <= 2:
            score += 1
        elif att_type == 'meaning_match':
            # meaning_match coverage capped at +1 regardless of skeleton
            if skel_len > 0 and abs(word_cons - skel_len) <= 1:
                score += 1

        return min(score, 10)

    # ── v3.2: PARALLEL ORIG2 CHECK ─────────────────────────────────────────────

    def _try_orig2_parallel(self, term: str, is_russian: bool) -> Optional[dict]:
        """
        Parallel ORIG2 check — runs ALONGSIDE ORIG1, not as fallback.

        v3.2: For words where ORIG1 passes but the word might actually be
        ORIG2 (Bitig/Turkic).  Especially critical for Russian words where
        >50% are Bitig-corridor.

        IMPORTANT: This parallel check searches ONLY KNOWN_ORIG2_ENTRIES
        (manually verified Kashgari entries), NOT the full KashgariIndex.
        The full index is too broad and produces false positives.
        The full Kashgari search is reserved for the ORIG2 fallback path
        (when ORIG1 fails completely).

        Converts Russian consonants to Latin for Kashgari search.
        Returns dict with ORIG2 details if found, None otherwise.
        """
        if self.kashgari_gate is None:
            return None

        # Extract consonants and convert to Latin for Kashgari search
        if is_russian:
            cyrillic_cons = self.ru_reversal.extract_consonants(term)
            if not cyrillic_cons:
                return None
            # Generate all Latin skeleton variants (main + voicing + suffix-stripped)
            latin_variants = self.ru_reversal.to_latin_skeleton_variants(cyrillic_cons)
        else:
            consonants = self.reversal.extract_consonants(term)
            if not consonants:
                return None
            # For English: also generate suffix-stripped and voicing variants
            main_skel = ''.join(consonants).lower()
            latin_variants = [main_skel]
            # Add suffix-stripped (1 consonant)
            if len(main_skel) >= 3:
                latin_variants.append(main_skel[:-1])

        # Search ONLY KNOWN_ORIG2_ENTRIES (not the full KashgariIndex)
        # This prevents false positives from broad skeleton matching.
        known_entries = KashgariGate.KNOWN_ORIG2_ENTRIES
        for latin_skel in latin_variants:
            hits = known_entries.get(latin_skel, [])
            if hits:
                best = hits[0]
                # Build a GateResult-like dict for scoring
                orig2_score = 3 + 2  # Kashgari attested (+3) + skeleton_match (+2)
                if len(hits) > 1:
                    orig2_score += 1  # multiple hits
                # Check consonant coverage
                skel_len = len(best.get('skeleton', ''))
                word_cons = len(latin_skel)
                if skel_len > 0 and abs(word_cons - skel_len) <= 1:
                    orig2_score += 2
                elif skel_len > 0 and abs(word_cons - skel_len) <= 2:
                    orig2_score += 1
                orig2_score = min(orig2_score, 10)

                return {
                    'passed': True,
                    'kashgari_translit': best.get('translit', ''),
                    'kashgari_meaning':  best.get('meaning', ''),
                    'kashgari_line':     best.get('line', 0),
                    'attestation_type':  'known_orig2_match',
                    'skeleton':          best.get('skeleton', latin_skel),
                    'orig2_score':       orig2_score,
                    'all_hits':          len(hits),
                    'bitig_warnings':    [],
                    'search_skeleton':   latin_skel,
                }
        return None

    # ── v3.3: COGNATE CROSS-REFERENCING ───────────────────────────────────────
    def _try_cognate_crossref(self, term: str, is_russian: bool) -> Optional[dict]:
        """
        Sibling Database Principle — automated.

        When processing a Russian word, check COGNATE_CROSSREF_RU_TO_EN for
        a known English equivalent.  If found, run the ENGLISH form through
        the English PhoneticReversal → Q → U → Scorer pipeline.

        Returns dict with:
          - en_cousin:      the English word used
          - root_letters:   root found by English pipeline
          - token_count:    Q-gate tokens
          - score:          English pipeline score
          - phonetic_chain: English U-gate chain
          - candidates:     full candidate list from English reversal
        Or None if no cognate found or English pipeline fails.

        Also works EN→RU: when processing English, checks if there's a
        Russian entry already confirmed.
        """
        if is_russian:
            en_cousin = COGNATE_CROSSREF_RU_TO_EN.get(term.lower())
            if not en_cousin:
                return None

            # Check if English cousin already in lattice (strongest signal)
            existing_en = self.existing_terms.get(en_cousin.upper())
            if existing_en:
                return {
                    'en_cousin': en_cousin,
                    'source': 'LATTICE_ENTRY',
                    'entry_id': existing_en,
                    'note': f'{en_cousin} already in A1_ENTRIES #{existing_en}',
                }

            # Process English cousin through English pipeline
            # v3.3b: Try the full word first, then prefix-stripped variants,
            # then suffix-stripped variants (-ent/-ent which main pipeline skips).
            # DESCENT → DE-SCENT → S-C-N → سَكَنَ ; TALENT → TAL-ENT → T-L-N
            LATIN_PREFIXES = ['de', 're', 'in', 'con', 'dis', 'ex', 'pre',
                              'pro', 'ad', 'com', 'per', 'sub', 'trans', 'un']
            COGNATE_EXTRA_SUFFIXES = ['ent', 'ment', 'ure', 'ude',
                                      'ance', 'ence', 'ant']

            try:
                # Build list of word variants to try: (word_form, extra_ops, label)
                variants = [(en_cousin, [], 'FULL')]

                # ── Prefix-stripped variants ─────────────────────────────────
                lw = en_cousin.lower()
                for pfx in LATIN_PREFIXES:
                    if lw.startswith(pfx) and len(lw) - len(pfx) >= 3:
                        stem = en_cousin[len(pfx):]
                        variants.append((stem, [f'OP_PREFIX({pfx}-)'], f'PREFIX_{pfx}'))

                # ── Suffix-stripped variants (for -ent/-ence not in main list) ─
                for sfx in COGNATE_EXTRA_SUFFIXES:
                    if lw.endswith(sfx) and len(lw) - len(sfx) >= 3:
                        stem = en_cousin[:-len(sfx)]
                        variants.append((stem, [f'OP_SUFFIX(-{sfx})'], f'SUFFIX_{sfx}'))

                # ── Prefix+suffix combined (e.g. DE-SCEN-T with -t suffix) ──
                for pfx in LATIN_PREFIXES:
                    if lw.startswith(pfx):
                        inner = en_cousin[len(pfx):]
                        for sfx in COGNATE_EXTRA_SUFFIXES:
                            if inner.lower().endswith(sfx) and len(inner) - len(sfx) >= 2:
                                stem = inner[:-len(sfx)]
                                if len(stem) >= 2:
                                    variants.append((stem,
                                        [f'OP_PREFIX({pfx}-)', f'OP_SUFFIX(-{sfx})'],
                                        f'BOTH_{pfx}_{sfx}'))

                # ── Score all variants, keep the absolute best ──────────────
                best_score = -1
                best_cand = None
                best_q = None
                best_u = None
                best_variant_label = ''
                best_ops = []
                best_word_used = en_cousin

                for (word_form, extra_ops, label) in variants:
                    if len(word_form) < 2:
                        continue
                    cands = self.reversal.reverse(word_form)
                    if not cands:
                        continue
                    for ci in range(min(3, len(cands))):
                        cand = cands[ci]
                        cq = self.q_gate.check(cand.letters)
                        if not cq.passed:
                            continue
                        cu = self.u_gate.verify(word_form, cand.letters,
                                                cand.operations + extra_ops)
                        cf = self.f_gate.assign(word_form, cand.letters,
                                                cu.details.get('phonetic_chain', ''))
                        cs, cb = self.scorer.score(cand, word_form, cq, cu, cf)
                        if cs > best_score:
                            best_score = cs
                            best_cand = cand
                            best_q = cq
                            best_u = cu
                            best_variant_label = label
                            best_ops = extra_ops + (cand.operations or [])
                            best_word_used = word_form

                if best_cand is None or best_score < SCORE_QUEUE:
                    return None

                return {
                    'en_cousin':      en_cousin,
                    'source':         'EN_PIPELINE',
                    'root_letters':   best_cand.letters,
                    'token_count':    best_cand.token_count,
                    'score':          best_score,
                    'phonetic_chain': best_u.details.get('phonetic_chain', '') if best_u else '',
                    'ar_word':        best_cand.ar_word or '',
                    'operations':     best_ops,
                    'positional':     getattr(best_cand, 'positional_score', 0.5),
                    'variant_used':   best_variant_label,
                    'word_form_used': best_word_used,
                }
            except Exception as e:
                return {'en_cousin': en_cousin, 'source': 'ERROR', 'error': str(e)}

        else:
            # EN→RU: check if any Russian siblings exist
            ru_siblings = COGNATE_CROSSREF_EN_TO_RU.get(term.upper(), [])
            if not ru_siblings:
                return None
            for sib in ru_siblings:
                existing_ru = self.existing_ru_terms.get(sib.upper())
                if existing_ru:
                    return {
                        'ru_sibling': sib,
                        'source': 'LATTICE_ENTRY',
                        'entry_id': existing_ru,
                        'note': f'Russian sibling {sib.upper()} in A1_ЗАПИСИ #{existing_ru}',
                    }
            return None


# ═══════════════════════════════════════════════════════════════════════════════
# INTERACTIVE CLI — main()
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("""
╔══════════════════════════════════════════════════════════════╗
║         USLaP Autonomous Engine v1.0                         ║
║         Unified Source Language Proof                        ║
║         بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ              ║
╚══════════════════════════════════════════════════════════════╝

Menu:
  1. Process word/root/ratio/phrase  (full 360-degree analysis + lattice placement)
  2. Dry run  (analysis only, no writes)
  3. Export ENGINE_QUEUE to JSON  (for Oversight Dashboard)
  4. Quit

Or type your query directly (no menu number needed).
""")

    master = sys.argv[1] if len(sys.argv) > 1 else MASTER_FILE
    if not os.path.exists(master):
        print(f"ERROR: Master file not found:\n  {master}")
        sys.exit(1)

    try:
        engine = USLaPEngine(master_file=master)
    except Exception as e:
        print(f"Engine initialisation failed: {e}")
        sys.exit(1)

    while True:
        try:
            user_input = input("\nQuery > ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nExiting.")
            break

        if not user_input:
            continue

        if user_input == '1':
            term = input("Enter query: ").strip()
            if term:
                engine.process(term, dry_run=False)

        elif user_input == '2':
            term = input("Enter query (DRY RUN): ").strip()
            if term:
                engine.process(term, dry_run=True)

        elif user_input == '3':
            engine.writer.export_queue_json(WORKSPACE_DIR)

        elif user_input in ('4', 'quit', 'exit', 'q'):
            print("Exiting engine.")
            break

        else:
            # Direct query — no menu number
            engine.process(user_input, dry_run=False)


if __name__ == '__main__':
    main()
