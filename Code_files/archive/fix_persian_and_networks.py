#!/usr/bin/env python3
"""
Fix two issues in the USLaP database:
1. Persian fa_term not populated — 59 rows exist in entries with all term columns NULL (entry IDs 230–288)
   Read sheet PERSIAN_A1_MADĀKHIL, UPDATE those rows with fa_term from the sheet data.
   Match by row order or by entry_id sequence starting at the first NULL-term row.
2. networks table empty — read sheet M4_NETWORKS from the master Excel file, populate the networks table.
   This will resolve 57 of the 64 remaining FK violations.
   Columns needed: network_id (N01, N02...), name, description, key_verse, entry_ids.
"""

import sqlite3
try:
    from uslap_db_connect import connect as _uslap_connect
    _HAS_WRAPPER = True
except ImportError:
    _HAS_WRAPPER = False
import openpyxl
import re
import sys
import os

# ============================================================================
# GLOBAL CONFIGURATION
# ============================================================================

EXCEL_PATH = "USLaP_Final_Data_Consolidated_Master_v3.xlsx"
DB_PATH = "Code_files/uslap_lattice.db"

# ============================================================================
# EXTRACT_CONSONANTS UDF (from USLaP_Engine.py)
# ============================================================================

def extract_consonants(word):
    """
    Python UDF for SQLite: Extract consonant skeleton from a word.
    Must be registered before any operations on word_fingerprints table.
    
    This is the same logic as PhoneticReversal.extract_consonants() in USLaP_Engine.py.
    """
    if not word:
        return ""
    
    # Import the function directly from USLaP_Engine if available
    try:
        from USLaP_Engine import PhoneticReversal
        # Create a minimal instance just to use the method
        # We'll implement the logic directly here to avoid dependencies
        pass
    except ImportError:
        pass
    
    # Direct implementation from USLaP_Engine.py
    vowels = set('aeiou')
    result = []
    i = 0
    word_lower = word.lower()
    
    while i < len(word_lower):
        digraph = word_lower[i:i+2] if i + 1 < len(word_lower) else ''
        if digraph in ('sh', 'ch', 'gh', 'th', 'ph', 'wh', 'qu'):
            result.append(digraph)
            i += 2
        elif word_lower[i] not in vowels:
            result.append(word_lower[i])
            i += 1
        else:
            i += 1
    
    return ''.join(result)

def clean_column_name(col):
    """Convert Excel column header to valid SQLite column name."""
    if col is None:
        return "unknown"
    col = str(col).strip()
    col = re.sub(r'[^\w\s]', '', col)  # Remove punctuation
    col = re.sub(r'\s+', '_', col)     # Replace spaces with underscore
    col = col.lower()
    if not col:
        return "unknown"
    return col

def fix_persian_terms(conn, cursor, wb):
    """Fix Persian fa_term not populated issue."""
    print("Fixing Persian fa_term not populated...")
    
    if "PERSIAN_A1_MADĀKHIL" not in wb.sheetnames:
        print("  ERROR: Sheet PERSIAN_A1_MADĀKHIL not found")
        return 0
    
    ws = wb["PERSIAN_A1_MADĀKHIL"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("  ERROR: No data in PERSIAN_A1_MADĀKHIL")
        return 0
    
    # First row has complex headers with Persian and English
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    # Debug: print cleaned headers to see what we're working with
    print(f"  Cleaned headers (first 5): {headers[:5]}")
    
    # Get the list of entry IDs that need fa_term (entries 230-288 with NULL terms)
    cursor.execute('''
        SELECT entry_id FROM entries 
        WHERE en_term IS NULL AND ru_term IS NULL AND fa_term IS NULL
        ORDER BY entry_id
    ''')
    null_entry_ids = [row[0] for row in cursor.fetchall()]
    
    if len(null_entry_ids) != 59:
        print(f"  WARNING: Expected 59 NULL-term entries, found {len(null_entry_ids)}")
    
    print(f"  Found {len(null_entry_ids)} entries needing fa_term (IDs: {null_entry_ids[0]} to {null_entry_ids[-1]})")
    
    count = 0
    # Process each row in the Persian sheet (skip header row)
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        if i-1 >= len(null_entry_ids):
            print(f"  WARNING: More Persian rows ({i}) than NULL-term entries ({len(null_entry_ids)})")
            break
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Try multiple possible column names for Persian term
        # Based on inspection, the cleaned header will be something like 'vāzhe_fārsī_واژه_فارسی_persian_term'
        # Let's find the right column by checking for 'persian' in the key
        fa_term = None
        for key in row_dict:
            if 'persian' in key.lower():
                fa_term = row_dict.get(key)
                break
        
        # If not found by 'persian', try looking for column at index 1 (based on inspection)
        if not fa_term and len(row) > 1:
            fa_term = row[1]  # Column index 1 based on inspection
        
        if not fa_term:
            print(f"  Row {i}: No Persian term found")
            continue
        
        # Parse the Persian term from the format "Persian / transliteration / meaning"
        # Example: 'سَلطَنَت / salṭanat / sovereignty'
        fa_term_str = str(fa_term).strip()
        if ' / ' in fa_term_str:
            # Extract just the Persian part (before first '/')
            persian_part = fa_term_str.split(' / ')[0].strip()
            if persian_part:
                fa_term = persian_part
            else:
                print(f"  Row {i}: Could not extract Persian part from: {fa_term_str}")
                continue
        else:
            # Use as-is
            fa_term = fa_term_str
        
        entry_id = null_entry_ids[i-1]  # Match by row order
        
        # Update the entry with fa_term
        try:
            cursor.execute('''
                UPDATE entries 
                SET fa_term = ?, modified_at = CURRENT_TIMESTAMP
                WHERE entry_id = ?
            ''', (fa_term, entry_id))
            
            if cursor.rowcount > 0:
                count += 1
                print(f"    Updated entry {entry_id} with Persian term: {fa_term}")
            else:
                print(f"    No rows updated for entry_id {entry_id}")
                
        except Exception as e:
            print(f"    Error updating entry {entry_id}: {e}")
    
    print(f"  Updated {count} Persian entries with fa_term")
    return count

def populate_networks(conn, cursor, wb):
    """Populate networks table from M4_NETWORKS sheet."""
    print("Populating networks table...")
    
    if "M4_NETWORKS" not in wb.sheetnames:
        print("  ERROR: Sheet M4_NETWORKS not found")
        return 0
    
    ws = wb["M4_NETWORKS"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("  ERROR: No data in M4_NETWORKS")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map columns according to Excel structure
        network_data = {
            'network_id': row_dict.get('network_id'),
            'name': row_dict.get('name'),
            'title': row_dict.get('title'),
            'link_verse': row_dict.get('link_verse'),
            'description': row_dict.get('description'),
            'mechanism': row_dict.get('mechanism'),
            'entry_ids': row_dict.get('entry_ids'),
            'status': row_dict.get('status'),
            'foundation_ref': row_dict.get('foundation_ref'),
        }
        
        # Check if network already exists
        cursor.execute('SELECT 1 FROM networks WHERE network_id = ?', (network_data['network_id'],))
        if cursor.fetchone():
            print(f"    Network {network_data['network_id']} already exists, skipping")
            continue
        
        # Insert into networks table
        try:
            cursor.execute('''
                INSERT INTO networks (
                    network_id, name, title, link_verse, description,
                    mechanism, entry_ids, status, foundation_ref
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                network_data['network_id'], network_data['name'], network_data['title'],
                network_data['link_verse'], network_data['description'], network_data['mechanism'],
                network_data['entry_ids'], network_data['status'], network_data['foundation_ref']
            ))
            count += 1
            print(f"    Inserted network {network_data['network_id']}: {network_data['name']}")
        except Exception as e:
            print(f"    Error inserting network row {i}: {e}")
            print(f"    Data: {network_data}")
    
    print(f"  Inserted {count} networks")
    return count

def check_foreign_keys(conn, cursor):
    """Check foreign key constraints and return count of violations."""
    print("\nChecking foreign key constraints...")
    cursor.execute("PRAGMA foreign_key_check")
    fk_errors = cursor.fetchall()
    
    if fk_errors:
        print(f"  Found {len(fk_errors)} foreign key violations:")
        for error in fk_errors[:5]:  # Show first 5
            print(f"    Table: {error[0]}, Row: {error[1]}, Referenced: {error[2]}, FK Index: {error[3]}")
        if len(fk_errors) > 5:
            print(f"    ... and {len(fk_errors) - 5} more")
    else:
        print("  ✓ All foreign key constraints satisfied")
    
    return len(fk_errors)

def main():
    print("\n" + "=" * 70)
    print("  Fixing Persian Terms and Populating Networks")
    print("=" * 70)
    
    # Check if files exist
    if not os.path.exists(EXCEL_PATH):
        print(f"\n❌ ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)
    
    if not os.path.exists(DB_PATH):
        print(f"\n❌ ERROR: Database file not found: {DB_PATH}")
        sys.exit(1)
    
    # Load Excel workbook
    print(f"\n📖 Loading Excel file: {EXCEL_PATH}")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        print(f"  Found {len(wb.sheetnames)} sheets")
    except Exception as e:
        print(f"\n❌ ERROR: Failed to load Excel file: {e}")
        sys.exit(1)
    
    # Connect to database
    print(f"\n🗄️  Connecting to database: {DB_PATH}")
    try:
        conn = _uslap_connect(DB_PATH) if _HAS_WRAPPER else sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Enable foreign keys
        cursor.execute("PRAGMA foreign_keys = ON")
        
        # Register extract_consonants UDF (CRITICAL: must be done before any updates due to triggers)
        print("  Registering extract_consonants() UDF...")
        conn.create_function("extract_consonants", 1, extract_consonants)
        
        # Begin transaction
        conn.execute("BEGIN TRANSACTION")
        
        # Fix Persian terms
        persian_count = fix_persian_terms(conn, cursor, wb)
        
        # Populate networks
        networks_count = populate_networks(conn, cursor, wb)
        
        # Commit transaction
        conn.commit()
        
        # Check foreign key constraints
        violation_count = check_foreign_keys(conn, cursor)
        
        # Verify Persian terms were updated
        cursor.execute("SELECT COUNT(*) FROM entries WHERE fa_term IS NOT NULL AND fa_term != ''")
        fa_count = cursor.fetchone()[0]
        print(f"\n📊 Verification:")
        print(f"  Persian entries with fa_term: {fa_count} (target: ~59)")
        print(f"  Networks inserted: {networks_count}")
        print(f"  Foreign key violations: {violation_count} (target: ≤7)")
        
        if violation_count <= 7:
            print("\n✅ SUCCESS: Target of ≤7 foreign key violations achieved!")
        else:
            print(f"\n⚠️  WARNING: {violation_count} foreign key violations exceeds target of ≤7")
        
        # Close workbook and connection
        wb.close()
        conn.close()
        
    except Exception as e:
        print(f"\n❌ ERROR during fix: {e}")
        import traceback
        traceback.print_exc()
        
        # Rollback if connection is still open
        try:
            conn.rollback()
            conn.close()
        except:
            pass
        
        print("\n⚠️  Fix failed. Database has been rolled back.")
        sys.exit(1)

if __name__ == "__main__":
    main()