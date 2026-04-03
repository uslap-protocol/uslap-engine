#!/usr/bin/env python3
"""
USLaP SAFE ACCELERATED FOREST GROWTH
Accelerated growth with backup protection and error handling
"""

import openpyxl
from pathlib import Path
from datetime import datetime
import time
import shutil
import sys
import os

# ============================================================================
# SAFE FILE HANDLER
# ============================================================================

class SafeFileHandler:
    """Handles file operations safely with backups"""
    
    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self.backup_path = None
        self.last_successful_write = None
        
    def create_backup(self):
        """Create a timestamped backup before starting operations"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{self.filepath.stem}_backup_{timestamp}{self.filepath.suffix}"
        self.backup_path = self.filepath.parent / backup_name
        
        try:
            shutil.copy2(self.filepath, self.backup_path)
            print(f"📁 Created backup: {self.backup_path.name}")
            return True
        except Exception as e:
            print(f"❌ Failed to create backup: {e}")
            return False
    
    def verify_file_integrity(self):
        """Verify the file can be loaded and has valid structure"""
        try:
            wb = openpyxl.load_workbook(self.filepath, data_only=True, read_only=True)
            
            # Check if file has content
            total_cells = 0
            for ws in wb.worksheets:
                total_cells += ws.max_row * ws.max_column
            
            wb.close()
            
            if total_cells == 0:
                print(f"⚠️  Warning: File appears empty")
                return False
            
            print(f"✅ File integrity check passed: {len(wb.sheetnames)} sheets")
            return True
            
        except Exception as e:
            print(f"❌ File integrity check failed: {e}")
            return False
    
    def get_target_sheet(self, sheet_name='A1_ENTRIES'):
        """Get or create target sheet safely"""
        try:
            # Load workbook for editing
            wb = openpyxl.load_workbook(self.filepath)
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"✓ Using existing sheet: {sheet_name}")
            else:
                # Create new sheet at the end
                ws = wb.create_sheet(sheet_name)
                print(f"✓ Created new sheet: {sheet_name}")
                
                # Add headers if it's a new sheet
                headers = ['ENTRY_ID', 'SCORE', 'EN_TERM', 'AR_WORD', 'ROOT_ID', 
                          'ROOT_LETTERS', 'QUR_MEANING', 'PATTERN', 'NETWORK_ID',
                          'PHONETIC_CHAIN', 'INVERSION_TYPE', 'SOURCE_FORM', 'FOUNDATION_REF']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                print(f"✓ Added headers to new sheet")
            
            wb.close()
            return True
            
        except Exception as e:
            print(f"❌ Error accessing sheet {sheet_name}: {e}")
            return False
    
    def write_entries_safely(self, entries, sheet_name='A1_ENTRIES'):
        """Write entries with proper error handling and recovery"""
        if not entries:
            print("💾 No entries to write")
            return 0
        
        # Create temp file for safe writing
        temp_file = self.filepath.parent / f"temp_{self.filepath.name}"
        
        try:
            print(f"💾 Writing {len(entries)} entries safely...")
            
            # Load source workbook
            wb = openpyxl.load_workbook(self.filepath)
            
            # Get or create target sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                # Add headers
                headers = ['ENTRY_ID', 'SCORE', 'EN_TERM', 'AR_WORD', 'ROOT_ID', 
                          'ROOT_LETTERS', 'QUR_MEANING', 'PATTERN', 'NETWORK_ID',
                          'PHONETIC_CHAIN', 'INVERSION_TYPE', 'SOURCE_FORM', 'FOUNDATION_REF']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
            
            # Find next empty row
            next_row = ws.max_row + 1 if ws.max_row > 1 else 2
            
            # Write entries
            written_count = 0
            for entry in entries:
                try:
                    ws.cell(row=next_row, column=1, value=entry.get('ENTRY_ID', ''))
                    ws.cell(row=next_row, column=2, value=entry.get('SCORE', 0))
                    ws.cell(row=next_row, column=3, value=entry.get('EN_TERM', ''))
                    ws.cell(row=next_row, column=4, value=entry.get('AR_WORD', ''))
                    ws.cell(row=next_row, column=5, value=entry.get('ROOT_ID', ''))
                    ws.cell(row=next_row, column=6, value=entry.get('ROOT_LETTERS', '—'))
                    ws.cell(row=next_row, column=7, value=entry.get('QUR_MEANING', '—'))
                    ws.cell(row=next_row, column=8, value=entry.get('PATTERN', 'A'))
                    ws.cell(row=next_row, column=9, value=entry.get('NETWORK_ID', ''))
                    ws.cell(row=next_row, column=10, value=entry.get('PHONETIC_CHAIN', '—'))
                    ws.cell(row=next_row, column=11, value=entry.get('INVERSION_TYPE', 'HIDDEN'))
                    ws.cell(row=next_row, column=12, value=entry.get('SOURCE_FORM', '—'))
                    ws.cell(row=next_row, column=13, value=entry.get('FOUNDATION_REF', '—'))
                    
                    next_row += 1
                    written_count += 1
                    
                except Exception as e:
                    print(f"  ⚠️  Skipping entry due to error: {e}")
                    continue
            
            # Save to temp file first
            wb.save(temp_file)
            
            # Verify temp file
            try:
                test_wb = openpyxl.load_workbook(temp_file, read_only=True)
                test_wb.close()
                
                # If verification passes, replace original
                shutil.move(temp_file, self.filepath)
                self.last_successful_write = datetime.now()
                
                print(f"  ✓ Successfully wrote {written_count} entries")
                return written_count
                
            except Exception as e:
                print(f"  ❌ Temp file verification failed: {e}")
                if temp_file.exists():
                    temp_file.unlink()
                return 0
                
        except Exception as e:
            print(f"  ❌ Error during write operation: {e}")
            # Clean up temp file if it exists
            if temp_file.exists():
                temp_file.unlink()
            return 0
        
        finally:
            # Ensure temp file is cleaned up
            if temp_file.exists():
                try:
                    temp_file.unlink()
                except:
                    pass


# ============================================================================
# ACCELERATED PATTERN DETECTOR (Modified)
# ============================================================================

class AcceleratedDetector:
    """Finds gaps with safety limits"""
    
    def detect_safely(self, reader, max_ratios=20, max_networks=10):
        print("\n🔍 Detecting patterns (safe mode)...")
        
        gaps = []
        candidates = []
        
        # Find roots with multiple entries (limit to max_networks)
        all_roots = {}
        for node_id, node in reader.graph.items():
            root = node.get('root', '')
            if root:
                if root not in all_roots:
                    all_roots[root] = []
                all_roots[root].append(node_id)
        
        # Create limited number of network candidates
        root_count = 0
        for root, entries in all_roots.items():
            if len(entries) >= 2 and root_count < max_networks:
                candidates.append({
                    'type': 'instant_network',
                    'root': root,
                    'entries': entries,
                    'count': len(entries)
                })
                root_count += 1
        
        # Find limited number of ratio gaps
        ratios = [
            (4,3,'musical fourth', 'N08'),
            (5,3,'musical sixth', 'N08'),
            (7,3,'maqam proportion', 'N08'),
            (8,3,'angelic proportion', 'N07'),
            (10,3,'sacred tenth', 'N05'),
            (11,3,'falaq third', 'N16'),
            (7,5,'prayer kernel', 'N06'),
            (8,5,'sacred 8/5', 'N07'),
            (11,5,'dombra sacred', 'N16'),
            (22,7,'circle perfect', 'N07'),
            (28,7,'lunar perfect', 'N05')
        ]
        
        for i, (num, den, name, net) in enumerate(ratios):
            if i < max_ratios:
                gaps.append({
                    'type': 'ratio_needed',
                    'ratio': f"{num}/{den}",
                    'numerator': num,
                    'denominator': den,
                    'name': name,
                    'network': net,
                    'priority': 10
                })
        
        print(f"  Found {len(gaps)} ratio gaps, {len(candidates)} network candidates")
        return gaps, candidates


# ============================================================================
# ACCELERATED GENERATOR (Modified)
# ============================================================================

class AcceleratedGenerator:
    """Generates entries safely with validation"""
    
    def __init__(self, reader):
        self.reader = reader
        self.next_id = self._find_next_id()
        self.new_entries = []
        
    def _find_next_id(self):
        """Find the next available ENTRY_ID"""
        max_num = 0
        for node_id in self.reader.graph.keys():
            if node_id and node_id.startswith('F'):
                try:
                    num = int(node_id[1:])
                    max_num = max(max_num, num)
                except:
                    pass
        return max_num + 1
    
    def generate_safely(self, gaps, candidates, max_entries_per_cycle=15):
        """Generate entries with limits to avoid overwhelming the file"""
        print(f"\n🌱 Generating up to {max_entries_per_cycle} entries...")
        
        self.new_entries = []
        
        # Generate ratio entries from gaps
        ratio_count = 0
        for gap in gaps:
            if gap['type'] == 'ratio_needed' and ratio_count < max_entries_per_cycle // 2:
                entry = self._create_ratio_entry(
                    gap['numerator'], 
                    gap['denominator'],
                    gap.get('name', f"ratio {gap['numerator']}/{gap['denominator']}"),
                    gap.get('network', 'N08')
                )
                if entry:
                    self.new_entries.append(entry)
                    ratio_count += 1
        
        # Generate network entries from candidates
        network_count = 0
        for cand in candidates:
            if cand['type'] == 'instant_network' and network_count < max_entries_per_cycle // 2:
                entry = self._create_network_entry(cand)
                if entry:
                    self.new_entries.append(entry)
                    network_count += 1
        
        print(f"  Generated {len(self.new_entries)} entries (ratios: {ratio_count}, networks: {network_count})")
        return self.new_entries
    
    def _create_ratio_entry(self, num, den, name, network):
        """Create a ratio entry with validation"""
        try:
            entry_id = f"F{self.next_id:04d}"
            self.next_id += 1
            
            # Calculate ratio value for shadow detection
            ratio_value = num / den if den != 0 else 0
            
            # Determine which Western shadow this replaces
            shadow = ''
            if abs(ratio_value - 1.4142) < 0.1:
                shadow = '√2'
            elif abs(ratio_value - 1.732) < 0.1:
                shadow = '√3'
            elif abs(ratio_value - 2.236) < 0.1:
                shadow = '√5'
            elif abs(ratio_value - 3.14159) < 0.1:
                shadow = 'π'
            elif abs(ratio_value - 2.71828) < 0.1:
                shadow = 'e'
            elif abs(ratio_value - 1.618) < 0.1:
                shadow = 'φ (REJECTED)'
            
            shadow_note = f" — replaces {shadow}" if shadow else ''
            
            return {
                'ENTRY_ID': entry_id,
                'SCORE': min(10, 8 + (len(shadow) > 0)),  # Higher score if replaces a Western constant
                'EN_TERM': f"{name} ({num}/{den})",
                'AR_WORD': f"نسبة {num}/{den}",
                'ROOT_ID': f"R{self.next_id}",
                'ROOT_LETTERS': '—',
                'QUR_MEANING': f"Divine ratio {num}/{den} — created by precise measure (Q54:49){shadow_note}",
                'PATTERN': 'A',
                'NETWORK_ID': network if network else 'N08',
                'PHONETIC_CHAIN': '→'.join(['✓'] * 3),
                'INVERSION_TYPE': 'DIVINE',
                'SOURCE_FORM': f"ratio:{num}/{den}",
                'FOUNDATION_REF': f"Q54:49 + {den}-denominator family"
            }
        except Exception as e:
            print(f"  ⚠️  Error creating ratio entry {num}/{den}: {e}")
            return None
    
    def _create_network_entry(self, cand):
        """Create a network entry with validation"""
        try:
            entry_id = f"F{self.next_id:04d}"
            self.next_id += 1
            
            root = cand.get('root', 'UNKNOWN')
            entries = cand.get('entries', [])
            
            return {
                'ENTRY_ID': entry_id,
                'SCORE': min(10, 7 + min(len(entries), 3)),  # Higher score for more connections
                'EN_TERM': f"Network for root {root}",
                'AR_WORD': f"شبكة {root}",
                'ROOT_ID': root,
                'ROOT_LETTERS': root,
                'QUR_MEANING': f"Network connecting {len(entries)} entries with same root",
                'PATTERN': 'A',
                'NETWORK_ID': f"N{self.next_id}",
                'PHONETIC_CHAIN': '→'.join(['✓'] * 3),
                'INVERSION_TYPE': 'DIVINE',
                'SOURCE_FORM': f"root:{root}",
                'FOUNDATION_REF': f"F2: All scripts from same root"
            }
        except Exception as e:
            print(f"  ⚠️  Error creating network entry: {e}")
            return None


# ============================================================================
# READER FROM uslap_forest.py (Modified to work without A1_ENTRIES)
# ============================================================================

class SafeSelfReader:
    """Modified reader that works with any sheet structure"""
    
    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self.wb = None
        self.graph = {}
        self.sheets_data = {}
        self.total_entries = 0
        
    def read_all(self):
        """Read the workbook and build a simple graph from any data"""
        print(f"\n📖 Reading {self.filepath.name}...")
        
        if not self.filepath.exists():
            raise FileNotFoundError(f"File not found: {self.filepath}")
        
        try:
            self.wb = openpyxl.load_workbook(self.filepath, data_only=True, read_only=True)
            
            # Read all sheets
            for sheet_name in self.wb.sheetnames:
                sheet = self.wb[sheet_name]
                data = []
                
                # Try to find data rows (skip empty rows)
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        # Convert row to dict with column letters as keys
                        row_dict = {}
                        for i, value in enumerate(row, 1):
                            if value is not None:
                                col_letter = openpyxl.utils.get_column_letter(i)
                                row_dict[col_letter] = value
                        if row_dict:
                            data.append(row_dict)
                
                self.sheets_data[sheet_name] = data
                print(f"  ✓ {sheet_name}: {len(data)} data rows")
            
            # Build a simple graph from any data that looks like entries
            self._build_simple_graph()
            
            self.wb.close()
            return self
            
        except Exception as e:
            print(f"❌ Error reading workbook: {e}")
            if self.wb:
                self.wb.close()
            raise
    
    def _build_simple_graph(self):
        """Build a simple graph from any data that has ID-like fields"""
        self.graph = {}
        
        for sheet_name, data in self.sheets_data.items():
            for row_idx, row in enumerate(data):
                # Look for any field that might be an ID
                for key, value in row.items():
                    if value and isinstance(value, str):
                        # Check if value looks like an ID (F followed by numbers, or similar)
                        if (value.startswith('F') and value[1:].isdigit()) or \
                           (value.startswith('N') and value[1:].isdigit()) or \
                           (value.startswith('R') and value[1:].isdigit()):
                            
                            node_id = f"{sheet_name}_{row_idx}_{value}"
                            self.graph[node_id] = {
                                'id': value,
                                'sheet': sheet_name,
                                'row': row_idx + 1,
                                'data': row
                            }
        
        self.total_entries = len(self.graph)
        print(f"📊 Found {self.total_entries} entry-like nodes across all sheets")
        
        return self.graph


# ============================================================================
# SAFE ACCELERATED SYSTEM
# ============================================================================

class SafeAcceleratedSystem:
    """Runs accelerated growth with safety features"""
    
    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self.file_handler = SafeFileHandler(filepath)
        self.cycle_count = 0
        self.total_written = 0
        
    def run_safely(self, cycles=10, entries_per_cycle=15):
        """Run accelerated growth with safety limits"""
        print("\n" + "="*60)
        print(f"🛡️  SAFE ACCELERATED GROWTH: {cycles} CYCLES")
        print("="*60)
        
        # Step 1: Create backup
        print("\n📁 Step 1: Creating backup...")
        if not self.file_handler.create_backup():
            print("⚠️  Continuing without backup (user choice)")
            response = input("Continue without backup? (y/n): ").strip().lower()
            if response != 'y':
                print("Aborted by user")
                return
        
        # Step 2: Verify file integrity
        print("\n✅ Step 2: Verifying file integrity...")
        if not self.file_handler.verify_file_integrity():
            print("⚠️  File integrity check failed. Consider restoring from backup.")
            response = input("Continue anyway? (y/n): ").strip().lower()
            if response != 'y':
                print("Aborted by user")
                return
        
        # Step 3: Ensure target sheet exists
        print("\n📋 Step 3: Preparing target sheet...")
        if not self.file_handler.get_target_sheet('A1_ENTRIES'):
            print("⚠️  Could not prepare target sheet")
            return
        
        # Step 4: Initialize reader
        print("\n📖 Step 4: Initializing reader...")
        try:
            reader = SafeSelfReader(self.filepath)
            reader.read_all()
        except Exception as e:
            print(f"❌ Failed to initialize reader: {e}")
            return
        
        # Step 5: Run cycles
        print("\n🚀 Step 5: Running growth cycles...")
        start_time = time.time()
        
        for cycle in range(1, cycles + 1):
            print(f"\n--- Cycle {cycle}/{cycles} ---")
            
            try:
                # Detect patterns
                detector = AcceleratedDetector()
                gaps, candidates = detector.detect_safely(reader, max_ratios=10, max_networks=5)
                
                # Generate entries
                generator = AcceleratedGenerator(reader)
                new_entries = generator.generate_safely(gaps, candidates, max_entries_per_cycle=entries_per_cycle)
                
                # Write entries
                if new_entries:
                    written = self.file_handler.write_entries_safely(new_entries, 'A1_ENTRIES')
                    self.total_written += written
                    
                    # Update reader with new data
                    if written > 0:
                        print(f"  ↻ Refreshing reader...")
                        reader = SafeSelfReader(self.filepath)
                        reader.read_all()
                else:
                    print("  ⏭️  No new entries generated this cycle")
                    written = 0
                
                # Progress tracking
                elapsed = time.time() - start_time
                avg_time = elapsed / cycle
                remaining = avg_time * (cycles - cycle)
                
                print(f"  📈 Cycle {cycle}: +{written} entries | Total: {self.total_written} | Elapsed: {elapsed:.1f}s | Remaining: {remaining:.1f}s")
                
                # Safety pause between cycles
                if cycle < cycles:
                    time.sleep(1.0)  # Longer pause for safety
                
            except Exception as e:
                print(f"  ❌ Error in cycle {cycle}: {e}")
                print(f"  ⚠️  Skipping to next cycle...")
                time.sleep(2.0)  # Longer pause after error
                continue
        
        # Final summary
        total_time = time.time() - start_time
        print("\n" + "="*60)
        print(f"✅ SAFE GROWTH COMPLETE")
        print(f"   Cycles: {cycles}")
        print(f"   Total entries added: {self.total_written}")
        print(f"   Total time: {total_time:.1f} seconds")
        print(f"   Average: {total_time/cycles:.1f} seconds per cycle")
        print(f"   Backup: {self.file_handler.backup_path.name if self.file_handler.backup_path else 'None'}")
        print("="*60)
        
        # Final file check
        print("\n🔍 Final file check...")
        final_size = self.filepath.stat().st_size if self.filepath.exists() else 0
        print(f"   Final file size: {final_size:,} bytes")
        
