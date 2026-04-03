#!/usr/bin/env python3
"""
USLaP SAFE ACCELERATED FOREST GROWTH - FINAL VERSION
Safe accelerated growth with automatic backups and corruption protection

Usage:
  python3 uslap_accelerated_safe_final.py [cycles=10] [entries_per_cycle=10]

Example:
  python3 uslap_accelerated_safe_final.py 5 10  # 5 cycles, 10 entries per cycle
  python3 uslap_accelerated_safe_final.py 100   # 100 cycles, default 10 entries per cycle
  python3 uslap_accelerated_safe_final.py       # Default: 10 cycles, 10 entries per cycle
"""

import openpyxl
from pathlib import Path
from datetime import datetime
import time
import shutil
import sys
import os

# ============================================================================
# SAFE FILE HANDLER
# ============================================================================

class SafeFileHandler:
    """Handles file operations safely with backups"""
    
    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self.backup_path = None
        
    def create_backup(self):
        """Create a timestamped backup before starting operations"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{self.filepath.stem}_backup_{timestamp}{self.filepath.suffix}"
        self.backup_path = self.filepath.parent / backup_name
        
        try:
            shutil.copy2(self.filepath, self.backup_path)
            print(f"📁 Created backup: {self.backup_path.name}")
            return True
        except Exception as e:
            print(f"❌ Failed to create backup: {e}")
            return False
    
    def verify_file_integrity(self):
        """Verify the file can be loaded and has valid structure"""
        try:
            wb = openpyxl.load_workbook(self.filepath, data_only=True, read_only=True)
            wb.close()
            print(f"✅ File integrity check passed: {len(wb.sheetnames)} sheets")
            return True
        except Exception as e:
            print(f"❌ File integrity check failed: {e}")
            return False
    
    def write_entries_safely(self, entries, sheet_name='A1_ENTRIES'):
        """Write entries with proper error handling and recovery"""
        if not entries:
            return 0
        
        # Create temp file for safe writing
        temp_file = self.filepath.parent / f"temp_{self.filepath.name}"
        
        try:
            # Load source workbook
            wb = openpyxl.load_workbook(self.filepath)
            
            # Get or create target sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                # Add headers
                headers = ['ENTRY_ID', 'SCORE', 'EN_TERM', 'AR_WORD', 'ROOT_ID', 
                          'ROOT_LETTERS', 'QUR_MEANING', 'PATTERN', 'NETWORK_ID',
                          'PHONETIC_CHAIN', 'INVERSION_TYPE', 'SOURCE_FORM', 'FOUNDATION_REF']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
            
            # Find next empty row
            next_row = ws.max_row + 1 if ws.max_row > 1 else 2
            
            # Write entries
            written_count = 0
            for entry in entries:
                try:
                    ws.cell(row=next_row, column=1, value=entry.get('ENTRY_ID', ''))
                    ws.cell(row=next_row, column=2, value=entry.get('SCORE', 0))
                    ws.cell(row=next_row, column=3, value=entry.get('EN_TERM', ''))
                    ws.cell(row=next_row, column=4, value=entry.get('AR_WORD', ''))
                    ws.cell(row=next_row, column=5, value=entry.get('ROOT_ID', ''))
                    ws.cell(row=next_row, column=6, value=entry.get('ROOT_LETTERS', '—'))
                    ws.cell(row=next_row, column=7, value=entry.get('QUR_MEANING', '—'))
                    ws.cell(row=next_row, column=8, value=entry.get('PATTERN', 'A'))
                    ws.cell(row=next_row, column=9, value=entry.get('NETWORK_ID', ''))
                    ws.cell(row=next_row, column=10, value=entry.get('PHONETIC_CHAIN', '—'))
                    ws.cell(row=next_row, column=11, value=entry.get('INVERSION_TYPE', 'HIDDEN'))
                    ws.cell(row=next_row, column=12, value=entry.get('SOURCE_FORM', '—'))
                    ws.cell(row=next_row, column=13, value=entry.get('FOUNDATION_REF', '—'))
                    
                    next_row += 1
                    written_count += 1
                except Exception as e:
                    print(f"  ⚠️  Skipping entry due to error: {e}")
                    continue
            
            # Save to temp file first
            wb.save(temp_file)
            
            # Verify temp file
            try:
                test_wb = openpyxl.load_workbook(temp_file, read_only=True)
                test_wb.close()
                
                # If verification passes, replace original
                shutil.move(temp_file, self.filepath)
                return written_count
            except Exception as e:
                print(f"  ❌ Temp file verification failed: {e}")
                if temp_file.exists():
                    temp_file.unlink()
                return 0
                
        except Exception as e:
            print(f"  ❌ Error during write operation: {e}")
            return 0
        finally:
            # Clean up temp file if it exists
            if temp_file.exists():
                try:
                    temp_file.unlink()
                except:
                    pass

# ============================================================================
# SIMPLE DATA GENERATOR
# ============================================================================

class SimpleGenerator:
    """Generates simple ratio entries for testing"""
    
    def __init__(self):
        self.next_id = 1000  # Start from a high number to avoid conflicts
        
    def generate_ratio_entries(self, count=10):
        """Generate simple ratio entries"""
        entries = []
        
        # Common divine ratios from USLaP
        ratios = [
            (4, 3, 'musical fourth', 'N08'),
            (5, 3, 'musical sixth', 'N08'),
            (7, 3, 'maqam proportion', 'N08'),
            (7, 5, 'prayer kernel', 'N06'),
            (11, 5, 'dombra sacred', 'N16'),
            (22, 7, 'circle perfect', 'N07'),
            (28, 7, 'lunar perfect', 'N05'),
            (99, 70, 'high precision √2', 'N05'),
            (355, 113, 'high precision π', 'N05'),
            (19, 7, 'growth constant', 'N10'),
        ]
        
        for i, (num, den, name, network) in enumerate(ratios[:count]):
            entry_id = f"F{self.next_id:04d}"
            self.next_id += 1
            
            entries.append({
                'ENTRY_ID': entry_id,
                'SCORE': 9,
                'EN_TERM': f"{name} ({num}/{den})",
                'AR_WORD': f"نسبة {num}/{den}",
                'ROOT_ID': f"R{self.next_id}",
                'ROOT_LETTERS': '—',
                'QUR_MEANING': f"Divine ratio {num}/{den} — precise measure (Q54:49)",
                'PATTERN': 'A',
                'NETWORK_ID': network,
                'PHONETIC_CHAIN': '→'.join(['✓'] * 3),
                'INVERSION_TYPE': 'DIVINE',
                'SOURCE_FORM': f"ratio:{num}/{den}",
                'FOUNDATION_REF': f"Q54:49 + {den}-denominator"
            })
        
        return entries

# ============================================================================
# MAIN ACCELERATED GROWTH SYSTEM
# ============================================================================

def run_safe_growth(filepath, cycles=10, entries_per_cycle=10):
    """Run safe accelerated growth cycles"""
    
    filepath = Path(filepath)
    if not filepath.exists():
        print(f"❌ File not found: {filepath}")
        return False
    
    print("\n" + "="*60)
    print(f"🛡️  SAFE ACCELERATED GROWTH")
    print("="*60)
    print(f"📁 File: {filepath.name}")
    print(f"📦 Size: {filepath.stat().st_size / 1024 / 1024:.2f} MB")
    print(f"🚀 Cycles: {cycles}")
    print(f"🌱 Entries per cycle: {entries_per_cycle}")
    print("="*60)
    
    # Initialize handlers
    file_handler = SafeFileHandler(filepath)
    generator = SimpleGenerator()
    
    # Step 1: Create backup
    print("\n📁 Step 1: Creating backup...")
    if not file_handler.create_backup():
        print("⚠️  Backup creation failed!")
        response = input("Continue without backup? (y/n): ").strip().lower()
        if response != 'y':
            print("Aborted by user")
            return False
    
    # Step 2: Verify file integrity
    print("\n✅ Step 2: Verifying file integrity...")
    if not file_handler.verify_file_integrity():
        print("⚠️  File integrity check failed!")
        response = input("Continue anyway? (y/n): ").strip().lower()
        if response != 'y':
            print("Aborted by user")
            return False
    
    # Step 3: Run growth cycles
    print("\n🚀 Step 3: Running growth cycles...")
    start_time = time.time()
    total_written = 0
    
    for cycle in range(1, cycles + 1):
        print(f"\n--- Cycle {cycle}/{cycles} ---")
        
        try:
            # Generate entries for this cycle
            entries = generator.generate_ratio_entries(entries_per_cycle)
            
            # Write entries safely
            written = file_handler.write_entries_safely(entries, 'A1_ENTRIES')
            total_written += written
            
            # Progress tracking
            elapsed = time.time() - start_time
            avg_time = elapsed / cycle
            remaining = avg_time * (cycles - cycle)
            
            print(f"  📈 +{written} entries | Total: {total_written} | Elapsed: {elapsed:.1f}s | Remaining: {remaining:.1f}s")
            
            # Safety pause between cycles
            if cycle < cycles:
                time.sleep(0.5)  # Short pause for safety
            
        except Exception as e:
            print(f"  ❌ Error in cycle {cycle}: {e}")
            print(f"  ⚠️  Skipping to next cycle...")
            time.sleep(1.0)
            continue
    
    # Final summary
    total_time = time.time() - start_time
    print("\n" + "="*60)
    print(f"✅ SAFE GROWTH COMPLETE")
    print(f"   Cycles completed: {cycles}")
    print(f"   Total entries added: {total_written}")
    print(f"   Total time: {total_time:.1f} seconds")
    print(f"   Average: {total_time/cycles:.1f} seconds per cycle")
    
    if file_handler.backup_path and file_handler.backup_path.exists():
        print(f"   Backup saved: {file_handler.backup_path.name}")
    
    # Final file check
    final_size = filepath.stat().st_size if filepath.exists() else 0
    print(f"   Final file size: {final_size:,} bytes")
    print("="*60)
    
    return True

# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

def main():
    """Main command line interface"""
    
    # Default values
    cycles = 10
    entries_per_cycle = 10
    filepath = "USLaP_Final_Data_Consolidated_Master.xlsx"
    
    # Parse command line arguments
    if len(sys.argv) > 1:
        try:
            cycles = int(sys.argv[1])
        except ValueError:
            print(f"⚠️  Invalid cycle count: {sys.argv[1]}. Using default: {cycles}")
    
    if len(sys.argv) > 2:
        try:
            entries_per_cycle = int(sys.argv[2])
        except ValueError:
            print(f"⚠️  Invalid entries per cycle: {sys.argv[2]}. Using default: {entries_per_cycle}")
    
    # Check if file exists
    if not os.path.exists(filepath):
        print(f"\n❌ File not found: {filepath}")
        print("\nPlease provide the path to your USLaP Excel file:")
        print(f"  python3 {sys.argv[0]} /path/to/your/file.xlsx [cycles] [entries_per_cycle]")
        print("\nOr place this script in the same folder as your file.")
        return 1
    
    # Show configuration
    print(f"\n⚙️  Configuration:")
    print(f"   File: {filepath}")
    print(f"   Cycles: {cycles}")
    print(f"   Entries per cycle: {entries_per_cycle}")
    print(f"   Total entries to add: ~{cycles * entries_per_cycle}")
    
    # Confirm with user
    print(f"\n⚠️  This will modify {filepath} and create a backup.")
    response = input("Continue? (y/n): ").strip().lower()
    if response != 'y':
        print("Operation cancelled")
        return 0
    
    # Run safe growth
    try:
        success = run_safe_growth(filepath, cycles, entries_per_cycle)
        return 0 if success else 1
    except KeyboardInterrupt:
        print("\n\n👋 Operation interrupted by user")
        return 1
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return 1

# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    sys.exit(main())