#!/usr/bin/env python3
"""
USLaP Migration with All Fixes:
1. Creates database with corrected schema (xref_id as TEXT in cross_refs)
2. Registers extract_consonants() UDF before any operations
3. Migrates all data from Excel
4. Fixes Russian, Persian, and Bitig entries as NEW entries (not updates)
5. Includes quran_refs migration (A3_QURAN_REFS sheet)
6. Handles BITIG entries as new entries with ORIG2 prefix

This script combines migrate_to_sqlite_fixed.py and fix_missing_migration.py
with all necessary fixes applied.
"""

import sqlite3
import openpyxl
import re
import sys
import os
import shutil
import json
from datetime import datetime
from pathlib import Path

# Add the current directory to sys.path to import USLaP_Engine
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ============================================================================
# GLOBAL CONFIGURATION
# ============================================================================

EXCEL_PATH = "USLaP_Final_Data_Consolidated_Master_v3.xlsx"
DB_PATH = "Code_files/uslap_lattice.db"
SCHEMA_PATH = "Code_files/create_uslap_db.sql"
BACKUP_DIR = "Code_files/backups"

# Sheets to migrate (structured data only, skip EXCEL_DATA_CONSOLIDATED)
SHEETS_TO_MIGRATE = [
    "A1_ENTRIES",
    "A1_ЗАПИСИ",
    "PERSIAN_A1_MADĀKHIL", 
    "BITIG_A1_ENTRIES",
    "CHILD_SCHEMA",
    "A4_DERIVATIVES",
    "A5_CROSS_REFS",
    "A3_QURAN_REFS",
    "M1_PHONETIC_SHIFTS",
    "M2_DETECTION_PATTERNS",
    "M4_NETWORKS",
    "M3_SCHOLARS",
    "M5_QUR_VERIFICATION",
    "A2_NAMES_OF_ALLAH",
]

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def clean_column_name(col):
    """Convert Excel column header to valid SQLite column name."""
    if col is None:
        return "unknown"
    col = str(col).strip()
    col = re.sub(r'[^\w\s]', '', col)  # Remove punctuation
    col = re.sub(r'\s+', '_', col)     # Replace spaces with underscore
    col = col.lower()
    if not col:
        return "unknown"
    return col

def backup_database(db_path):
    """Create a timestamped backup of existing database."""
    if not os.path.exists(db_path):
        return None
    
    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(BACKUP_DIR, f"{os.path.basename(db_path)}_backup_{timestamp}.db")
    shutil.copy2(db_path, backup_path)
    print(f"  Created backup: {backup_path}")
    return backup_path

# ============================================================================
# EXTRACT_CONSONANTS UDF (from USLaP_Engine.py)
# ============================================================================

def extract_consonants(word):
    """
    Python UDF for SQLite: Extract consonant skeleton from a word.
    Must be registered before any operations on word_fingerprints table.
    
    This is the same logic as PhoneticReversal.extract_consonants() in USLaP_Engine.py.
    """
    if not word:
        return ""
    
    # Direct implementation
    vowels = set('aeiou')
    result = []
    i = 0
    word_lower = word.lower()
    
    while i < len(word_lower):
        digraph = word_lower[i:i+2] if i + 1 < len(word_lower) else ''
        if digraph in ('sh', 'ch', 'gh', 'th', 'ph', 'wh', 'qu'):
            result.append(digraph)
            i += 2
        elif word_lower[i] not in vowels:
            result.append(word_lower[i])
            i += 1
        else:
            i += 1
    
    return ''.join(result)

# ============================================================================
# FIXED SCHEMA CREATION (with xref_id as TEXT)
# ============================================================================

def create_fixed_schema(conn, cursor):
    """Create database schema with corrected cross_refs table (xref_id as TEXT)."""
    print("  Creating database with fixed schema...")
    
    # Read the original schema
    if not os.path.exists(SCHEMA_PATH):
        print(f"    ERROR: Schema file not found: {SCHEMA_PATH}")
        return False
    
    with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
        schema_sql = f.read()
    
    # Fix the cross_refs table definition
    # Original: xref_id INTEGER PRIMARY KEY AUTOINCREMENT,
    # Fixed: xref_id TEXT PRIMARY KEY,
    fixed_schema = schema_sql.replace(
        'xref_id             INTEGER PRIMARY KEY AUTOINCREMENT,',
        'xref_id             TEXT PRIMARY KEY,'  # Changed to TEXT, removed AUTOINCREMENT
    )
    
    # Execute the fixed schema
    cursor.executescript(fixed_schema)
    
    # Verify the fix
    cursor.execute("PRAGMA table_info(cross_refs)")
    columns = cursor.fetchall()
    xref_col = [col for col in columns if col[1] == 'xref_id'][0]
    if xref_col[2] == 'TEXT':  # column type
        print("    ✓ Fixed: cross_refs.xref_id is now TEXT (was INTEGER)")
    else:
        print(f"    ⚠ Warning: cross_refs.xref_id type is {xref_col[2]}, expected TEXT")
    
    return True

# ============================================================================
# MIGRATION FUNCTIONS (from migrate_to_sqlite_fixed.py)
# ============================================================================

def migrate_a1_entries(conn, cursor, wb):
    """Migrate A1_ENTRIES sheet to entries table."""
    print("  Migrating A1_ENTRIES...")
    
    if "A1_ENTRIES" not in wb.sheetnames:
        print("    Sheet A1_ENTRIES not found")
        return 0
    
    ws = wb["A1_ENTRIES"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A1_ENTRIES")
        return 0
    
    # First row is headers
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map to entries table schema
        entry_data = {
            'entry_id': row_dict.get('entry_id'),
            'score': row_dict.get('score'),
            'en_term': row_dict.get('en_term'),
            'ar_word': row_dict.get('ar_word'),
            'root_letters': row_dict.get('root_letters'),
            'qur_meaning': row_dict.get('qur_meaning'),
            'pattern': row_dict.get('pattern'),
            'allah_name_id': row_dict.get('allah_name_id'),
            'network_id': row_dict.get('network_id'),
            'phonetic_chain': row_dict.get('phonetic_chain'),
            'inversion_type': row_dict.get('inversion_type'),
            'source_form': row_dict.get('source_form'),
            'foundation_refs': row_dict.get('foundation_ref'),
        }
        
        # Extract root ID from root_letters or root_id
        root_id = row_dict.get('root_id')
        if root_id:
            # Fix root ID padding: convert R01 to R001 format
            if root_id and re.match(r'^R\d{1,3}$', root_id):
                match = re.match(r'^R(\d+)$', root_id)
                if match:
                    num = int(match.group(1))
                    root_id = f'R{num:03d}'
            entry_data['root_id'] = root_id
        
        # Extract DS corridor from foundation_ref
        foundation_ref = row_dict.get('foundation_ref', '')
        ds_match = re.search(r'DS\d+', str(foundation_ref))
        if ds_match:
            entry_data['ds_corridor'] = ds_match.group(0)
        
        # Insert into entries table
        try:
            cursor.execute('''
                INSERT INTO entries (
                    entry_id, score, en_term, ar_word, root_id, root_letters,
                    qur_meaning, pattern, allah_name_id, network_id, phonetic_chain,
                    inversion_type, source_form, foundation_refs, ds_corridor
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                entry_data['entry_id'], entry_data['score'], entry_data['en_term'],
                entry_data['ar_word'], entry_data.get('root_id'), entry_data['root_letters'],
                entry_data['qur_meaning'], entry_data['pattern'], entry_data['allah_name_id'],
                entry_data['network_id'], entry_data['phonetic_chain'], entry_data['inversion_type'],
                entry_data['source_form'], entry_data['foundation_refs'], entry_data.get('ds_corridor')
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {entry_data}")
    
    print(f"    Migrated {count} entries")
    return count

def migrate_a1_zapisi(conn, cursor, wb):
    """Migrate A1_ЗАПИСИ sheet to entries table (Russian entries)."""
    print("  Migrating A1_ЗАПИСИ (Russian entries)...")
    
    if "A1_ЗАПИСИ" not in wb.sheetnames:
        print("    Sheet A1_ЗАПИСИ not found")
        return 0
    
    ws = wb["A1_ЗАПИСИ"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A1_ЗАПИСИ")
        return 0
    
    # First row is headers (Russian column names)
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map Russian column names to entries table schema
        entry_data = {
            'entry_id': row_dict.get('запись_id'),
            'score': row_dict.get('балл'),
            'ru_term': row_dict.get('рус_термин'),
            'ar_word': row_dict.get('ар_слово'),
            'root_letters': row_dict.get('корневые_буквы'),
            'qur_meaning': row_dict.get('коранич_значение'),
            'pattern': row_dict.get('паттерн'),
            'allah_name_id': row_dict.get('имя_аллаха_id'),
            'network_id': row_dict.get('сеть_id'),
            'phonetic_chain': row_dict.get('фонетическая_цепь'),
            'inversion_type': row_dict.get('тип_инверсии'),
            'source_form': row_dict.get('исходная_форма'),
            'foundation_refs': row_dict.get('основание'),
        }
        
        # Extract root ID
        root_id = row_dict.get('корень_id')
        if root_id:
            # Fix root ID padding
            if root_id and re.match(r'^R\d{1,3}$', root_id):
                match = re.match(r'^R(\d+)$', root_id)
                if match:
                    num = int(match.group(1))
                    root_id = f'R{num:03d}'
            entry_data['root_id'] = root_id
        
        # Each Russian row is a NEW entry (not an update to existing English entry)
        # Generate a new entry_id if the one from Excel conflicts with existing entries
        excel_entry_id = entry_data['entry_id']
        new_entry_id = excel_entry_id
        
        # Check if entry_id already exists
        cursor.execute('SELECT 1 FROM entries WHERE entry_id = ?', (excel_entry_id,))
        if cursor.fetchone():
            # Conflict: generate new sequential entry_id
            cursor.execute('SELECT MAX(entry_id) FROM entries')
            max_id = cursor.fetchone()[0]
            if max_id:
                new_entry_id = max_id + 1
            else:
                new_entry_id = 1000  # Start Russian entries from 1000 to avoid conflicts
            print(f"    Entry ID conflict: {excel_entry_id} → {new_entry_id} for Russian term: {entry_data['ru_term']}")
        
        # Insert new Russian entry (en_term = NULL, ru_term populated)
        try:
            cursor.execute('''
                INSERT INTO entries (
                    entry_id, score, ru_term, ar_word, root_id, root_letters,
                    qur_meaning, pattern, allah_name_id, network_id, phonetic_chain,
                    inversion_type, source_form, foundation_refs, en_term
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL)
            ''', (
                new_entry_id, entry_data['score'], entry_data['ru_term'],
                entry_data['ar_word'], entry_data.get('root_id'), entry_data['root_letters'],
                entry_data['qur_meaning'], entry_data['pattern'], entry_data['allah_name_id'],
                entry_data['network_id'], entry_data['phonetic_chain'], entry_data['inversion_type'],
                entry_data['source_form'], entry_data['foundation_refs']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {entry_data}")
    
    print(f"    Migrated {count} Russian entries (as new entries with en_term = NULL)")
    return count

def migrate_persian_a1_madakhil(conn, cursor, wb):
    """Migrate PERSIAN_A1_MADĀKHIL sheet to entries table (Persian entries)."""
    print("  Migrating PERSIAN_A1_MADĀKHIL (Persian entries)...")
    
    if "PERSIAN_A1_MADĀKHIL" not in wb.sheetnames:
        print("    Sheet PERSIAN_A1_MADĀKHIL not found")
        return 0
    
    ws = wb["PERSIAN_A1_MADĀKHIL"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in PERSIAN_A1_MADĀKHIL")
        return 0
    
    # First row has complex headers with Persian and English
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map to entries table schema
        entry_data = {
            'entry_id': row_dict.get('madkhal_identry_id'),
            'score': row_dict.get('nomrescore'),
            'fa_term': row_dict.get('vazhe_farsipersian_term'),
            'source_word': row_dict.get('kalame_aslisource_word'),
            'root_letters': row_dict.get('horuf_e_risheroot_letters'),
            'qur_meaning': row_dict.get('mana_ye_quraniquot_meaning'),
            'pattern': row_dict.get('olgu_pattern'),
            'allah_name_id': row_dict.get('esm_e_allah_idallah_name_id'),
            'network_id': row_dict.get('shabake_idnetwork_id'),
            'phonetic_chain': row_dict.get('zanjire_sawtiphoneic_chain'),
            'inversion_type': row_dict.get('now_e_vazhguniinversion_type'),
            'source_form': row_dict.get('shakl_e_aslisource_form'),
            'foundation_refs': row_dict.get('boniyanfoundation_ref'),
        }
        
        # Extract root ID
        root_id = row_dict.get('rishe_idroot_id')
        if root_id:
            # Fix root ID padding
            if root_id and re.match(r'^R\d{1,3}$', root_id):
                match = re.match(r'^R(\d+)$', root_id)
                if match:
                    num = int(match.group(1))
                    root_id = f'R{num:03d}'
            entry_data['root_id'] = root_id
        
        # Each Persian row is a NEW entry (not an update to existing English entry)
        # Generate a new entry_id if the one from Excel conflicts with existing entries
        excel_entry_id = entry_data['entry_id']
        new_entry_id = excel_entry_id
        
        # Check if entry_id already exists
        cursor.execute('SELECT 1 FROM entries WHERE entry_id = ?', (excel_entry_id,))
        if cursor.fetchone():
            # Conflict: generate new sequential entry_id
            cursor.execute('SELECT MAX(entry_id) FROM entries')
            max_id = cursor.fetchone()[0]
            if max_id:
                new_entry_id = max_id + 1
            else:
                new_entry_id = 2000  # Start Persian entries from 2000 to avoid conflicts
            print(f"    Entry ID conflict: {excel_entry_id} → {new_entry_id} for Persian term: {entry_data['fa_term']}")
        
        # Insert new Persian entry (en_term = NULL, fa_term populated)
        try:
            cursor.execute('''
                INSERT INTO entries (
                    entry_id, score, fa_term, ar_word, root_id, root_letters,
                    qur_meaning, pattern, allah_name_id, network_id, phonetic_chain,
                    inversion_type, source_form, foundation_refs, en_term
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL)
            ''', (
                new_entry_id, entry_data['score'], entry_data['fa_term'],
                entry_data['source_word'], entry_data.get('root_id'), entry_data['root_letters'],
                entry_data['qur_meaning'], entry_data['pattern'], entry_data['allah_name_id'],
                entry_data['network_id'], entry_data['phonetic_chain'], entry_data['inversion_type'],
                entry_data['source_form'], entry_data['foundation_refs']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {entry_data}")
    
    print(f"    Migrated {count} Persian entries (as new entries with en_term = NULL)")
    return count

def migrate_bitig_a1_entries(conn, cursor, wb):
    """Migrate BITIG_A1_ENTRIES sheet to entries table (ORIG2 entries)."""
    print("  Migrating BITIG_A1_ENTRIES (ORIG2 entries)...")
    
    if "BITIG_A1_ENTRIES" not in wb.sheetnames:
        print("    Sheet BITIG_A1_ENTRIES not found")
        return 0
    
    ws = wb["BITIG_A1_ENTRIES"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in BITIG_A1_ENTRIES")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map to entries table schema
        excel_entry_id = row_dict.get('entry_id')
        en_term = row_dict.get('orig2_term')
        ar_word = row_dict.get('orig2_script')
        root_letters = row_dict.get('root_letters')
        phonetic_chain = row_dict.get('phonetic_chain')
        notes = row_dict.get('notes')
        score = row_dict.get('score', 5)
        
        # For ORIG2 entries, mark with ORIG2 prefix in foundation_refs
        kashgari_attestation = row_dict.get('kashgari_attestation', '')
        foundation_refs = f"ORIG2: {kashgari_attestation}"
        
        # Each Bitig (ORIG2) row is a NEW entry (not an update to existing English entry)
        # Generate a new entry_id if the one from Excel conflicts with existing entries
        new_entry_id = excel_entry_id
        
        # Check if entry_id already exists
        cursor.execute('SELECT 1 FROM entries WHERE entry_id = ?', (excel_entry_id,))
        if cursor.fetchone():
            # Conflict: generate new sequential entry_id
            cursor.execute('SELECT MAX(entry_id) FROM entries')
            max_id = cursor.fetchone()[0]
            if max_id:
                new_entry_id = max_id + 1
            else:
                new_entry_id = 3000  # Start ORIG2 entries from 3000 to avoid conflicts
            print(f"    Entry ID conflict: {excel_entry_id} → {new_entry_id} for ORIG2 term: {en_term}")
        
        # Extract root ID if present
        root_id = None
        if root_letters:
            # Try to find existing root with these letters
            cursor.execute('SELECT root_id FROM roots WHERE root_letters = ?', (root_letters,))
            result = cursor.fetchone()
            if result:
                root_id = result[0]
            else:
                # Create new root if needed
                cursor.execute('SELECT COUNT(*) FROM roots')
                root_count = cursor.fetchone()[0]
                root_id = f"R{root_count + 1:03d}"
                
                # Create bare root (without hyphens)
                root_bare = re.sub(r'[\-\s]', '', root_letters) if root_letters else ''
                root_type = 'TRILITERAL' if len(root_bare) == 3 else 'QUADRILITERAL' if len(root_bare) == 4 else 'OTHER'
                
                cursor.execute('''
                    INSERT INTO roots (root_id, root_letters, root_bare, root_type, notes)
                    VALUES (?, ?, ?, ?, ?)
                ''', (
                    root_id, root_letters, root_bare, root_type,
                    f"Created during ORIG2 entries migration for entry {new_entry_id}"
                ))
        
        # Insert new ORIG2 entry (en_term populated, foundation_refs has ORIG2 prefix)
        try:
            cursor.execute('''
                INSERT INTO entries (
                    entry_id, score, en_term, ar_word, root_id, root_letters,
                    phonetic_chain, foundation_refs, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                new_entry_id, score, en_term, ar_word, root_id, root_letters,
                phonetic_chain, foundation_refs, notes
            ))
            count += 1
            print(f"    Inserted new ORIG2 entry {new_entry_id}: {en_term}")
        except Exception as e:
            print(f"    Error inserting ORIG2 row {i}: {e}")
            print(f"    Data: entry_id={new_entry_id}, en_term={en_term}")
    
    print(f"    Migrated {count} ORIG2 entries (as new entries with ORIG2 prefix)")
    return count

def migrate_a2_names_of_allah(conn, cursor, wb):
    """Migrate A2_NAMES_OF_ALLAH sheet to names_of_allah table."""
    print("  Migrating A2_NAMES_OF_ALLAH...")
    
    if "A2_NAMES_OF_ALLAH" not in wb.sheetnames:
        print("    Sheet A2_NAMES_OF_ALLAH not found")
        return 0
    
    ws = wb["A2_NAMES_OF_ALLAH"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A2_NAMES_OF_ALLAH")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map columns based on actual Excel headers
        name_data = {
            'name_id': row_dict.get('allah_id'),  # ALLAH_ID column
            'arabic_name': row_dict.get('arabic_name'),
            'transliteration': row_dict.get('transliteration'),
            'meaning': row_dict.get('meaning'),
            'qur_ref': row_dict.get('qur_ref'),
            'entry_ids': row_dict.get('entry_ids'),
            'root_id': row_dict.get('root_id'),  # ROOT_ID column
        }
        
        # Fix root ID padding for root_id if present
        if name_data['root_id']:
            if re.match(r'^R\d{1,3}$', name_data['root_id']):
                match = re.match(r'^R(\d+)$', name_data['root_id'])
                if match:
                    num = int(match.group(1))
                    name_data['root_id'] = f'R{num:03d}'
        
        try:
            cursor.execute('''
                INSERT INTO names_of_allah (
                    allah_id, arabic_name, transliteration, meaning,
                    qur_ref, entry_ids, root_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                name_data['name_id'], name_data['arabic_name'], name_data['transliteration'],
                name_data['meaning'], name_data['qur_ref'], name_data['entry_ids'],
                name_data.get('root_id')
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {name_data}")
    
    print(f"    Migrated {count} names of Allah")
    return count

def migrate_a5_cross_refs(conn, cursor, wb):
    """Migrate A5_CROSS_REFS sheet to cross_refs table (with TEXT xref_id)."""
    print("  Migrating A5_CROSS_REFS...")
    
    if "A5_CROSS_REFS" not in wb.sheetnames:
        print("    Sheet A5_CROSS_REFS not found")
        return 0
    
    ws = wb["A5_CROSS_REFS"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A5_CROSS_REFS")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map columns according to Excel structure:
        # XREF_ID → xref_id (string)
        # FROM_ID → from_entry_id (integer)
        # TO_ID → to_entry_id (integer)
        # LINK_TYPE → link_type
        # DESCRIPTION → description
        # LAYER_REF → layer_ref
        xref_id = row_dict.get('xref_id')
        from_id = row_dict.get('from_id')
        to_id = row_dict.get('to_id')
        link_type = row_dict.get('link_type')
        description = row_dict.get('description')
        layer_ref = row_dict.get('layer_ref')
        
        # Skip if missing critical data
        if not xref_id or from_id is None or to_id is None:
            continue
        
        # Convert IDs to integers
        try:
            from_id = int(from_id)
            to_id = int(to_id)
        except (ValueError, TypeError):
            print(f"    Row {i}: Invalid ID format - from_id={from_id}, to_id={to_id}")
            continue
        
        # Check if entry IDs exist
        cursor.execute('SELECT 1 FROM entries WHERE entry_id = ?', (from_id,))
        if not cursor.fetchone():
            print(f"    Row {i}: from_entry_id {from_id} not found in entries")
            continue
        
        cursor.execute('SELECT 1 FROM entries WHERE entry_id = ?', (to_id,))
        if not cursor.fetchone():
            print(f"    Row {i}: to_entry_id {to_id} not found in entries")
            continue
        
        # Check if cross-ref already exists
        cursor.execute('''
            SELECT 1 FROM cross_refs 
            WHERE from_entry_id = ? AND to_entry_id = ? AND link_type = ?
        ''', (from_id, to_id, link_type))
        
        if cursor.fetchone():
            # Already exists, skip
            continue
        
        try:
            cursor.execute('''
                INSERT INTO cross_refs (
                    xref_id, from_entry_id, to_entry_id, link_type, description, layer_ref
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                xref_id, from_id, to_id, link_type, description, layer_ref
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: xref_id={xref_id}, from_id={from_id}, to_id={to_id}, link_type={link_type}")
    
    print(f"    Migrated {count} cross-references")
    return count

def migrate_a3_quran_refs(conn, cursor, wb):
    """Migrate A3_QURAN_REFS sheet to quran_refs table."""
    print("  Migrating A3_QURAN_REFS...")
    
    if "A3_QURAN_REFS" not in wb.sheetnames:
        print("    Sheet A3_QURAN_REFS not found")
        return 0
    
    ws = wb["A3_QURAN_REFS"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A3_QURAN_REFS")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        quran_data = {
            'ref_id': row_dict.get('ref_id'),
            'surah': row_dict.get('surah'),
            'ayah': row_dict.get('ayah'),
            'arabic_text': row_dict.get('arabic_text'),
            'relevance': row_dict.get('relevance'),
            'entry_ids': row_dict.get('entry_ids'),
            'network_id': row_dict.get('network_id'),
            'layer_ref': row_dict.get('layer_ref'),
            'qv_id': row_dict.get('qv_id'),
        }
        
        try:
            cursor.execute('''
                INSERT INTO quran_refs (
                    ref_id, surah, ayah, arabic_text, relevance,
                    entry_ids, network_id, layer_ref, qv_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                quran_data['ref_id'], quran_data['surah'], quran_data['ayah'],
                quran_data['arabic_text'], quran_data['relevance'],
                quran_data['entry_ids'], quran_data['network_id'],
                quran_data['layer_ref'], quran_data['qv_id']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {quran_data}")
    
    print(f"    Migrated {count} Qur'an references")
    return count

# ============================================================================
# OTHER MIGRATION FUNCTIONS (simplified - include only essential ones)
# ============================================================================

def migrate_child_schema(conn, cursor, wb):
    """Migrate CHILD_SCHEMA sheet to child_entries table."""
    print("  Migrating CHILD_SCHEMA...")
    
    if "CHILD_SCHEMA" not in wb.sheetnames:
        print("    Sheet CHILD_SCHEMA not found")
        return 0
    
    ws = wb["CHILD_SCHEMA"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 4:
        print("    No data in CHILD_SCHEMA")
        return 0
    
    # Row 2 is headers (0: title, 1: description, 2: headers)
    headers = [clean_column_name(cell) for cell in rows[2]]
    
    count = 0
    for i, row in enumerate(rows[3:], start=3):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        child_data = {
            'child_id': row_dict.get('entry_id'),
            'shell_name': row_dict.get('shell_name'),
            'shell_language': row_dict.get('shell_language'),
            'orig_class': row_dict.get('orig_class'),
            'orig_root': row_dict.get('orig_root'),
            'orig_lemma': row_dict.get('orig_lemma'),
            'orig_meaning': row_dict.get('orig_meaning'),
            'operation_role': row_dict.get('operation_role'),
            'shell_meaning': row_dict.get('shell_meaning'),
            'inversion_direction': row_dict.get('inversion_direction'),
            'phonetic_chain': row_dict.get('phonetic_chain'),
            'qur_anchors': row_dict.get('qur_anchors'),
            'dp_codes': row_dict.get('dp_codes'),
            'nt_code': row_dict.get('nt_code'),
            'pattern': row_dict.get('pattern'),
            'parent_op': row_dict.get('parent_op'),
            'gate_status': row_dict.get('gate_status'),
            'notes': row_dict.get('notes'),
        }
        
        # Clean up pattern field
        pattern = child_data['pattern']
        if pattern and '+' in pattern:
            pattern = pattern.split('+')[0]
        child_data['pattern'] = pattern
        
        try:
            cursor.execute('''
                INSERT INTO child_entries (
                    child_id, shell_name, shell_language, orig_class, orig_root,
                    orig_lemma, orig_meaning, operation_role, shell_meaning,
                    inversion_direction, phonetic_chain, qur_anchors, dp_codes,
                    nt_code, pattern, parent_op, gate_status, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                child_data['child_id'], child_data['shell_name'], child_data['shell_language'],
                child_data['orig_class'], child_data['orig_root'], child_data['orig_lemma'],
                child_data['orig_meaning'], child_data['operation_role'], child_data['shell_meaning'],
                child_data['inversion_direction'], child_data['phonetic_chain'], child_data['qur_anchors'],
                child_data['dp_codes'], child_data['nt_code'], child_data['pattern'],
                child_data['parent_op'], child_data['gate_status'], child_data['notes']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {child_data}")
    
    print(f"    Migrated {count} child entries")
    return count

def migrate_a4_derivatives(conn, cursor, wb):
    """Migrate A4_DERIVATIVES sheet to derivatives table."""
    print("  Migrating A4_DERIVATIVES...")
    
    if "A4_DERIVATIVES" not in wb.sheetnames:
        print("    Sheet A4_DERIVATIVES not found")
        return 0
    
    ws = wb["A4_DERIVATIVES"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A4_DERIVATIVES")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        entry_id = row_dict.get('entry_id')
        
        # If entry_id is not available, try to get it from en_term via lookup
        if not entry_id and row_dict.get('en_term'):
            cursor.execute('SELECT entry_id FROM entries WHERE en_term = ? LIMIT 1', (row_dict.get('en_term'),))
            result = cursor.fetchone()
            if result:
                entry_id = result[0]
        
        derivative_term = row_dict.get('derivative')
        link_type = row_dict.get('link_type')
        
        # Skip if missing critical data
        if not entry_id or not derivative_term:
            continue
        
        try:
            cursor.execute('''
                INSERT INTO derivatives (
                    entry_id, derivative_term, language, link_type
                ) VALUES (?, ?, ?, ?)
            ''', (
                entry_id, derivative_term, 'en', link_type
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: entry_id={entry_id}, derivative_term={derivative_term}, link_type={link_type}")
    
    print(f"    Migrated {count} derivatives")
    return count

def extract_and_insert_roots(conn, cursor):
    """Extract unique roots from all entries and insert into roots table."""
    print("  Extracting roots from entries...")
    
    # First, collect all root_id values already present in entries
    cursor.execute('''
        SELECT DISTINCT root_id 
        FROM entries 
        WHERE root_id IS NOT NULL AND root_id != ''
    ''')
    existing_root_ids = {row[0] for row in cursor.fetchall()}
    
    # Fix root ID padding: convert R01 to R001 format
    count_fixed = 0
    for root_id in list(existing_root_ids):
        if root_id and re.match(r'^R\d{1,3}$', root_id):
            # Extract numeric part and pad to 3 digits
            match = re.match(r'^R(\d+)$', root_id)
            if match:
                num = int(match.group(1))
                padded_id = f'R{num:03d}'
                if padded_id != root_id:
                    # Update entries table to use padded ID
                    cursor.execute('''
                        UPDATE entries 
                        SET root_id = ?
                        WHERE root_id = ?
                    ''', (padded_id, root_id))
                    count_fixed += 1
                    print(f"    Fixed root ID padding: {root_id} → {padded_id}")
    
    # Re-fetch existing root_ids after fixing
    cursor.execute('''
        SELECT DISTINCT root_id 
        FROM entries 
        WHERE root_id IS NOT NULL AND root_id != ''
    ''')
    existing_root_ids = {row[0] for row in cursor.fetchall()}
    
    # For each existing root_id, ensure it exists in roots table
    count_existing = 0
    for root_id in existing_root_ids:
        cursor.execute('SELECT 1 FROM roots WHERE root_id = ?', (root_id,))
        if not cursor.fetchone():
            # Root doesn't exist, create a minimal entry
            # Extract numeric part for root_bare
            root_bare = root_id
            if root_id.startswith('R'):
                root_bare = root_id[1:]  # Remove 'R' prefix
            
            cursor.execute('''
                INSERT INTO roots (root_id, root_letters, root_bare, notes)
                VALUES (?, ?, ?, ?)
            ''', (root_id, f"Root {root_id}", root_bare, "Created during migration for existing entries"))
            count_existing += 1
    
    # Get all unique root_letters from entries that don't have root_id yet
    cursor.execute('''
        SELECT DISTINCT root_letters 
        FROM entries 
        WHERE root_letters IS NOT NULL AND root_letters != ''
        AND (root_id IS NULL OR root_id = '')
    ''')
    root_rows = cursor.fetchall()
    
    count_new = 0
    for root_letters, in root_rows:
        if not root_letters:
            continue
        
        # Generate root_id (R{count+1:03d})
        # But first check if we already have a root with these letters
        cursor.execute('SELECT root_id FROM roots WHERE root_letters = ?', (root_letters,))
        existing = cursor.fetchone()
        
        if existing:
            root_id = existing[0]
        else:
            # Count existing roots to generate new ID
            cursor.execute('SELECT COUNT(*) FROM roots')
            root_count = cursor.fetchone()[0]
            root_id = f"R{root_count + 1:03d}"
            
            # Create bare root (without hyphens)
            root_bare = re.sub(r'[\-\s]', '', root_letters)
            
            # Count entries with this root
            cursor.execute('SELECT COUNT(*) FROM entries WHERE root_letters = ?', (root_letters,))
            entry_count = cursor.fetchone()[0]
            
            # Try to extract root type from pattern
            root_type = 'TRILITERAL' if len(root_bare) == 3 else 'QUADRILITERAL' if len(root_bare) == 4 else 'OTHER'
            
            # Insert into roots table
            try:
                cursor.execute('''
                    INSERT INTO roots (
                        root_id, root_letters, root_bare, root_type, notes
                    ) VALUES (?, ?, ?, ?, ?)
                ''', (
                    root_id, root_letters, root_bare, root_type,
                    f"Extracted from {entry_count} entries during migration"
                ))
                count_new += 1
            except Exception as e:
                print(f"    Error inserting root {root_letters}: {e}")
                continue
        
        # Update entries with the root_id
        cursor.execute('''
            UPDATE entries 
            SET root_id = ? 
            WHERE root_letters = ? AND (root_id IS NULL OR root_id = '')
        ''', (root_id, root_letters))
    
    print(f"    Fixed {count_fixed} root IDs, created {count_existing} roots for existing IDs, {count_new} new roots from root_letters")
    return count_existing + count_new

# ============================================================================
# MAIN MIGRATION FUNCTION
# ============================================================================

def main():
    print("\n" + "═" * 70)
    print("  USLaP Migration with All Fixes")
    print("  بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ")
    print("═" * 70)
    
    # Check if files exist
    if not os.path.exists(EXCEL_PATH):
        print(f"\n❌ ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)
    
    if not os.path.exists(SCHEMA_PATH):
        print(f"\n❌ ERROR: Schema file not found: {SCHEMA_PATH}")
        sys.exit(1)
    
    # Backup existing database if it exists
    if os.path.exists(DB_PATH):
        backup_path = backup_database(DB_PATH)
        if backup_path:
            print(f"\n✓ Backed up existing database to: {backup_path}")
        # Remove the existing database file to start fresh
        os.remove(DB_PATH)
        print(f"  Removed existing database file to start fresh")
    
    # Load Excel workbook
    print(f"\n📖 Loading Excel file: {EXCEL_PATH}")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        print(f"  Found {len(wb.sheetnames)} sheets")
    except Exception as e:
        print(f"\n❌ ERROR: Failed to load Excel file: {e}")
        sys.exit(1)
    
    # Create database and execute fixed schema
    print(f"\n🗄️  Creating database: {DB_PATH}")
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Enable foreign keys
        cursor.execute("PRAGMA foreign_keys = ON")
        
        # Create fixed schema (with TEXT xref_id in cross_refs)
        if not create_fixed_schema(conn, cursor):
            print("❌ Failed to create fixed schema")
            sys.exit(1)
        
        # Register extract_consonants UDF (CRITICAL: must be done before any inserts)
        print("  Registering extract_consonants() UDF...")
        conn.create_function("extract_consonants", 1, extract_consonants)
        
        # Disable foreign keys during data migration to avoid constraint violations
        cursor.execute("PRAGMA foreign_keys = OFF")
        
        # Begin transaction
        conn.execute("BEGIN TRANSACTION")
        
        # Migrate each sheet
        print("\n📊 Migrating data...")
        
        total_counts = {}
        
        # Order matters: create entries first, then related tables
        total_counts['A1_ENTRIES'] = migrate_a1_entries(conn, cursor, wb)
        total_counts['A1_ЗАПИСИ'] = migrate_a1_zapisi(conn, cursor, wb)
        total_counts['PERSIAN_A1_MADĀKHIL'] = migrate_persian_a1_madakhil(conn, cursor, wb)
        total_counts['BITIG_A1_ENTRIES'] = migrate_bitig_a1_entries(conn, cursor, wb)
        total_counts['A2_NAMES_OF_ALLAH'] = migrate_a2_names_of_allah(conn, cursor, wb)
        
        # Extract and insert roots (must be done before other tables that reference roots)
        total_counts['roots'] = extract_and_insert_roots(conn, cursor)
        
        # Migrate remaining tables
        total_counts['CHILD_SCHEMA'] = migrate_child_schema(conn, cursor, wb)
        total_counts['A4_DERIVATIVES'] = migrate_a4_derivatives(conn, cursor, wb)
        total_counts['A5_CROSS_REFS'] = migrate_a5_cross_refs(conn, cursor, wb)
        total_counts['A3_QURAN_REFS'] = migrate_a3_quran_refs(conn, cursor, wb)
        
        # Commit transaction
        conn.commit()
        
        # Re-enable foreign keys and check constraints
        cursor.execute("PRAGMA foreign_keys = ON")
        
        # Generate statistics
        print("\n📈 Migration Statistics:")
        print("─" * 40)
        
        cursor.execute("SELECT COUNT(*) FROM entries")
        entries_count = cursor.fetchone()[0]
        print(f"  Total entries: {entries_count}")
        
        cursor.execute("SELECT COUNT(*) FROM cross_refs")
        crossrefs_count = cursor.fetchone()[0]
        print(f"  Total cross-references: {crossrefs_count}")
        
        cursor.execute("SELECT COUNT(*) FROM entries WHERE ru_term IS NOT NULL AND ru_term != ''")
        ru_count = cursor.fetchone()[0]
        print(f"  Entries with ru_term: {ru_count}")
        
        cursor.execute("SELECT COUNT(*) FROM entries WHERE fa_term IS NOT NULL AND fa_term != ''")
        fa_count = cursor.fetchone()[0]
        print(f"  Entries with fa_term: {fa_count}")
        
        cursor.execute("SELECT COUNT(*) FROM entries WHERE foundation_refs LIKE 'ORIG2:%'")
        orig2_count = cursor.fetchone()[0]
        print(f"  ORIG2 entries: {orig2_count}")
        
        cursor.execute("SELECT COUNT(*) FROM word_fingerprints")
        fingerprints_count = cursor.fetchone()[0]
        print(f"  Word fingerprints: {fingerprints_count}")
        
        # Verify foreign keys
        print("\n🔍 Verifying foreign key constraints...")
        cursor.execute("PRAGMA foreign_key_check")
        fk_errors = cursor.fetchall()
        
        if fk_errors:
            print("❌ Foreign key errors found:")
            for error in fk_errors:
                print(f"  Table: {error[0]}, Row: {error[1]}, Referenced: {error[2]}, FK Index: {error[3]}")
        else:
            print("✓ All foreign key constraints satisfied")
        
        # Close workbook and connection
        wb.close()
        conn.close()
        
        print("\n" + "═" * 70)
        print("✅ MIGRATION WITH ALL FIXES COMPLETED SUCCESSFULLY")
        print("═" * 70)
        
        # Print next steps
        print("\n📋 Next steps:")
        print("  1. Test the database with: sqlite3 Code_files/uslap_lattice.db '.tables'")
        print("  2. Verify data integrity with queries")
        print(f"  3. Target counts should be: ~392 entries, ~155 cross-references")
        
        # Show actual counts vs targets
        print(f"\n📊 Actual vs Target:")
        print(f"  Entries: {entries_count} / ~392")
        print(f"  Cross-references: {crossrefs_count} / ~155")
        print(f"  Russian entries: {ru_count} / ~91")
        print(f"  Persian entries: {fa_count} / ~59")
        print(f"  ORIG2 entries: {orig2_count} / ~26")
        
    except Exception as e:
        print(f"\n❌ ERROR during migration: {e}")
        import traceback
        traceback.print_exc()
        
        # Rollback if connection is still open
        try:
            conn.rollback()
            conn.close()
        except:
            pass
        
        print("\n⚠️  Migration failed. Database has been rolled back.")
        sys.exit(1)

if __name__ == "__main__":
    main()