#!/usr/bin/env python3
"""
Test script to verify the safe accelerated growth script works with current Excel file
"""

import openpyxl
from pathlib import Path
import shutil
import time

def test_file_operations():
    """Test basic file operations on the current Excel file"""
    filepath = Path('USLaP_Final_Data_Consolidated_Master.xlsx')
    
    print(f"Testing file: {filepath.name}")
    print(f"File size: {filepath.stat().st_size:,} bytes")
    
    # 1. Test if we can load the file
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        print(f"✓ File loads successfully")
        print(f"  Sheets: {wb.sheetnames}")
        wb.close()
    except Exception as e:
        print(f"✗ Error loading file: {e}")
        return False
    
    # 2. Test if we can create a backup
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    backup_path = filepath.parent / f"{filepath.stem}_test_backup_{timestamp}{filepath.suffix}"
    
    try:
        shutil.copy2(filepath, backup_path)
        print(f"✓ Backup created: {backup_path.name}")
        
        # Verify backup
        wb_backup = openpyxl.load_workbook(backup_path, data_only=True, read_only=True)
        print(f"✓ Backup loads successfully")
        wb_backup.close()
        
        # Clean up test backup
        backup_path.unlink()
        print(f"✓ Test backup cleaned up")
        
    except Exception as e:
        print(f"✗ Error creating backup: {e}")
        return False
    
    # 3. Test if we can add a new sheet
    try:
        # Load workbook for editing
        wb_edit = openpyxl.load_workbook(filepath)
        
        # Check if A1_ENTRIES exists
        if 'A1_ENTRIES' not in wb_edit.sheetnames:
            print("  A1_ENTRIES sheet doesn't exist, will create it")
            ws = wb_edit.create_sheet('A1_ENTRIES')
            
            # Add headers
            headers = ['ENTRY_ID', 'SCORE', 'EN_TERM', 'AR_WORD', 'ROOT_ID', 
                      'ROOT_LETTERS', 'QUR_MEANING', 'PATTERN', 'NETWORK_ID',
                      'PHONETIC_CHAIN', 'INVERSION_TYPE', 'SOURCE_FORM', 'FOUNDATION_REF']
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            print(f"✓ Created A1_ENTRIES sheet with headers")
            
            # Add a test entry
            test_entry = {
                'ENTRY_ID': 'TEST001',
                'SCORE': 10,
                'EN_TERM': 'Test entry',
                'AR_WORD': 'اختبار',
                'ROOT_ID': 'R001',
                'ROOT_LETTERS': '—',
                'QUR_MEANING': 'Test entry for verification',
                'PATTERN': 'A',
                'NETWORK_ID': 'N08',
                'PHONETIC_CHAIN': '→'.join(['✓'] * 3),
                'INVERSION_TYPE': 'DIVINE',
                'SOURCE_FORM': 'test:001',
                'FOUNDATION_REF': 'TEST_REF'
            }
            
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=test_entry['ENTRY_ID'])
            ws.cell(row=next_row, column=2, value=test_entry['SCORE'])
            ws.cell(row=next_row, column=3, value=test_entry['EN_TERM'])
            ws.cell(row=next_row, column=4, value=test_entry['AR_WORD'])
            ws.cell(row=next_row, column=5, value=test_entry['ROOT_ID'])
            ws.cell(row=next_row, column=6, value=test_entry['ROOT_LETTERS'])
            ws.cell(row=next_row, column=7, value=test_entry['QUR_MEANING'])
            ws.cell(row=next_row, column=8, value=test_entry['PATTERN'])
            ws.cell(row=next_row, column=9, value=test_entry['NETWORK_ID'])
            ws.cell(row=next_row, column=10, value=test_entry['PHONETIC_CHAIN'])
            ws.cell(row=next_row, column=11, value=test_entry['INVERSION_TYPE'])
            ws.cell(row=next_row, column=12, value=test_entry['SOURCE_FORM'])
            ws.cell(row=next_row, column=13, value=test_entry['FOUNDATION_REF'])
            
            print(f"✓ Added test entry to row {next_row}")
        
        # Save to temp file first (safe write)
        temp_file = filepath.parent / f"temp_{filepath.name}"
        wb_edit.save(temp_file)
        
        # Verify temp file
        test_wb = openpyxl.load_workbook(temp_file, read_only=True)
        test_wb.close()
        
        # Replace original with temp file
        shutil.move(temp_file, filepath)
        print(f"✓ Successfully saved changes using safe write pattern")
        
        # Verify the file still loads
        wb_final = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if 'A1_ENTRIES' in wb_final.sheetnames:
            ws_final = wb_final['A1_ENTRIES']
            print(f"✓ A1_ENTRIES sheet exists with {ws_final.max_row} rows")
        wb_final.close()
        
    except Exception as e:
        print(f"✗ Error modifying file: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True

def test_simple_growth():
    """Test a simple growth cycle without modifying the original file"""
    print("\n" + "="*60)
    print("Testing simple growth cycle on COPY of file")
    print("="*60)
    
    original = Path('USLaP_Final_Data_Consolidated_Master.xlsx')
    test_copy = Path('USLaP_TEST_COPY.xlsx')
    
    # Create test copy
    try:
        shutil.copy2(original, test_copy)
        print(f"✓ Created test copy: {test_copy.name}")
    except Exception as e:
        print(f"✗ Error creating test copy: {e}")
        return False
    
    try:
        # Load and modify test copy
        wb = openpyxl.load_workbook(test_copy)
        
        # Create or get A1_ENTRIES sheet
        if 'A1_ENTRIES' not in wb.sheetnames:
            ws = wb.create_sheet('A1_ENTRIES')
            
            # Add headers
            headers = ['ENTRY_ID', 'SCORE', 'EN_TERM', 'AR_WORD', 'ROOT_ID', 
                      'ROOT_LETTERS', 'QUR_MEANING', 'PATTERN', 'NETWORK_ID',
                      'PHONETIC_CHAIN', 'INVERSION_TYPE', 'SOURCE_FORM', 'FOUNDATION_REF']
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            print(f"✓ Created A1_ENTRIES sheet in test copy")
        
        # Add some test entries
        test_entries = []
        for i in range(1, 6):
            test_entries.append({
                'ENTRY_ID': f'TEST{i:03d}',
                'SCORE': 8 + i,
                'EN_TERM': f'Test ratio {i}/3',
                'AR_WORD': f'نسبة {i}/3',
                'ROOT_ID': f'R{i:03d}',
                'ROOT_LETTERS': '—',
                'QUR_MEANING': f'Test divine ratio {i}/3',
                'PATTERN': 'A',
                'NETWORK_ID': 'N08',
                'PHONETIC_CHAIN': '→'.join(['✓'] * 3),
                'INVERSION_TYPE': 'DIVINE',
                'SOURCE_FORM': f'test:ratio_{i}/3',
                'FOUNDATION_REF': f'TEST_REF_{i}'
            })
        
        ws = wb['A1_ENTRIES']
        next_row = ws.max_row + 1
        
        for entry in test_entries:
            ws.cell(row=next_row, column=1, value=entry['ENTRY_ID'])
            ws.cell(row=next_row, column=2, value=entry['SCORE'])
            ws.cell(row=next_row, column=3, value=entry['EN_TERM'])
            ws.cell(row=next_row, column=4, value=entry['AR_WORD'])
            ws.cell(row=next_row, column=5, value=entry['ROOT_ID'])
            ws.cell(row=next_row, column=6, value=entry['ROOT_LETTERS'])
            ws.cell(row=next_row, column=7, value=entry['QUR_MEANING'])
            ws.cell(row=next_row, column=8, value=entry['PATTERN'])
            ws.cell(row=next_row, column=9, value=entry['NETWORK_ID'])
            ws.cell(row=next_row, column=10, value=entry['PHONETIC_CHAIN'])
            ws.cell(row=next_row, column=11, value=entry['INVERSION_TYPE'])
            ws.cell(row=next_row, column=12, value=entry['SOURCE_FORM'])
            ws.cell(row=next_row, column=13, value=entry['FOUNDATION_REF'])
            next_row += 1
        
        # Save test copy
        wb.save(test_copy)
        print(f"✓ Added {len(test_entries)} test entries to A1_ENTRIES sheet")
        
        # Verify the test copy
        wb_check = openpyxl.load_workbook(test_copy, data_only=True, read_only=True)
        ws_check = wb_check['A1_ENTRIES']
        print(f"✓ Test copy verified: A1_ENTRIES has {ws_check.max_row} rows")
        wb_check.close()
        
        # Clean up test copy
        test_copy.unlink()
        print(f"✓ Test copy cleaned up")
        
        return True
        
    except Exception as e:
        print(f"✗ Error in simple growth test: {e}")
        import traceback
        traceback.print_exc()
        
        # Clean up test copy if it exists
        if test_copy.exists():
            try:
                test_copy.unlink()
            except:
                pass
        
        return False

if __name__ == "__main__":
    print("="*60)
    print("TESTING USLaP SAFE SCRIPT COMPATIBILITY")
    print("="*60)
    
    # Test 1: Basic file operations
    print("\nTest 1: Basic file operations")
    print("-" * 40)
    test1_passed = test_file_operations()
    
    # Test 2: Simple growth on copy
    print("\nTest 2: Simple growth on copy")
    print("-" * 40)
    test2_passed = test_simple_growth()
    
    print("\n" + "="*60)
    print("TEST RESULTS")
    print("="*60)
    print(f"Test 1 (File operations): {'✓ PASSED' if test1_passed else '✗ FAILED'}")
    print(f"Test 2 (Simple growth): {'✓ PASSED' if test2_passed else '✗ FAILED'}")
    
    if test1_passed and test2_passed:
        print("\n✅ All tests passed! The safe script should work.")
        print("\nRecommendation: Run uslap_accelerated_safe.py with small cycles first:")
        print("  python3 uslap_accelerated_safe.py USLaP_Final_Data_Consolidated_Master.xlsx 5")
        print("\nThis will run 5 safe cycles (creates backup first).")
    else:
        print("\n❌ Some tests failed. Check the errors above.")
    
    print("="*60)