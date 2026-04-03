#!/usr/bin/env python3
import openpyxl
import sys

wb = openpyxl.load_workbook("USLaP_Final_Data_Consolidated_Master_v3.xlsx", read_only=True)
print("Sheets in v3:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")

# Also check row counts for new sheets
for sheet in ["PROTOCOL_CORRECTIONS", "SCHOLAR_WARNINGS"]:
    if sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        print(f"\n{sheet}: {len(rows)} rows")
        if rows:
            print("First few rows:")
            for i, row in enumerate(rows[:5]):
                print(f"  Row {i}: {row}")
wb.close()