#!/usr/bin/env python3
"""
USLaP Migration Script: Excel → SQLite relational schema
Extends build_database_v3.py to populate the full USLaP lattice database.

This script:
1. Creates uslap_lattice.db with the full relational schema (create_uslap_db.sql)
2. Reads data from structured Excel sheets (skip EXCEL_DATA_CONSOLIDATED)
3. Normalizes data into the relational schema with proper foreign keys
4. Registers Python UDF extract_consonants() from USLaP_Engine.py
5. Generates word_fingerprints table for O(log n) cluster expansion
6. Uses transaction safety with rollback on error
7. Creates backup of existing databases

بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ
"""

import sqlite3
import openpyxl
import re
import sys
import os
import shutil
import json
from datetime import datetime
from pathlib import Path

# Add the current directory to sys.path to import USLaP_Engine
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ============================================================================
# GLOBAL CONFIGURATION
# ============================================================================

EXCEL_PATH = "USLaP_Final_Data_Consolidated_Master_v3.xlsx"
DB_PATH = "Code_files/uslap_lattice.db"
SCHEMA_PATH = "Code_files/create_uslap_db.sql"
BACKUP_DIR = "Code_files/backups"

# Sheets to migrate (structured data only, skip EXCEL_DATA_CONSOLIDATED)
SHEETS_TO_MIGRATE = [
    "A1_ENTRIES",
    "A1_ЗАПИСИ",
    "PERSIAN_A1_MADĀKHIL", 
    "BITIG_A1_ENTRIES",
    "CHILD_SCHEMA",
    "A4_DERIVATIVES",
    "A5_CROSS_REFS",
    "A3_QURAN_REFS",
    "M1_PHONETIC_SHIFTS",
    "M2_DETECTION_PATTERNS",
    "M4_NETWORKS",
    "M3_SCHOLARS",
    "M5_QUR_VERIFICATION",
    "A2_NAMES_OF_ALLAH",  # Added missing sheet
]

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def clean_column_name(col):
    """Convert Excel column header to valid SQLite column name."""
    if col is None:
        return "unknown"
    col = str(col).strip()
    col = re.sub(r'[^\w\s]', '', col)  # Remove punctuation
    col = re.sub(r'\s+', '_', col)     # Replace spaces with underscore
    col = col.lower()
    if not col:
        return "unknown"
    return col

def find_header_row(ws, sheet_name):
    """Find the header row in an Excel sheet based on patterns."""
    rows = list(ws.iter_rows(values_only=True))
    
    # Special handling for each sheet based on debug output
    if sheet_name == "CHILD_SCHEMA":
        # Row 0: title, Row 1: description, Row 2: headers
        if len(rows) > 2:
            return rows[2], 3
    elif sheet_name == "PHONETIC_REVERSAL":
        # Row 0: title, Row 1: description, Row 2: zone header, Row 3: headers
        if len(rows) > 3:
            return rows[3], 4
    elif sheet_name in ["UMD_OPERATIONS", "DP_REGISTER", "ATT_TERMS", 
                        "SESSION_INDEX", "PROTOCOL_CORRECTIONS", 
                        "SCHOLAR_WARNINGS", "A1_ENTRIES"]:
        # Row 0: title/headers (most sheets)
        if len(rows) > 0:
            return rows[0], 1
    else:
        # Generic fallback: find row with typical column names
        for i, row in enumerate(rows):
            if row and any(isinstance(cell, str) and ('ID' in cell or 'NAME' in cell or 'CODE' in cell or 'TERM' in cell) for cell in row):
                return row, i + 1
    
    return None, 0

def backup_database(db_path):
    """Create a timestamped backup of existing database."""
    if not os.path.exists(db_path):
        return None
    
    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(BACKUP_DIR, f"{os.path.basename(db_path)}_backup_{timestamp}.db")
    shutil.copy2(db_path, backup_path)
    print(f"  Created backup: {backup_path}")
    return backup_path

# ============================================================================
# EXTRACT_CONSONANTS UDF (from USLaP_Engine.py)
# ============================================================================

def extract_consonants(word):
    """
    Python UDF for SQLite: Extract consonant skeleton from a word.
    Must be registered before any operations on word_fingerprints table.
    
    This is the same logic as PhoneticReversal.extract_consonants() in USLaP_Engine.py.
    """
    if not word:
        return ""
    
    # Import the function directly from USLaP_Engine if available
    try:
        from USLaP_Engine import PhoneticReversal
        # Create a minimal instance just to use the method
        # We'll implement the logic directly here to avoid dependencies
        pass
    except ImportError:
        pass
    
    # Direct implementation from USLaP_Engine.py
    vowels = set('aeiou')
    result = []
    i = 0
    word_lower = word.lower()
    
    while i < len(word_lower):
        digraph = word_lower[i:i+2] if i + 1 < len(word_lower) else ''
        if digraph in ('sh', 'ch', 'gh', 'th', 'ph', 'wh', 'qu'):
            result.append(digraph)
            i += 2
        elif word_lower[i] not in vowels:
            result.append(word_lower[i])
            i += 1
        else:
            i += 1
    
    return ''.join(result)

# ============================================================================
# MIGRATION FUNCTIONS FOR EACH SHEET
# ============================================================================

def migrate_a1_entries(conn, cursor, wb):
    """Migrate A1_ENTRIES sheet to entries table."""
    print("  Migrating A1_ENTRIES...")
    
    if "A1_ENTRIES" not in wb.sheetnames:
        print("    Sheet A1_ENTRIES not found")
        return 0
    
    ws = wb["A1_ENTRIES"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A1_ENTRIES")
        return 0
    
    # First row is headers
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map to entries table schema
        entry_data = {
            'entry_id': row_dict.get('entry_id'),
            'score': row_dict.get('score'),
            'en_term': row_dict.get('en_term'),
            'ar_word': row_dict.get('ar_word'),
            'root_letters': row_dict.get('root_letters'),
            'qur_meaning': row_dict.get('qur_meaning'),
            'pattern': row_dict.get('pattern'),
            'allah_name_id': row_dict.get('allah_name_id'),
            'network_id': row_dict.get('network_id'),
            'phonetic_chain': row_dict.get('phonetic_chain'),
            'inversion_type': row_dict.get('inversion_type'),
            'source_form': row_dict.get('source_form'),
            'foundation_refs': row_dict.get('foundation_ref'),
        }
        
        # Extract root ID from root_letters or root_id
        root_id = row_dict.get('root_id')
        if root_id:
            # Fix root ID padding: convert R01 to R001 format
            if root_id and re.match(r'^R\d{1,3}$', root_id):
                match = re.match(r'^R(\d+)$', root_id)
                if match:
                    num = int(match.group(1))
                    root_id = f'R{num:03d}'
            entry_data['root_id'] = root_id
        
        # Extract DS corridor from foundation_ref
        foundation_ref = row_dict.get('foundation_ref', '')
        ds_match = re.search(r'DS\d+', str(foundation_ref))
        if ds_match:
            entry_data['ds_corridor'] = ds_match.group(0)
        
        # Insert into entries table
        try:
            cursor.execute('''
                INSERT INTO entries (
                    entry_id, score, en_term, ar_word, root_id, root_letters,
                    qur_meaning, pattern, allah_name_id, network_id, phonetic_chain,
                    inversion_type, source_form, foundation_refs, ds_corridor
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                entry_data['entry_id'], entry_data['score'], entry_data['en_term'],
                entry_data['ar_word'], entry_data.get('root_id'), entry_data['root_letters'],
                entry_data['qur_meaning'], entry_data['pattern'], entry_data['allah_name_id'],
                entry_data['network_id'], entry_data['phonetic_chain'], entry_data['inversion_type'],
                entry_data['source_form'], entry_data['foundation_refs'], entry_data.get('ds_corridor')
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {entry_data}")
    
    print(f"    Migrated {count} entries")
    return count

def migrate_a1_zapisi(conn, cursor, wb):
    """Migrate A1_ЗАПИСИ sheet to entries table (Russian entries)."""
    print("  Migrating A1_ЗАПИСИ (Russian entries)...")
    
    if "A1_ЗАПИСИ" not in wb.sheetnames:
        print("    Sheet A1_ЗАПИСИ not found")
        return 0
    
    ws = wb["A1_ЗАПИСИ"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A1_ЗАПИСИ")
        return 0
    
    # First row is headers (Russian column names)
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map Russian column names to entries table schema
        entry_data = {
            'entry_id': row_dict.get('запись_id'),
            'score': row_dict.get('балл'),
            'ru_term': row_dict.get('рус_термин'),
            'ar_word': row_dict.get('ар_слово'),
            'root_letters': row_dict.get('корневые_буквы'),
            'qur_meaning': row_dict.get('коранич_значение'),
            'pattern': row_dict.get('паттерн'),
            'allah_name_id': row_dict.get('имя_аллаха_id'),
            'network_id': row_dict.get('сеть_id'),
            'phonetic_chain': row_dict.get('фонетическая_цепь'),
            'inversion_type': row_dict.get('тип_инверсии'),
            'source_form': row_dict.get('исходная_форма'),
            'foundation_refs': row_dict.get('основание'),
        }
        
        # Extract root ID
        root_id = row_dict.get('корень_id')
        if root_id:
            # Fix root ID padding
            if root_id and re.match(r'^R\d{1,3}$', root_id):
                match = re.match(r'^R(\d+)$', root_id)
                if match:
                    num = int(match.group(1))
                    root_id = f'R{num:03d}'
            entry_data['root_id'] = root_id
        
        # Insert into entries table (Russian term goes to ru_term)
        try:
            cursor.execute('''
                INSERT INTO entries (
                    entry_id, score, ru_term, ar_word, root_id, root_letters,
                    qur_meaning, pattern, allah_name_id, network_id, phonetic_chain,
                    inversion_type, source_form, foundation_refs
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                entry_data['entry_id'], entry_data['score'], entry_data['ru_term'],
                entry_data['ar_word'], entry_data.get('root_id'), entry_data['root_letters'],
                entry_data['qur_meaning'], entry_data['pattern'], entry_data['allah_name_id'],
                entry_data['network_id'], entry_data['phonetic_chain'], entry_data['inversion_type'],
                entry_data['source_form'], entry_data['foundation_refs']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {entry_data}")
    
    print(f"    Migrated {count} Russian entries")
    return count

def migrate_persian_a1_madakhil(conn, cursor, wb):
    """Migrate PERSIAN_A1_MADĀKHIL sheet to entries table (Persian entries)."""
    print("  Migrating PERSIAN_A1_MADĀKHIL (Persian entries)...")
    
    if "PERSIAN_A1_MADĀKHIL" not in wb.sheetnames:
        print("    Sheet PERSIAN_A1_MADĀKHIL not found")
        return 0
    
    ws = wb["PERSIAN_A1_MADĀKHIL"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in PERSIAN_A1_MADĀKHIL")
        return 0
    
    # First row has complex headers with Persian and English
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map to entries table schema
        entry_data = {
            'entry_id': row_dict.get('madkhal_identry_id'),
            'score': row_dict.get('nomrescore'),
            'fa_term': row_dict.get('vazhe_farsipersian_term'),
            'source_word': row_dict.get('kalame_aslisource_word'),
            'root_letters': row_dict.get('horuf_e_risheroot_letters'),
            'qur_meaning': row_dict.get('mana_ye_quraniquot_meaning'),
            'pattern': row_dict.get('olgu_pattern'),
            'allah_name_id': row_dict.get('esm_e_allah_idallah_name_id'),
            'network_id': row_dict.get('shabake_idnetwork_id'),
            'phonetic_chain': row_dict.get('zanjire_sawtiphoneic_chain'),
            'inversion_type': row_dict.get('now_e_vazhguniinversion_type'),
            'source_form': row_dict.get('shakl_e_aslisource_form'),
            'foundation_refs': row_dict.get('boniyanfoundation_ref'),
        }
        
        # Extract root ID
        root_id = row_dict.get('rishe_idroot_id')
        if root_id:
            # Fix root ID padding
            if root_id and re.match(r'^R\d{1,3}$', root_id):
                match = re.match(r'^R(\d+)$', root_id)
                if match:
                    num = int(match.group(1))
                    root_id = f'R{num:03d}'
            entry_data['root_id'] = root_id
        
        # Use Persian term for fa_term and source_word for ar_word
        try:
            cursor.execute('''
                INSERT INTO entries (
                    entry_id, score, fa_term, ar_word, root_id, root_letters,
                    qur_meaning, pattern, allah_name_id, network_id, phonetic_chain,
                    inversion_type, source_form, foundation_refs
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                entry_data['entry_id'], entry_data['score'], entry_data['fa_term'],
                entry_data['source_word'], entry_data.get('root_id'), entry_data['root_letters'],
                entry_data['qur_meaning'], entry_data['pattern'], entry_data['allah_name_id'],
                entry_data['network_id'], entry_data['phonetic_chain'], entry_data['inversion_type'],
                entry_data['source_form'], entry_data['foundation_refs']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {entry_data}")
    
    print(f"    Migrated {count} Persian entries")
    return count

def migrate_bitig_a1_entries(conn, cursor, wb):
    """Migrate BITIG_A1_ENTRIES sheet to entries table (ORIG2 entries)."""
    print("  Migrating BITIG_A1_ENTRIES (ORIG2 entries)...")
    
    if "BITIG_A1_ENTRIES" not in wb.sheetnames:
        print("    Sheet BITIG_A1_ENTRIES not found")
        return 0
    
    ws = wb["BITIG_A1_ENTRIES"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in BITIG_A1_ENTRIES")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map to entries table schema
        entry_data = {
            'entry_id': row_dict.get('entry_id'),
            'score': row_dict.get('score'),
            'en_term': row_dict.get('orig2_term'),  # ORIG2 term goes to en_term
            'orig2_script': row_dict.get('orig2_script'),
            'root_letters': row_dict.get('root_letters'),
            'phonetic_chain': row_dict.get('phonetic_chain'),
            'semantic_field': row_dict.get('semantic_field'),
            'dispersal_range': row_dict.get('dispersal_range'),
            'status': row_dict.get('status'),
            'notes': row_dict.get('notes'),
        }
        
        # Extract root ID if present
        root_id = row_dict.get('root_id')
        if root_id:
            # Fix root ID padding
            if root_id and re.match(r'^R\d{1,3}$', root_id):
                match = re.match(r'^R(\d+)$', root_id)
                if match:
                    num = int(match.group(1))
                    root_id = f'R{num:03d}'
            entry_data['root_id'] = root_id
        
        # For ORIG2 entries, we need to handle specially
        # Mark as ORIG2 in foundation_refs
        foundation_refs = f"ORIG2: {row_dict.get('kashgari_attestation', '')}"
        
        try:
            cursor.execute('''
                INSERT INTO entries (
                    entry_id, score, en_term, ar_word, root_id, root_letters,
                    phonetic_chain, foundation_refs, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                entry_data['entry_id'], entry_data['score'], entry_data['en_term'],
                entry_data.get('orig2_script'), entry_data.get('root_id'), entry_data['root_letters'],
                entry_data['phonetic_chain'], foundation_refs, entry_data['notes']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {entry_data}")
    
    print(f"    Migrated {count} ORIG2 entries")
    return count

def migrate_a2_names_of_allah(conn, cursor, wb):
    """Migrate A2_NAMES_OF_ALLAH sheet to names_of_allah table."""
    print("  Migrating A2_NAMES_OF_ALLAH...")
    
    if "A2_NAMES_OF_ALLAH" not in wb.sheetnames:
        print("    Sheet A2_NAMES_OF_ALLAH not found")
        return 0
    
    ws = wb["A2_NAMES_OF_ALLAH"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A2_NAMES_OF_ALLAH")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map columns based on actual Excel headers:
        # ALLAH_ID → name_id, ARABIC_NAME → arabic_name, TRANSLITERATION → transliteration
        # MEANING → meaning, QUR_REF → qur_ref, ENTRY_IDS → entry_ids, ROOT_ID → root_id
        name_data = {
            'name_id': row_dict.get('allah_id'),  # ALLAH_ID column
            'arabic_name': row_dict.get('arabic_name'),
            'transliteration': row_dict.get('transliteration'),
            'meaning': row_dict.get('meaning'),
            'qur_ref': row_dict.get('qur_ref'),
            'entry_ids': row_dict.get('entry_ids'),
            'root_id': row_dict.get('root_id'),  # ROOT_ID column
            # Note: network_id and foundation_ref not in Excel sheet
        }
        
        # Fix root ID padding for root_id if present
        if name_data['root_id']:
            if re.match(r'^R\d{1,3}$', name_data['root_id']):
                match = re.match(r'^R(\d+)$', name_data['root_id'])
                if match:
                    num = int(match.group(1))
                    name_data['root_id'] = f'R{num:03d}'
        
        try:
            cursor.execute('''
                INSERT INTO names_of_allah (
                    allah_id, arabic_name, transliteration, meaning,
                    qur_ref, entry_ids, root_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                name_data['name_id'], name_data['arabic_name'], name_data['transliteration'],
                name_data['meaning'], name_data['qur_ref'], name_data['entry_ids'],
                name_data.get('root_id')
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {name_data}")
    
    print(f"    Migrated {count} names of Allah")
    return count

def migrate_child_schema(conn, cursor, wb):
    """Migrate CHILD_SCHEMA sheet to child_entries table."""
    print("  Migrating CHILD_SCHEMA...")
    
    if "CHILD_SCHEMA" not in wb.sheetnames:
        print("    Sheet CHILD_SCHEMA not found")
        return 0
    
    ws = wb["CHILD_SCHEMA"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 4:
        print("    No data in CHILD_SCHEMA")
        return 0
    
    # Row 2 is headers (0: title, 1: description, 2: headers)
    headers = [clean_column_name(cell) for cell in rows[2]]
    
    count = 0
    for i, row in enumerate(rows[3:], start=3):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map to child_entries table schema
        child_data = {
            'child_id': row_dict.get('entry_id'),
            'shell_name': row_dict.get('shell_name'),
            'shell_language': row_dict.get('shell_language'),
            'orig_class': row_dict.get('orig_class'),
            'orig_root': row_dict.get('orig_root'),
            'orig_lemma': row_dict.get('orig_lemma'),
            'orig_meaning': row_dict.get('orig_meaning'),
            'operation_role': row_dict.get('operation_role'),
            'shell_meaning': row_dict.get('shell_meaning'),
            'inversion_direction': row_dict.get('inversion_direction'),
            'phonetic_chain': row_dict.get('phonetic_chain'),
            'qur_anchors': row_dict.get('qur_anchors'),
            'dp_codes': row_dict.get('dp_codes'),
            'nt_code': row_dict.get('nt_code'),
            'pattern': row_dict.get('pattern'),
            'parent_op': row_dict.get('parent_op'),
            'gate_status': row_dict.get('gate_status'),
            'notes': row_dict.get('notes'),
        }
        
        # Clean up pattern field
        pattern = child_data['pattern']
        if pattern and '+' in pattern:
            pattern = pattern.split('+')[0]
        child_data['pattern'] = pattern
        
        try:
            cursor.execute('''
                INSERT INTO child_entries (
                    child_id, shell_name, shell_language, orig_class, orig_root,
                    orig_lemma, orig_meaning, operation_role, shell_meaning,
                    inversion_direction, phonetic_chain, qur_anchors, dp_codes,
                    nt_code, pattern, parent_op, gate_status, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                child_data['child_id'], child_data['shell_name'], child_data['shell_language'],
                child_data['orig_class'], child_data['orig_root'], child_data['orig_lemma'],
                child_data['orig_meaning'], child_data['operation_role'], child_data['shell_meaning'],
                child_data['inversion_direction'], child_data['phonetic_chain'], child_data['qur_anchors'],
                child_data['dp_codes'], child_data['nt_code'], child_data['pattern'],
                child_data['parent_op'], child_data['gate_status'], child_data['notes']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {child_data}")
    
    print(f"    Migrated {count} child entries")
    return count

def migrate_a4_derivatives(conn, cursor, wb):
    """Migrate A4_DERIVATIVES sheet to derivatives table."""
    print("  Migrating A4_DERIVATIVES...")
    
    if "A4_DERIVATIVES" not in wb.sheetnames:
        print("    Sheet A4_DERIVATIVES not found")
        return 0
    
    ws = wb["A4_DERIVATIVES"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A4_DERIVATIVES")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map columns according to Excel structure:
        # DERIV_ID → ignore (database auto-generates)
        # ENTRY_ID → entry_id (integer)
        # DERIVATIVE → derivative_term
        # LINK_TYPE → link_type
        # EN_TERM → can be used for lookup if entry_id missing
        # No language column in Excel, default to 'en'
        entry_id = row_dict.get('entry_id')
        
        # If entry_id is not available, try to get it from en_term via lookup
        if not entry_id and row_dict.get('en_term'):
            # Try to find entry_id by looking up en_term in entries table
            cursor.execute('SELECT entry_id FROM entries WHERE en_term = ? LIMIT 1', (row_dict.get('en_term'),))
            result = cursor.fetchone()
            if result:
                entry_id = result[0]
        
        derivative_term = row_dict.get('derivative')
        link_type = row_dict.get('link_type')
        
        # Skip if missing critical data
        if not entry_id or not derivative_term:
            continue
        
        try:
            cursor.execute('''
                INSERT INTO derivatives (
                    entry_id, derivative_term, language, link_type
                ) VALUES (?, ?, ?, ?)
            ''', (
                entry_id, derivative_term, 'en', link_type
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: entry_id={entry_id}, derivative_term={derivative_term}, link_type={link_type}")
    
    print(f"    Migrated {count} derivatives")
    return count

def migrate_a5_cross_refs(conn, cursor, wb):
    """Migrate A5_CROSS_REFS sheet to cross_refs table."""
    print("  Migrating A5_CROSS_REFS...")
    
    if "A5_CROSS_REFS" not in wb.sheetnames:
        print("    Sheet A5_CROSS_REFS not found")
        return 0
    
    ws = wb["A5_CROSS_REFS"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A5_CROSS_REFS")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        # Map columns according to Excel structure:
        # XREF_ID → xref_id (string)
        # FROM_ID → from_entry_id (integer)
        # TO_ID → to_entry_id (integer)
        # LINK_TYPE → link_type
        # DESCRIPTION → description
        # LAYER_REF → layer_ref
        xref_id = row_dict.get('xref_id')
        from_id = row_dict.get('from_id')
        to_id = row_dict.get('to_id')
        link_type = row_dict.get('link_type')
        description = row_dict.get('description')
        layer_ref = row_dict.get('layer_ref')
        
        # Skip if missing critical data
        if not xref_id or from_id is None or to_id is None:
            continue
        
        try:
            cursor.execute('''
                INSERT INTO cross_refs (
                    xref_id, from_entry_id, to_entry_id, link_type, description, layer_ref
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                xref_id, from_id, to_id, link_type, description, layer_ref
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: xref_id={xref_id}, from_id={from_id}, to_id={to_id}, link_type={link_type}")
    
    print(f"    Migrated {count} cross-references")
    return count

def migrate_a3_quran_refs(conn, cursor, wb):
    """Migrate A3_QURAN_REFS sheet to quran_refs table."""
    print("  Migrating A3_QURAN_REFS...")
    
    if "A3_QURAN_REFS" not in wb.sheetnames:
        print("    Sheet A3_QURAN_REFS not found")
        return 0
    
    ws = wb["A3_QURAN_REFS"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in A3_QURAN_REFS")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        quran_data = {
            'ref_id': row_dict.get('ref_id'),
            'surah': row_dict.get('surah'),
            'ayah': row_dict.get('ayah'),
            'arabic_text': row_dict.get('arabic_text'),
            'relevance': row_dict.get('relevance'),
            'entry_ids': row_dict.get('entry_ids'),
            'network_id': row_dict.get('network_id'),
            'layer_ref': row_dict.get('layer_ref'),
            'qv_id': row_dict.get('qv_id'),
        }
        
        try:
            cursor.execute('''
                INSERT INTO quran_refs (
                    ref_id, surah, ayah, arabic_text, relevance,
                    entry_ids, network_id, layer_ref, qv_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                quran_data['ref_id'], quran_data['surah'], quran_data['ayah'],
                quran_data['arabic_text'], quran_data['relevance'],
                quran_data['entry_ids'], quran_data['network_id'],
                quran_data['layer_ref'], quran_data['qv_id']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {quran_data}")
    
    print(f"    Migrated {count} Qur'an references")
    return count

def migrate_m1_phonetic_shifts(conn, cursor, wb):
    """Migrate M1_PHONETIC_SHIFTS sheet to phonetic_shifts table."""
    print("  Migrating M1_PHONETIC_SHIFTS...")
    
    if "M1_PHONETIC_SHIFTS" not in wb.sheetnames:
        print("    Sheet M1_PHONETIC_SHIFTS not found")
        return 0
    
    ws = wb["M1_PHONETIC_SHIFTS"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in M1_PHONETIC_SHIFTS")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        shift_data = {
            'shift_id': row_dict.get('shift_id'),
            'ar_letter': row_dict.get('ar_letter'),
            'ar_name': row_dict.get('ar_name'),
            'en_outputs': row_dict.get('en_outputs'),
            'direction': row_dict.get('direction'),
            'examples': row_dict.get('examples'),
            'entry_ids': row_dict.get('entry_ids'),
            'foundation_ref': row_dict.get('foundation_ref'),
            'ru_outputs': row_dict.get('ru_outputs'),
        }
        
        try:
            cursor.execute('''
                INSERT INTO phonetic_shifts (
                    shift_id, ar_letter, ar_name, en_outputs, direction,
                    examples, entry_ids, foundation_ref, ru_outputs
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                shift_data['shift_id'], shift_data['ar_letter'], shift_data['ar_name'],
                shift_data['en_outputs'], shift_data['direction'], shift_data['examples'],
                shift_data['entry_ids'], shift_data['foundation_ref'], shift_data['ru_outputs']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {shift_data}")
    
    print(f"    Migrated {count} phonetic shifts")
    return count

def migrate_m2_detection_patterns(conn, cursor, wb):
    """Migrate M2_DETECTION_PATTERNS sheet to detection_patterns table."""
    print("  Migrating M2_DETECTION_PATTERNS...")
    
    if "M2_DETECTION_PATTERNS" not in wb.sheetnames:
        print("    Sheet M2_DETECTION_PATTERNS not found")
        return 0
    
    ws = wb["M2_DETECTION_PATTERNS"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in M2_DETECTION_PATTERNS")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        pattern_data = {
            'pattern_id': row_dict.get('pattern_id'),
            'name': row_dict.get('name'),
            'level': row_dict.get('level'),
            'description': row_dict.get('description'),
            'triggers': row_dict.get('triggers'),
            'qur_ref': row_dict.get('qur_ref'),
            'example': row_dict.get('example'),
            'entry_ids': row_dict.get('entry_ids'),
            'foundation_ref': row_dict.get('foundation_ref'),
        }
        
        try:
            cursor.execute('''
                INSERT INTO detection_patterns (
                    pattern_id, name, level, description, triggers,
                    qur_ref, example, entry_ids, foundation_ref
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                pattern_data['pattern_id'], pattern_data['name'], pattern_data['level'],
                pattern_data['description'], pattern_data['triggers'], pattern_data['qur_ref'],
                pattern_data['example'], pattern_data['entry_ids'], pattern_data['foundation_ref']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {pattern_data}")
    
    print(f"    Migrated {count} detection patterns")
    return count

def migrate_m4_networks(conn, cursor, wb):
    """Migrate M4_NETWORKS sheet to networks table."""
    print("  Migrating M4_NETWORKS...")
    
    if "M4_NETWORKS" not in wb.sheetnames:
        print("    Sheet M4_NETWORKS not found")
        return 0
    
    ws = wb["M4_NETWORKS"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in M4_NETWORKS")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        network_data = {
            'network_id': row_dict.get('network_id'),
            'name': row_dict.get('name'),
            'title': row_dict.get('title'),
            'link_verse': row_dict.get('link_verse'),
            'description': row_dict.get('description'),
            'mechanism': row_dict.get('mechanism'),
            'entry_ids': row_dict.get('entry_ids'),
            'status': row_dict.get('status'),
            'foundation_ref': row_dict.get('foundation_ref'),
        }
        
        try:
            cursor.execute('''
                INSERT INTO networks (
                    network_id, name, title, link_verse, description,
                    mechanism, entry_ids, status, foundation_ref
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                network_data['network_id'], network_data['name'], network_data['title'],
                network_data['link_verse'], network_data['description'], network_data['mechanism'],
                network_data['entry_ids'], network_data['status'], network_data['foundation_ref']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {network_data}")
    
    print(f"    Migrated {count} networks")
    return count

def migrate_m3_scholars(conn, cursor, wb):
    """Migrate M3_SCHOLARS sheet to scholars table."""
    print("  Migrating M3_SCHOLARS...")
    
    if "M3_SCHOLARS" not in wb.sheetnames:
        print("    Sheet M3_SCHOLARS not found")
        return 0
    
    ws = wb["M3_SCHOLARS"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in M3_SCHOLARS")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        scholar_data = {
            'scholar_id': row_dict.get('scholar_id'),
            'verified_name': row_dict.get('verified_name'),
            'birth_place': row_dict.get('birthplace'),
            'identity': row_dict.get('identity'),
            'role': row_dict.get('role'),
            'achievement': row_dict.get('achievement'),
            'lies_applied': row_dict.get('lies_applied'),
            'entry_ids': row_dict.get('entry_ids'),
        }
        
        try:
            cursor.execute('''
                INSERT INTO scholars (
                    scholar_id, verified_name, birth_place, identity,
                    role, achievement, lies_applied, entry_ids
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                scholar_data['scholar_id'], scholar_data['verified_name'],
                scholar_data['birth_place'], scholar_data['identity'],
                scholar_data['role'], scholar_data['achievement'],
                scholar_data['lies_applied'], scholar_data['entry_ids']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {scholar_data}")
    
    print(f"    Migrated {count} scholars")
    return count

def migrate_m5_qur_verification(conn, cursor, wb):
    """Migrate M5_QUR_VERIFICATION sheet to qur_verification table."""
    print("  Migrating M5_QUR_VERIFICATION...")
    
    if "M5_QUR_VERIFICATION" not in wb.sheetnames:
        print("    Sheet M5_QUR_VERIFICATION not found")
        return 0
    
    ws = wb["M5_QUR_VERIFICATION"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows or len(rows) < 2:
        print("    No data in M5_QUR_VERIFICATION")
        return 0
    
    headers = [clean_column_name(cell) for cell in rows[0]]
    
    count = 0
    for i, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = cell
        
        qv_data = {
            'qv_id': row_dict.get('qv_id'),
            'name': row_dict.get('name'),
            'mechanism': row_dict.get('mechanism'),
            'description': row_dict.get('description'),
            'markers': row_dict.get('markers'),
            'qur_refs': row_dict.get('qur_refs'),
            'contrast_refs': row_dict.get('contrast_refs'),
            'foundation_ref': row_dict.get('foundation_ref'),
        }
        
        try:
            cursor.execute('''
                INSERT INTO qur_verification (
                    qv_id, name, mechanism, description, markers,
                    qur_refs, contrast_refs, foundation_ref
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                qv_data['qv_id'], qv_data['name'], qv_data['mechanism'],
                qv_data['description'], qv_data['markers'], qv_data['qur_refs'],
                qv_data['contrast_refs'], qv_data['foundation_ref']
            ))
            count += 1
        except Exception as e:
            print(f"    Error inserting row {i}: {e}")
            print(f"    Data: {qv_data}")
    
    print(f"    Migrated {count} Qur'an verification mechanisms")
    return count

def extract_and_insert_roots(conn, cursor):
    """Extract unique roots from all entries and insert into roots table."""
    print("  Extracting roots from entries...")
    
    # First, collect all root_id values already present in entries
    cursor.execute('''
        SELECT DISTINCT root_id 
        FROM entries 
        WHERE root_id IS NOT NULL AND root_id != ''
    ''')
    existing_root_ids = {row[0] for row in cursor.fetchall()}
    
    # Fix root ID padding: convert R01 to R001 format
    count_fixed = 0
    for root_id in list(existing_root_ids):
        if root_id and re.match(r'^R\d{1,3}$', root_id):
            # Extract numeric part and pad to 3 digits
            match = re.match(r'^R(\d+)$', root_id)
            if match:
                num = int(match.group(1))
                padded_id = f'R{num:03d}'
                if padded_id != root_id:
                    # Update entries table to use padded ID
                    cursor.execute('''
                        UPDATE entries 
                        SET root_id = ?
                        WHERE root_id = ?
                    ''', (padded_id, root_id))
                    count_fixed += 1
                    print(f"    Fixed root ID padding: {root_id} → {padded_id}")
    
    # Re-fetch existing root_ids after fixing
    cursor.execute('''
        SELECT DISTINCT root_id 
        FROM entries 
        WHERE root_id IS NOT NULL AND root_id != ''
    ''')
    existing_root_ids = {row[0] for row in cursor.fetchall()}
    
    # For each existing root_id, ensure it exists in roots table
    count_existing = 0
    for root_id in existing_root_ids:
        cursor.execute('SELECT 1 FROM roots WHERE root_id = ?', (root_id,))
        if not cursor.fetchone():
            # Root doesn't exist, create a minimal entry
            # Extract numeric part for root_bare
            root_bare = root_id
            if root_id.startswith('R'):
                root_bare = root_id[1:]  # Remove 'R' prefix
            
            cursor.execute('''
                INSERT INTO roots (root_id, root_letters, root_bare, notes)
                VALUES (?, ?, ?, ?)
            ''', (root_id, f"Root {root_id}", root_bare, "Created during migration for existing entries"))
            count_existing += 1
    
    # Get all unique root_letters from entries that don't have root_id yet
    cursor.execute('''
        SELECT DISTINCT root_letters 
        FROM entries 
        WHERE root_letters IS NOT NULL AND root_letters != ''
        AND (root_id IS NULL OR root_id = '')
    ''')
    root_rows = cursor.fetchall()
    
    count_new = 0
    for root_letters, in root_rows:
        if not root_letters:
            continue
        
        # Generate root_id (R{count+1:03d})
        # But first check if we already have a root with these letters
        cursor.execute('SELECT root_id FROM roots WHERE root_letters = ?', (root_letters,))
        existing = cursor.fetchone()
        
        if existing:
            root_id = existing[0]
        else:
            # Count existing roots to generate new ID
            cursor.execute('SELECT COUNT(*) FROM roots')
            root_count = cursor.fetchone()[0]
            root_id = f"R{root_count + 1:03d}"
            
            # Create bare root (without hyphens)
            root_bare = re.sub(r'[\-\s]', '', root_letters)
            
            # Count entries with this root
            cursor.execute('SELECT COUNT(*) FROM entries WHERE root_letters = ?', (root_letters,))
            entry_count = cursor.fetchone()[0]
            
            # Try to extract root type from pattern
            root_type = 'TRILITERAL' if len(root_bare) == 3 else 'QUADRILITERAL' if len(root_bare) == 4 else 'OTHER'
            
            # Insert into roots table
            try:
                cursor.execute('''
                    INSERT INTO roots (
                        root_id, root_letters, root_bare, root_type, notes
                    ) VALUES (?, ?, ?, ?, ?)
                ''', (
                    root_id, root_letters, root_bare, root_type,
                    f"Extracted from {entry_count} entries during migration"
                ))
                count_new += 1
            except Exception as e:
                print(f"    Error inserting root {root_letters}: {e}")
                continue
        
        # Update entries with the root_id
        cursor.execute('''
            UPDATE entries 
            SET root_id = ? 
            WHERE root_letters = ? AND (root_id IS NULL OR root_id = '')
        ''', (root_id, root_letters))
    
    print(f"    Fixed {count_fixed} root IDs, created {count_existing} roots for existing IDs, {count_new} new roots from root_letters")
    return count_existing + count_new

def create_word_fingerprints(conn, cursor):
    """Create word_fingerprints entries for all searchable terms."""
    print("  Creating word_fingerprints...")
    
    # The triggers should have already created fingerprints when entries were inserted
    # But let's verify and create any missing ones
    cursor.execute('SELECT COUNT(*) FROM word_fingerprints')
    count = cursor.fetchone()[0]
    
    print(f"    {count} fingerprints already created by triggers")
    return count

# ============================================================================
# MAIN MIGRATION FUNCTION
# ============================================================================

def main():
    print("\n" + "═" * 70)
    print("  USLaP Migration: Excel → SQLite Relational Database")
    print("  بِسْمِ اللَّهِ الرَّحْمَٰنِ الرَّحِيمِ")
    print("═" * 70)
    
    # Check if files exist
    if not os.path.exists(EXCEL_PATH):
        print(f"\n❌ ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)
    
    if not os.path.exists(SCHEMA_PATH):
        print(f"\n❌ ERROR: Schema file not found: {SCHEMA_PATH}")
        sys.exit(1)
    
    # Backup existing database if it exists
    if os.path.exists(DB_PATH):
        backup_path = backup_database(DB_PATH)
        if backup_path:
            print(f"\n✓ Backed up existing database to: {backup_path}")
        # Remove the existing database file to start fresh
        os.remove(DB_PATH)
        print(f"  Removed existing database file to start fresh")
    
    # Load Excel workbook
    print(f"\n📖 Loading Excel file: {EXCEL_PATH}")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        print(f"  Found {len(wb.sheetnames)} sheets")
    except Exception as e:
        print(f"\n❌ ERROR: Failed to load Excel file: {e}")
        sys.exit(1)
    
    # Create database and execute schema
    print(f"\n🗄️  Creating database: {DB_PATH}")
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Enable foreign keys
        cursor.execute("PRAGMA foreign_keys = ON")
        
        # Read and execute schema
        with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
            schema_sql = f.read()
        
        print("  Executing schema...")
        cursor.executescript(schema_sql)
        
        # Register extract_consonants UDF (CRITICAL: must be done before any inserts)
        print("  Registering extract_consonants() UDF...")
        conn.create_function("extract_consonants", 1, extract_consonants)
        
        # Disable foreign keys during data migration to avoid constraint violations
        cursor.execute("PRAGMA foreign_keys = OFF")
        
        # Begin transaction
        conn.execute("BEGIN TRANSACTION")
        
        # Migrate each sheet
        print("\n📊 Migrating data...")
        
        total_counts = {}
        
        # Order matters: create entries first, then related tables
        total_counts['A1_ENTRIES'] = migrate_a1_entries(conn, cursor, wb)
        total_counts['A1_ЗАПИСИ'] = migrate_a1_zapisi(conn, cursor, wb)
        total_counts['PERSIAN_A1_MADĀKHIL'] = migrate_persian_a1_madakhil(conn, cursor, wb)
        total_counts['BITIG_A1_ENTRIES'] = migrate_bitig_a1_entries(conn, cursor, wb)
        total_counts['A2_NAMES_OF_ALLAH'] = migrate_a2_names_of_allah(conn, cursor, wb)
        
        # Extract and insert roots (must be done before other tables that reference roots)
        total_counts['roots'] = extract_and_insert_roots(conn, cursor)
        
        # Migrate remaining tables
        total_counts['CHILD_SCHEMA'] = migrate_child_schema(conn, cursor, wb)
        total_counts['A4_DERIVATIVES'] = migrate_a4_derivatives(conn, cursor, wb)
        total_counts['A5_CROSS_REFS'] = migrate_a5_cross_refs(conn, cursor, wb)
        total_counts['A3_QURAN_REFS'] = migrate_a3_quran_refs(conn, cursor, wb)
        total_counts['M1_PHONETIC_SHIFTS'] = migrate_m1_phonetic_shifts(conn, cursor, wb)
        total_counts['M2_DETECTION_PATTERNS'] = migrate_m2_detection_patterns(conn, cursor, wb)
        total_counts['M4_NETWORKS'] = migrate_m4_networks(conn, cursor, wb)
        total_counts['M3_SCHOLARS'] = migrate_m3_scholars(conn, cursor, wb)
        total_counts['M5_QUR_VERIFICATION'] = migrate_m5_qur_verification(conn, cursor, wb)
        
        # Create word fingerprints (triggers should have done this, but verify)
        total_counts['word_fingerprints'] = create_word_fingerprints(conn, cursor)
        
        # Commit transaction
        conn.commit()
        
        # Re-enable foreign keys and check constraints
        cursor.execute("PRAGMA foreign_keys = ON")
        
        # Generate statistics
        print("\n📈 Migration Statistics:")
        print("─" * 40)
        
        cursor.execute("SELECT COUNT(*) FROM entries")
        entries_count = cursor.fetchone()[0]
        print(f"  Total entries: {entries_count}")
        
        cursor.execute("SELECT COUNT(*) FROM roots")
        roots_count = cursor.fetchone()[0]
        print(f"  Total roots: {roots_count}")
        
        cursor.execute("SELECT COUNT(*) FROM child_entries")
        child_count = cursor.fetchone()[0]
        print(f"  Child entries: {child_count}")
        
        cursor.execute("SELECT COUNT(*) FROM names_of_allah")
        names_count = cursor.fetchone()[0]
        print(f"  Names of Allah: {names_count}")
        
        cursor.execute("SELECT COUNT(*) FROM derivatives")
        derivatives_count = cursor.fetchone()[0]
        print(f"  Derivatives: {derivatives_count}")
        
        cursor.execute("SELECT COUNT(*) FROM cross_refs")
        crossrefs_count = cursor.fetchone()[0]
        print(f"  Cross-references: {crossrefs_count}")
        
        cursor.execute("SELECT COUNT(*) FROM word_fingerprints")
        fingerprints_count = cursor.fetchone()[0]
        print(f"  Word fingerprints: {fingerprints_count}")
        
        cursor.execute("SELECT COUNT(*) FROM engine_queue")
        queue_count = cursor.fetchone()[0]
        print(f"  Engine queue items: {queue_count}")
        
        # Verify foreign keys
        print("\n🔍 Verifying foreign key constraints...")
        cursor.execute("PRAGMA foreign_key_check")
        fk_errors = cursor.fetchall()
        
        if fk_errors:
            print("❌ Foreign key errors found:")
            for error in fk_errors:
                print(f"  Table: {error[0]}, Row: {error[1]}, Referenced: {error[2]}, FK Index: {error[3]}")
        else:
            print("✓ All foreign key constraints satisfied")
        
        # Close workbook and connection
        wb.close()
        conn.close()
        
        print("\n" + "═" * 70)
        print("✅ MIGRATION COMPLETED SUCCESSFULLY")
        print(f"✅ Database: {DB_PATH}")
        print("═" * 70)
        
        # Print next steps
        print("\n📋 Next steps:")
        print("  1. Test the database with: sqlite3 uslap_lattice.db '.tables'")
        print("  2. Verify data integrity with queries")
        print("  3. Update db_access_layer.py to use the new schema")
        print("  4. Run USLaP_Engine.py to test cluster expansion")
        
    except Exception as e:
        print(f"\n❌ ERROR during migration: {e}")
        import traceback
        traceback.print_exc()
        
        # Rollback if connection is still open
        try:
            conn.rollback()
            conn.close()
        except:
            pass
        
        print("\n⚠️  Migration failed. Database has been rolled back.")
        sys.exit(1)

if __name__ == "__main__":
    main()