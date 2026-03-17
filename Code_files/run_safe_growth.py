#!/usr/bin/env python3
"""
Run safe accelerated growth without interactive prompts
"""

import sys
import os
from pathlib import Path

# Add current directory to path to import our module
sys.path.insert(0, str(Path(__file__).parent))

def run_non_interactive(cycles=2, entries_per_cycle=3):
    """Run growth without interactive prompts"""
    filepath = "USLaP_Final_Data_Consolidated_Master.xlsx"
    
    if not os.path.exists(filepath):
        print(f"❌ File not found: {filepath}")
        return False
    
    print(f"Running safe growth: {cycles} cycles, {entries_per_cycle} entries per cycle")
    
    # Import the safe functions directly
    import openpyxl
    from datetime import datetime
    import time
    import shutil
    
    # Simplified version of SafeFileHandler
    class SimpleFileHandler:
        def __init__(self, filepath):
            self.filepath = Path(filepath)
            self.backup_path = None
            
        def create_backup(self):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"{self.filepath.stem}_backup_{timestamp}{self.filepath.suffix}"
            self.backup_path = self.filepath.parent / backup_name
            
            try:
                shutil.copy2(self.filepath, self.backup_path)
                print(f"📁 Created backup: {self.backup_path.name}")
                return True
            except Exception as e:
                print(f"❌ Failed to create backup: {e}")
                return False
        
        def write_entries_safely(self, entries, sheet_name='A1_ENTRIES'):
            if not entries:
                return 0
            
            temp_file = self.filepath.parent / f"temp_{self.filepath.name}"
            
            try:
                wb = openpyxl.load_workbook(self.filepath)
                
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
                    headers = ['ENTRY_ID', 'SCORE', 'EN_TERM', 'AR_WORD', 'ROOT_ID', 
                              'ROOT_LETTERS', 'QUR_MEANING', 'PATTERN', 'NETWORK_ID',
                              'PHONETIC_CHAIN', 'INVERSION_TYPE', 'SOURCE_FORM', 'FOUNDATION_REF']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                
                next_row = ws.max_row + 1 if ws.max_row > 1 else 2
                written_count = 0
                
                for entry in entries:
                    ws.cell(row=next_row, column=1, value=entry.get('ENTRY_ID', ''))
                    ws.cell(row=next_row, column=2, value=entry.get('SCORE', 0))
                    ws.cell(row=next_row, column=3, value=entry.get('EN_TERM', ''))
                    ws.cell(row=next_row, column=4, value=entry.get('AR_WORD', ''))
                    ws.cell(row=next_row, column=5, value=entry.get('ROOT_ID', ''))
                    ws.cell(row=next_row, column=6, value=entry.get('ROOT_LETTERS', '—'))
                    ws.cell(row=next_row, column=7, value=entry.get('QUR_MEANING', '—'))
                    ws.cell(row=next_row, column=8, value=entry.get('PATTERN', 'A'))
                    ws.cell(row=next_row, column=9, value=entry.get('NETWORK_ID', ''))
                    ws.cell(row=next_row, column=10, value=entry.get('PHONETIC_CHAIN', '—'))
                    ws.cell(row=next_row, column=11, value=entry.get('INVERSION_TYPE', 'HIDDEN'))
                    ws.cell(row=next_row, column=12, value=entry.get('SOURCE_FORM', '—'))
                    ws.cell(row=next_row, column=13, value=entry.get('FOUNDATION_REF', '—'))
                    
                    next_row += 1
                    written_count += 1
                
                # Safe save
                wb.save(temp_file)
                
                # Verify temp file
                test_wb = openpyxl.load_workbook(temp_file, read_only=True)
                test_wb.close()
                
                # Replace original
                shutil.move(temp_file, self.filepath)
                return written_count
                
            except Exception as e:
                print(f"❌ Error writing entries: {e}")
                return 0
            finally:
                if temp_file.exists():
                    try:
                        temp_file.unlink()
                    except:
                        pass
    
    # Simple generator
    class SimpleGenerator:
        def __init__(self):
            self.next_id = 1000
            
        def generate_ratio_entries(self, count=10):
            entries = []
            ratios = [
                (4, 3, 'musical fourth', 'N08'),
                (5, 3, 'musical sixth', 'N08'),
                (7, 3, 'maqam proportion', 'N08'),
                (7, 5, 'prayer kernel', 'N06'),
                (11, 5, 'dombra sacred', 'N16'),
                (22, 7, 'circle perfect', 'N07'),
                (28, 7, 'lunar perfect', 'N05'),
                (99, 70, 'high precision √2', 'N05'),
                (355, 113, 'high precision π', 'N05'),
                (19, 7, 'growth constant', 'N10'),
            ]
            
            for i, (num, den, name, network) in enumerate(ratios[:count]):
                entry_id = f"F{self.next_id:04d}"
                self.next_id += 1
                
                entries.append({
                    'ENTRY_ID': entry_id,
                    'SCORE': 9,
                    'EN_TERM': f"{name} ({num}/{den})",
                    'AR_WORD': f"نسبة {num}/{den}",
                    'ROOT_ID': f"R{self.next_id}",
                    'ROOT_LETTERS': '—',
                    'QUR_MEANING': f"Divine ratio {num}/{den} — precise measure (Q54:49)",
                    'PATTERN': 'A',
                    'NETWORK_ID': network,
                    'PHONETIC_CHAIN': '→'.join(['✓'] * 3),
                    'INVERSION_TYPE': 'DIVINE',
                    'SOURCE_FORM': f"ratio:{num}/{den}",
                    'FOUNDATION_REF': f"Q54:49 + {den}-denominator"
                })
            
            return entries
    
    # Main execution
    print("\n" + "="*60)
    print(f"🛡️  SAFE GROWTH (NON-INTERACTIVE)")
    print("="*60)
    
    file_handler = SimpleFileHandler(filepath)
    generator = SimpleGenerator()
    
    # Create backup
    print("\n📁 Creating backup...")
    if not file_handler.create_backup():
        print("⚠️  Backup creation failed, but continuing...")
    
    # Run cycles
    print(f"\n🚀 Running {cycles} cycles...")
    start_time = time.time()
    total_written = 0
    
    for cycle in range(1, cycles + 1):
        print(f"\n--- Cycle {cycle}/{cycles} ---")
        
        try:
            entries = generator.generate_ratio_entries(entries_per_cycle)
            written = file_handler.write_entries_safely(entries, 'A1_ENTRIES')
            total_written += written
            
            elapsed = time.time() - start_time
            print(f"  📈 +{written} entries | Total: {total_written} | Elapsed: {elapsed:.1f}s")
            
            if cycle < cycles:
                time.sleep(0.5)
                
        except Exception as e:
            print(f"  ❌ Error in cycle {cycle}: {e}")
            time.sleep(1.0)
            continue
    
    # Final summary
    total_time = time.time() - start_time
    print("\n" + "="*60)
    print(f"✅ GROWTH COMPLETE")
    print(f"   Cycles: {cycles}")
    print(f"   Total entries added: {total_written}")
    print(f"   Total time: {total_time:.1f} seconds")
    
    # Verify final file
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if 'A1_ENTRIES' in wb.sheetnames:
            ws = wb['A1_ENTRIES']
            print(f"   A1_ENTRIES now has {ws.max_row} rows")
        wb.close()
        
        final_size = Path(filepath).stat().st_size
        print(f"   Final file size: {final_size:,} bytes")
        print("="*60)
        
        return True
        
    except Exception as e:
        print(f"❌ Error verifying final file: {e}")
        print("="*60)
        return False

if __name__ == "__main__":
    # Parse command line arguments
    cycles = 2
    entries_per_cycle = 3
    
    if len(sys.argv) > 1:
        try:
            cycles = int(sys.argv[1])
        except:
            pass
    
    if len(sys.argv) > 2:
        try:
            entries_per_cycle = int(sys.argv[2])
        except:
            pass
    
    success = run_non_interactive(cycles, entries_per_cycle)
    sys.exit(0 if success else 1)